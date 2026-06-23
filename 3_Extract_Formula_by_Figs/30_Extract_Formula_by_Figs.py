import hashlib
import json
import sys
from pathlib import Path

import pandas as pd
from openpyxl.styles import Font, PatternFill

current_dir = Path(__file__).resolve().parent
parent_dir = current_dir.parent
if str(parent_dir) not in sys.path:
    sys.path.append(str(parent_dir))

from find_api import find_api_key_file, get_vertexai_client
from LLM_Batch import (
    append_batch_request,
    build_batch_request_metadata,
    build_generate_content_batch_request,
    count_requests_in_jsonl,
    create_batch_job_record,
    create_batch_request_file,
    download_batch_results,
    load_batch_results_as_map,
    poll_batch_job,
    submit_batch_job,
    upload_file_to_gcs,
)


DEFAULT_GCS_BATCH_BUCKET = "gs://lnpdb-articles-extraction-batch-results-geon"
TARGET_VISUAL_TYPES = ["barplot", "table", "chemical_structure", "heatmap"]
REQUIRED_RESULT_KEYS = [
    "Item_ID",
    "formulation_id",
    "IL_name",
    "IL_molarratio",
    "HL_name",
    "HL_molarratio",
    "CHL_name",
    "CHL_molarratio",
    "PEG_name",
    "PEG_molarratio",
    "Fifth_component_name",
    "Fifth_component_molarratio",
    "IL_to_nucleicacid_massratio",
    "_reference",
]


def load_total_mapping(paper_folder: Path, mapping_json_path=None):
    mapping_file = Path(mapping_json_path) if mapping_json_path is not None else paper_folder / "total_figure_mapping.json"
    if not mapping_file.exists():
        return {}

    with open(mapping_file, "r", encoding="utf-8") as f:
        data = json.load(f)

    # {paper_folder_name: {...}} 구조면 안쪽만 꺼냄
    if isinstance(data, dict) and paper_folder.name in data and isinstance(data[paper_folder.name], dict):
        return data[paper_folder.name]

    return data if isinstance(data, dict) else {}

def find_item_mapping(node, item_id: str):
    if isinstance(node, dict):
        if item_id in node and isinstance(node[item_id], dict):
            return node[item_id]
        for value in node.values():
            found = find_item_mapping(value, item_id)
            if found is not None:
                return found

    elif isinstance(node, list):
        for value in node:
            found = find_item_mapping(value, item_id)
            if found is not None:
                return found

    return None

def get_fig_selection_value(row):
    manual = str(row.get("manual_select", "")).strip().lower()
    if is_meaningful_manual_value(manual):
        if manual in {"yes", "y", "1", "true"}:
            return "yes"
        if manual in {"no", "n", "0", "false"}:
            return "no"
        if manual == "maybe":
            return "maybe"
        return "no"

    return str(row.get("need_for_lnpdb", "")).strip().lower()

EMPTY_MANUAL_VALUES = {"", "nan", "none", "null", "[]", "{}"}

def is_meaningful_manual_value(value) -> bool:
    if value is None:
        return False
    try:
        if pd.isna(value):
            return False
    except (TypeError, ValueError):
        pass
    return str(value).strip().lower() not in EMPTY_MANUAL_VALUES

def list_pdf_paths(folder: Path) -> list[Path]:
    return sorted(
        [
            f for f in folder.rglob("*.pdf")
            if f.is_file() and not f.name.startswith("~")
        ]
    )


def upload_pdfs_to_gcs(folder: Path, gcs_batch_bucket: str) -> list[dict]:
    uploaded = []
    for pdf_path in list_pdf_paths(folder):
        gcs_uri = f"{gcs_batch_bucket}/papers/{folder.name}/{pdf_path.name}"
        upload_file_to_gcs(pdf_path, gcs_uri)
        uploaded.append(
            {
                "local_path": str(pdf_path),
                "gcs_uri": gcs_uri,
                "mime_type": "application/pdf",
                "name": pdf_path.name,
            }
        )
    return uploaded


def build_pdf_file_parts(uploaded_pdfs: list[dict]) -> list[dict]:
    parts = []
    for item in uploaded_pdfs:
        gcs_uri = str(item.get("gcs_uri", "")).strip()
        mime_type = str(item.get("mime_type", "application/pdf")).strip() or "application/pdf"
        if not gcs_uri:
            continue
        parts.append({"fileData": {"fileUri": gcs_uri, "mimeType": mime_type}})
    return parts


def get_md_text(folder: Path) -> str:
    texts = []
    for path in folder.rglob("*.md"):
        if path.is_file() and not path.name.startswith("~"):
            try:
                texts.append(path.read_text(encoding="utf-8"))
            except UnicodeDecodeError:
                texts.append(path.read_text(encoding="utf-8", errors="replace"))
    return "\n\n".join(texts)


def guess_mime_type(file_path: Path) -> str | None:
    ext = file_path.suffix.lower()
    if ext == ".pdf":
        return "application/pdf"
    if ext == ".png":
        return "image/png"
    if ext in {".jpg", ".jpeg"}:
        return "image/jpeg"
    return None


def upload_media_part(file_path: Path, paper_folder: Path, upload_cache: dict[str, dict]) -> dict | None:
    resolved = str(file_path.resolve())
    if resolved in upload_cache:
        return upload_cache[resolved]

    mime_type = guess_mime_type(file_path)
    if not mime_type:
        return None

    digest = hashlib.sha1(resolved.encode("utf-8")).hexdigest()[:12]
    gcs_uri = f"{DEFAULT_GCS_BATCH_BUCKET}/formula_media/{paper_folder.name}/{digest}_{file_path.name}"
    upload_file_to_gcs(file_path, gcs_uri)
    part = {"fileData": {"fileUri": gcs_uri, "mimeType": mime_type}}
    upload_cache[resolved] = part
    return part


def build_item_image_parts(item_id: str, total_mapping: dict, paper_folder: Path, upload_cache: dict[str, dict]) -> list[dict]:
    item_map = find_item_mapping(total_mapping, item_id)

    if not isinstance(item_map, dict):
        return []

    panels = item_map.get("panels", {})
    if not isinstance(panels, dict):
        panels = {}

    img_paths = [item_map.get("full_image")] + list(panels.values())

    parts = []
    seen = set()
    for raw_path in img_paths:
        if not raw_path:
            continue

        img_path = Path(raw_path)
        if not img_path.exists():
            continue

        resolved = str(img_path.resolve())
        if resolved in seen:
            continue
        seen.add(resolved)

        part = upload_media_part(img_path, paper_folder, upload_cache)
        if part:
            parts.append(part)

    return parts


def make_formula_item_custom_id(paper_folder: Path, item_id: str) -> str:
    seed = f"{paper_folder.name}::{item_id.strip().lower()}"
    return "formula_fig__" + hashlib.sha1(seed.encode("utf-8")).hexdigest()[:12]


def build_formula_item_prompt(item_id: str, request_id: str) -> str:
    return f"""
    당신은 지질나노입자(LNP) 제형 전문가입니다. 제공된 문서와 이미지를 기반으로 지정된 figure의 LNP 조성비를 추출하세요.

    [Request Tracking]
    - request_id: {request_id}

    [분석 대상]
    - TARGET ITEM: {item_id}

    [추출 대상 항목]
    - Item_ID
    - formulation_id (LNP 제형 이름. 명확하지 않으면 "N/A")
    - IL_name, IL_molarratio
    - HL_name, HL_molarratio
    - CHL_name, CHL_molarratio
    - PEG_name, PEG_molarratio
    - IL_to_nucleicacid_massratio
    - Fifth_component_name, Fifth_component_molarratio (기본 4성분 외 추가 성분이 있으면 기록, 없으면 "N/A")
    - _reference

    [투입 대비 비율 추출 규칙]
    1. 반드시 `IL_to_nucleicacid_massratio`를 최대한 찾으세요.
    2. lipid:nucleic acid mass ratio, ionizable lipid:nucleic acid mass ratio, lipid:RNA ratio, lipid:mRNA ratio, wt/wt ratio 등이 직접 제시되면 그 값을 그대로 쓰세요.
    3. 정확한 mass ratio가 없으면 비슷한 개념의 값도 허용합니다. 예: N/P ratio, charge ratio, (+/-) ratio.
    4. 그런 경우 반드시 괄호로 종류를 표시하세요.
       - 예: `6 (N/P ratio)`
       - 예: `3 (charge ratio)`
       - 예: `10:1 (lipid:mRNA mass ratio)`
    5. 정확한 mass ratio가 있으면 괄호 없이 값만 쓰세요.
    6. 관련 정보가 없으면 빈 문자열 `""`로 쓰세요.
    8. Fifth_component_name, Fifth_component_molarratio
        - 기본 4성분(IL, HL, CHL, PEG) 외에 추가 성분이 존재하면 기록하세요.
        - 5번째 성분이 없으면 반드시 "N/A"로 기입하세요.
        - 빈 문자열 "", null, 생략은 허용하지 않습니다.


    [주의사항]
    1. 여러 제형이 발견되면 각각 별도의 JSON 객체로 분리하세요.
    2. 모든 결과 row의 `Item_ID`는 반드시 `{item_id}` 그대로 써야 합니다.
    3. 최상위 JSON에 `request_id`를 반드시 그대로 복사하세요.
    4. 코드블록 없이 JSON만 반환하세요.

    {{
      "request_id": "{request_id}",
      "results": [
        {{
          "Item_ID": "{item_id}",
          "formulation_id": "C8-200",
          "IL_name": "C8",
          "IL_molarratio": "50",
          "HL_name": "DOPE",
          "HL_molarratio": "10",
          "CHL_name": "Cholesterol",
          "CHL_molarratio": "38.5",
          "PEG_name": "DMG-PEG2000",
          "PEG_molarratio": "1.5",
          "Fifth_component_name": "N/A",
          "Fifth_component_molarratio": "N/A",
          "IL_to_nucleicacid_massratio": "6 (N/P ratio)",
          "_reference": "LNP formulations were prepared at a molar ratio of 50:10:38.5:1.5..."
        }}
      ]
    }}
    """


def validate_formula_item_results(results, expected_item_id: str):
    if not isinstance(results, list) or not results:
        raise ValueError("missing_results")

    expected_normalized = expected_item_id.strip().lower()
    for idx, row in enumerate(results):
        if not isinstance(row, dict):
            raise ValueError(f"result_not_dict::{idx}")
        missing_keys = [key for key in REQUIRED_RESULT_KEYS if key not in row]
        if missing_keys:
            raise ValueError(f"missing_keys::{idx}::{','.join(missing_keys)}")
        item_id = str(row.get("Item_ID", "")).strip()
        if not item_id:
            raise ValueError(f"missing_item_id::{idx}")
        if item_id.lower() != expected_normalized:
            raise ValueError(f"unexpected_item_id::{idx}::{item_id}")


def run_formula_item_batch(
    folder: Path,
    client,
    model_name: str,
    item_ids: list[str],
    shared_contents: list,
    item_image_parts_map: dict[str, list[dict]],
    task_suffix: str = "",
):
    request_file = create_batch_request_file(folder, f"formula_fig_{folder.name}{task_suffix}")
    request_count = 0

    for item_id in item_ids:
        custom_id = make_formula_item_custom_id(folder, item_id)
        prompt = build_formula_item_prompt(item_id=item_id, request_id=custom_id)
        metadata = build_batch_request_metadata(
            task_name="extract_formula_for_item",
            model_name=model_name,
            custom_id=custom_id,
            stage_name="extract_formula_for_item",
            item_id=item_id,
            paper_folder=str(folder),
            extra_metadata={"item_id": item_id},
        )
        request_body = build_generate_content_batch_request(
            model_name=model_name,
            contents=list(shared_contents) + list(item_image_parts_map.get(item_id, [])),
            prompt_text=prompt,
            response_mime_type="application/json",
        )
        append_batch_request(request_file, custom_id, request_body, metadata)
        request_count += 1

    if request_count == 0 or count_requests_in_jsonl(request_file) == 0:
        return {}

    local_job_id = create_batch_job_record(
        paper_folder=folder,
        task_name="extract_formula_for_item",
        model_name=model_name,
        request_file=request_file,
        metadata={
            "request_count": count_requests_in_jsonl(request_file),
            "gcs_input_uri": f"{DEFAULT_GCS_BATCH_BUCKET}/batch/{request_file.name}",
            "gcs_output_uri_prefix": f"{DEFAULT_GCS_BATCH_BUCKET}/batch_output/{request_file.stem}",
        },
    )

    batch_job = submit_batch_job(
        client=client,
        paper_folder=folder,
        local_job_id=local_job_id,
        display_name=f"formula-item-{folder.name}{task_suffix}",
    )
    print(f"  · figure batch 제출 완료: {batch_job.name}")

    finished_job = poll_batch_job(client, folder, local_job_id, poll_interval_seconds=30)
    state_name = getattr(getattr(finished_job, "state", None), "name", None) or str(getattr(finished_job, "state", "UNKNOWN"))
    print(f"  · figure batch 종료 상태: {state_name}")
    if state_name != "JOB_STATE_SUCCEEDED":
        raise RuntimeError(f"extract_formula_for_item batch 실패: {state_name}")

    result_file = download_batch_results(client, folder, local_job_id)
    return load_batch_results_as_map(result_file)


def consume_formula_item_batch_results(batch_results_map, folder: Path, item_ids: list[str], model_name: str):
    all_results = []
    usage_rows = []
    failed_items = []

    for item_id in item_ids:
        custom_id = make_formula_item_custom_id(folder, item_id)
        row = batch_results_map.get(custom_id)

        if not row:
            print(f"  ! batch 결과 누락: {item_id}")
            failed_items.append(item_id)
            continue
        if not row.get("success"):
            print(f"  ! batch success=false: {item_id} | {row.get('error')}")
            failed_items.append(item_id)
            continue

        response_text = str(row.get("response_text", "")).strip()
        if not response_text:
            print(f"  ! batch 빈 응답: {item_id}")
            failed_items.append(item_id)
            continue

        clean_text = response_text.replace("```json", "").replace("```", "").strip()
        try:
            payload = json.loads(clean_text)
            if not isinstance(payload, dict):
                raise ValueError("response_not_dict")

            response_request_id = str(payload.get("request_id", "")).strip()
            if not response_request_id:
                raise ValueError("missing_request_id")
            if response_request_id != custom_id:
                raise ValueError(f"request_id_mismatch::{response_request_id}")

            results = payload.get("results")
            validate_formula_item_results(results, item_id)
            all_results.extend(results)

            cost_info = row.get("cost_info") or {}
            usage_rows.append(
                {
                    "item_id": item_id,
                    "task_name": "extract_formula_for_item",
                    "model_name": model_name,
                    "input_tokens": row.get("input_tokens"),
                    "output_tokens": row.get("output_tokens"),
                    "total_tokens": row.get("total_tokens"),
                    "billed_output_tokens": row.get("billed_output_tokens"),
                    "total_cost_usd": cost_info.get("total_cost_usd", 0.0),
                }
            )
        except Exception as e:
            print(f"  ! batch 파싱/검증 실패 ({item_id}): {e}")
            failed_items.append(item_id)

    return all_results, usage_rows, failed_items


def fig_main(paper_folder_str, model_name, api_json_name="vertex.json", classified_csv_path=None, output_xlsx_path=None, batch_usage_csv_path=None, mapping_json_path=None):
    paper_folder = Path(paper_folder_str)
    print(f"🔬 [조성비 추출] 폴더 분석 시작: {paper_folder.name}")

    csv_path = Path(classified_csv_path) if classified_csv_path is not None else paper_folder / "fig_table_lnpdb_classified.csv"
    if not csv_path.exists():
        print(f"  ! 오류: 분류 CSV가 없습니다 ({csv_path}).")
        return

    df = pd.read_csv(csv_path)
    df["_fig_select"] = df.apply(get_fig_selection_value, axis=1)

    if "manual_select" in df.columns:
        manual_mask = df["manual_select"].apply(is_meaningful_manual_value)
    else:
        manual_mask = pd.Series(False, index=df.index)
    selected_mask = df["_fig_select"].isin(["yes", "maybe"])
    visual_mask = df["visual_type"].astype(str).str.lower().isin(TARGET_VISUAL_TYPES)
    target_mask = (manual_mask & selected_mask) | (~manual_mask & visual_mask & selected_mask)
    manual_selected_count = int((manual_mask & selected_mask).sum())
    manual_excluded_count = int((manual_mask & ~selected_mask).sum())
    auto_selected_count = int((~manual_mask & visual_mask & selected_mask).sum())
    print(f"  - rows selected by manual_select override: {manual_selected_count}")
    print(f"  - rows selected by automatic criteria: {auto_selected_count}")
    print(f"  - rows excluded by manual_select=no/invalid: {manual_excluded_count}")

    target_items = df[target_mask]["item_id"].astype(str).str.strip().tolist()

    if not target_items:
        print("  ! 분석 대상 아이템이 없습니다.")
        return

    api_key_path = find_api_key_file(api_json_name)
    with open(api_key_path, "r", encoding="utf-8") as f:
        cred_data = json.load(f)

    project_id = cred_data.get("project_id")
    if not project_id:
        raise ValueError(f"서비스 계정 JSON에 project_id가 없습니다: {api_key_path}")

    print(f"🔧 Vertex 프로젝트 설정: {project_id}")
    client = get_vertexai_client(api_key_path, project=project_id)
    total_mapping = load_total_mapping(paper_folder, mapping_json_path=mapping_json_path)

    print("  - PDF 및 figure 이미지 업로드 준비 중..")
    uploaded_pdfs = upload_pdfs_to_gcs(paper_folder, DEFAULT_GCS_BATCH_BUCKET)
    shared_contents = build_pdf_file_parts(uploaded_pdfs)
    md_text = get_md_text(paper_folder)
    if md_text.strip():
        shared_contents.append(md_text)

    upload_cache = {}
    item_image_parts_map = {}
    for item_id in target_items:
        item_image_parts_map[item_id] = build_item_image_parts(
            item_id=item_id,
            total_mapping=total_mapping,
            paper_folder=paper_folder,
            upload_cache=upload_cache,
        )

    batch_results_map = run_formula_item_batch(
        folder=paper_folder,
        client=client,
        model_name=model_name,
        item_ids=target_items,
        shared_contents=shared_contents,
        item_image_parts_map=item_image_parts_map,
    )

    all_results, usage_rows, failed_items = consume_formula_item_batch_results(
        batch_results_map=batch_results_map,
        folder=paper_folder,
        item_ids=target_items,
        model_name=model_name,
    )

    if failed_items:
        print(f"\n🔁 실패한 {len(failed_items)}개 item만 batch 재시도합니다..")
        retry_item_image_parts_map = {item_id: item_image_parts_map.get(item_id, []) for item_id in failed_items}
        retry_results_map = run_formula_item_batch(
            folder=paper_folder,
            client=client,
            model_name=model_name,
            item_ids=failed_items,
            shared_contents=shared_contents,
            item_image_parts_map=retry_item_image_parts_map,
            task_suffix="_retry",
        )
        retry_data, retry_usage_rows, retry_failed_items = consume_formula_item_batch_results(
            batch_results_map=retry_results_map,
            folder=paper_folder,
            item_ids=failed_items,
            model_name=model_name,
        )
        all_results.extend(retry_data)
        usage_rows.extend(retry_usage_rows)
        if retry_failed_items:
            print(f"  ! 재시도 후에도 실패한 item {len(retry_failed_items)}개")

    if all_results:
        final_df = pd.DataFrame(all_results)
        cols = final_df.columns.tolist()
        if "Item_ID" in cols:
            cols.insert(0, cols.pop(cols.index("Item_ID")))
        if "formulation_id" in cols:
            cols.insert(1, cols.pop(cols.index("formulation_id")))
        final_df = final_df[cols]

        output_xlsx = Path(output_xlsx_path) if output_xlsx_path is not None else paper_folder / f"3_{paper_folder.name}_formula.xlsx"

        with pd.ExcelWriter(output_xlsx, engine="openpyxl") as writer:
            final_df.to_excel(writer, index=False, sheet_name="Formula_Data")
            ws = writer.sheets["Formula_Data"]

            highlight = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")
            bold_font = Font(bold=True)

            target_cols = [
                c
                for c in final_df.columns
                if any(x in c.lower() for x in ["name", "molarratio", "formulation_id", "il_to_nucleicacid_massratio"])
            ]
            for col_name in target_cols:
                col_idx = final_df.columns.get_loc(col_name) + 1
                for row in range(1, len(final_df) + 2):
                    cell = ws.cell(row=row, column=col_idx)
                    cell.fill = highlight
                    if row > 1:
                        cell.font = bold_font

        print(f"\n✅ 추출 완료 및 저장 성공: {output_xlsx}")

    if usage_rows:
        usage_csv_path = Path(batch_usage_csv_path) if batch_usage_csv_path is not None else paper_folder / "30_formula_fig_batch_usage.csv"
        pd.DataFrame(usage_rows).to_csv(usage_csv_path, index=False, encoding="utf-8-sig")


if __name__ == "__main__":
    TEST_DIR = r"C:\Users\kogun\PycharmProjects\LNPDB_Articles_Extraction\Extraction_Examples\excel_x"
    model_name = "gemini-3.1-pro-preview"
    api_json_name = "vertex.json"
    fig_main(TEST_DIR, model_name, api_json_name)
