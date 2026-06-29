from __future__ import annotations

import argparse
import csv
import importlib.util
import json
import os
import re
import shlex
import shutil
import subprocess
import sys
import threading
import time
import traceback
from collections import Counter, defaultdict
from datetime import datetime, timezone
from pathlib import Path
from typing import Any


PROJECT_ROOT = Path(__file__).resolve().parent
WORKSPACE = PROJECT_ROOT / "agent_workspace"
STATE_PATH = WORKSPACE / "agent_state.json"
QUEUE_PATH = WORKSPACE / "task_queue.json"
LOG_DIR = WORKSPACE / "logs"
TASK_DIR = WORKSPACE / "tasks"
MANUAL_MARKER = ".manual_select_review_done"
PROJECT_EXCLUDE_DIRS = {
    ".git",
    ".idea",
    "__pycache__",
    "0_mark_down_gen",
    "agent_workspace",
}

FALLBACK_SOURCE_QUALITIES = {
    "suspect_crop",
    "missing_image",
    "caption_image_mismatch",
}

VALID_SOURCE_QUALITIES = FALLBACK_SOURCE_QUALITIES | {
    "ok",
    "pdf_page_render_fallback",
    "manual_required",
}


STAGE_ORDER = [
    "00_marker",
    "01_make_ft_csv",
    "02_ft_selector",
    "02b_manual_review",
    "03_figure_mapping",
    "03_split_excel_blocks",
    "03_split_excel_blocks_batch",
    "04_figure_separate",
    "04_ft_excel_matcher",
    "05_smiles_structure_resolution",
    "06_unified_lnpdb_extraction",
    "07_finalize_unified_table",
]

AGENT_STAGES = {
    "03_figure_mapping",
    "03_split_excel_blocks",
    "03_split_excel_blocks_batch",
    "04_figure_separate",
    "04_ft_excel_matcher",
    "05_smiles_structure_resolution",
    "06_unified_lnpdb_extraction",
    "07_finalize_unified_table",
}


STAGE_EXECUTION_MODE = {
    "03_figure_mapping": "external_agent",
    "03_split_excel_blocks_batch": "external_agent",
    "04_figure_separate": "external_agent",
    "04_ft_excel_matcher": "external_agent",
    "05_smiles_structure_resolution": "external_agent",
    "06_unified_lnpdb_extraction": "external_agent",
    "07_finalize_unified_table": "heuristic",
}

VALID_STAGE_EXECUTION_MODES = {"legacy", "external_agent", "heuristic"}

LEGACY_CONTEXT_BY_STAGE = {
    "00_marker": [
        "agent_workspace/legacy_context/by_stage/00_marker/README.md",
        "agent_workspace/legacy_context/full_copy/0_mark_down_gen/00_Marker.py",
    ],
    "02b_manual_review": [
        "agent_workspace/legacy_context/by_stage/02b_manual_review/README.md",
        "agent_workspace/legacy_context/full_copy/0_mark_down_gen/02B_FT_manual_selector_gui.py",
    ],
    "03_figure_mapping": [
        "agent_workspace/legacy_context/by_stage/03_figure_mapping/README.md",
        "agent_workspace/legacy_context/full_copy/0_mark_down_gen/03_figure_mapping.py",
    ],
    "03_split_excel_blocks_batch": [
        "agent_workspace/legacy_context/by_stage/03_split_excel_blocks_batch/README.md",
        "agent_workspace/legacy_context/full_copy/0_mark_down_gen/03_split_excel_blocks.py",
        "agent_workspace/legacy_context/full_copy/0_mark_down_gen/03_split_excel_blocks_batch.py",
        "agent_workspace/legacy_context/full_copy/0_mark_down_gen/sheet_block_splitter.py",
    ],
    "04_figure_separate": [
        "agent_workspace/legacy_context/by_stage/04_figure_separate/README.md",
        "agent_workspace/legacy_context/full_copy/0_mark_down_gen/04_figure_saperate_gemini.py",
    ],
    "04_ft_excel_matcher": [
        "agent_workspace/legacy_context/by_stage/04_ft_excel_matcher/README.md",
        "agent_workspace/legacy_context/full_copy/0_mark_down_gen/04_FT-Excel_matcher.py",
    ],
    "05_smiles_structure_resolution": [
        "agent_workspace/legacy_context/by_stage/05_smiles_structure_resolution/README.md",
        "agent_workspace/legacy_context/full_copy/2_Extract_SMILES/",
    ],
    "06_unified_lnpdb_extraction": [
        "agent_workspace/legacy_context/by_stage/06_unified_lnpdb_extraction/README.md",
        "agent_workspace/legacy_context/full_copy/1_Extract_Exp_Figs/",
        "agent_workspace/legacy_context/full_copy/3_Extract_Formula_by_Figs/",
        "agent_workspace/legacy_context/full_copy/4_Extract_Exp_Vals/",
    ],
    "07_finalize_unified_table": [
        "agent_workspace/legacy_context/by_stage/06_unified_lnpdb_extraction/README.md",
        "agent_workspace/legacy_context/full_copy/1_Extract_Exp_Figs/",
        "agent_workspace/legacy_context/full_copy/3_Extract_Formula_by_Figs/",
        "agent_workspace/legacy_context/full_copy/4_Extract_Exp_Vals/",
    ],
}

DEFAULT_AGENT_ACTIVE_STAGES = [
    "03_figure_mapping",
    "03_split_excel_blocks_batch",
    "04_figure_separate",
    "04_ft_excel_matcher",
    "05_smiles_structure_resolution",
    "06_unified_lnpdb_extraction",
    "07_finalize_unified_table",
]

DEFAULT_AGENT_COMMAND_TEMPLATES = {
    "codex": 'codex exec --cd "{project_root}" --dangerously-bypass-approvals-and-sandbox --add-dir "{paper_folder}" -',
    "claude": 'claude -p "{prompt_text}"',
}


STAGES: dict[str, dict[str, Any]] = {
    "00_marker": {
        "script": "0_mark_down_gen/00_Marker.py",
        "outputs": ["*.md"],
    },
    "01_make_ft_csv": {
        "script": "0_mark_down_gen/01_make_FT_csv.py",
        "outputs": ["fig_table_inventory.csv"],
    },
    "02_ft_selector": {
        "script": "0_mark_down_gen/02_FT_selector.py",
        "outputs": ["fig_table_lnpdb_classified.csv"],
    },
    "02b_manual_review": {
        "script": "0_mark_down_gen/02B_FT_manual_selector_gui.py",
        "outputs": [MANUAL_MARKER],
        "manual": True,
    },
    "03_figure_mapping": {
        "script": "0_mark_down_gen/03_figure_mapping.py",
        "outputs": ["total_figure_mapping.json"],
        "requires_manual_marker": True,
    },
    "03_split_excel_blocks": {
        "script": "0_mark_down_gen/03_split_excel_blocks.py",
        "outputs": [],
        "requires_manual_marker": True,
        "utility_only": True,
    },
    "03_split_excel_blocks_batch": {
        "script": "0_mark_down_gen/03_split_excel_blocks_batch.py",
        "outputs": ["excel_block_inventory.csv", "three_core_result_all.json", "Exp_Excel_Blocks"],
        "requires_manual_marker": True,
    },
    "04_figure_separate": {
        "script": "0_mark_down_gen/04_figure_saperate_gemini.py",
        "outputs": ["separated_panels_gemini"],
        "requires_manual_marker": True,
    },
    "04_ft_excel_matcher": {
        "script": "0_mark_down_gen/04_FT-Excel_matcher.py",
        "outputs": ["excel_mapping.json", "excel_mapping_rows.csv"],
        "requires_manual_marker": True,
    },
    "05_smiles_structure_resolution": {
        "script": "2_Extract_SMILES/FromIUPAC/Extract_Text_Lipid.py",
        "outputs": ["compound_inventory_standardized.csv", "smiles_resolved.csv", "smiles_resolution_qc.csv"],
        "requires_manual_marker": True,
    },
    "06_unified_lnpdb_extraction": {
        "script": "Agent_Task_Runner.py",
        "outputs": ["unified_extraction.csv", "unified_extraction.json", "unified_extraction_review_flags.csv"],
        "requires_manual_marker": True,
    },
    "07_finalize_unified_table": {
        "script": "Agent_Task_Runner.py",
        "outputs": [
            "unified_extraction_final.csv",
            "unified_extraction_lnpdb_like.csv",
            "unified_extraction_source_evidence.csv",
            "unified_extraction_figure_evidence_map.csv",
            "unified_extraction_qc_report.json",
        ],
        "requires_manual_marker": True,
    },
}

LEGACY_STAGE_SCRIPT_GROUPS: dict[str, list[str]] = {
    "05_smiles_structure_resolution": [
        "2_Extract_SMILES/FromIUPAC/20_Text_to_SMILES.py",
        "2_Extract_SMILES/FromIUPAC/Extract_Text_Lipid.py",
        "2_Extract_SMILES/FromIUPAC/Extract_Lipid_SMILES.py",
        "2_Extract_SMILES/FromImage/mol_annotator/*.py",
    ],
    "06_unified_lnpdb_extraction": [
        "1_Extract_Exp_Figs/10_Extract_from_Excel.py",
        "1_Extract_Exp_Figs/10_Extract_from_PDF_one.py",
        "1_Extract_Exp_Figs/10_Extract_from_PDF_grouped.py",
        "3_Extract_Formula_by_Figs/30_Extract_Formula_by_Excel.py",
        "3_Extract_Formula_by_Figs/30_Extract_Formula_by_Figs.py",
    ],
}

UNIFIED_EXTRACTION_COLUMNS = [
    "Paper_ID",
    "Item_ID",
    "visual_type",
    "source_type",
    "source_image",
    "source_pdf",
    "source_page",
    "selected_source_for_paneling",
    "excel_file",
    "excel_sheet",
    "block_id",
    "block_csv_path",
    "Aqueous_buffer",
    "Dialysis_buffer",
    "Mixing_method",
    "Model",
    "Model_type",
    "Model_target",
    "Route_of_administration",
    "Cargo",
    "Cargo_type",
    "Dose_ug_nucleicacid",
    "Experiment_method",
    "Experiment_batching",
    "formulation_id",
    "Formulation_Name",
    "IL_name",
    "IL_SMILES",
    "IL_molarratio",
    "HL_name",
    "HL_SMILES",
    "HL_molarratio",
    "CHL_name",
    "CHL_SMILES",
    "CHL_molarratio",
    "PEG_name",
    "PEG_SMILES",
    "PEG_molarratio",
    "Fifth_component_name",
    "Fifth_component_SMILES",
    "Fifth_component_molarratio",
    "IL_to_nucleicacid_massratio",
    "condition_1_name",
    "condition_1_value",
    "condition_2_name",
    "condition_2_value",
    "condition_3_name",
    "condition_3_value",
    "condition_4_name",
    "condition_4_value",
    "metric_type",
    "original_values",
    "aggregated_value",
    "unit",
    "replicate_type",
    "evidence_text",
    "evidence_image",
    "evidence_excel",
    "confidence",
    "manual_required",
    "reason",
]

UNIFIED_REVIEW_FLAG_COLUMNS = [
    "Paper_ID",
    "Item_ID",
    "block_id",
    "field",
    "issue",
    "severity",
    "reason",
]

OUTPUT_SMILES_COLUMNS = [
    "IL_SMILES",
    "HL_SMILES",
    "CHL_SMILES",
    "PEG_SMILES",
    "Fifth_component_SMILES",
]

STAGE_06_REFERENCE_COLUMNS = [
    "Aqueous_buffer",
    "Dialysis_buffer",
    "Mixing_method",
    "Model",
    "Model_type",
    "Model_target",
    "Route_of_administration",
    "Cargo",
    "Cargo_type",
    "Dose_ug_nucleicacid",
    "Experiment_method",
    "Experiment_batching",
    "formulation_id",
    "Formulation_Name",
    "IL_name",
    "IL_SMILES",
    "IL_molarratio",
    "HL_name",
    "HL_SMILES",
    "HL_molarratio",
    "CHL_name",
    "CHL_SMILES",
    "CHL_molarratio",
    "PEG_name",
    "PEG_SMILES",
    "PEG_molarratio",
    "Fifth_component_name",
    "Fifth_component_SMILES",
    "Fifth_component_molarratio",
    "IL_to_nucleicacid_massratio",
    "metric_type",
    "original_values",
    "aggregated_value",
    "unit",
    "replicate_type",
]

REFERENCE_TABLE_EXTENSIONS = {".csv", ".tsv", ".xlsx", ".xlsm"}
REFERENCE_TEXT_EXTENSIONS = {".txt", ".md", ".json"}
REFERENCE_EXTENSIONS = REFERENCE_TABLE_EXTENSIONS | REFERENCE_TEXT_EXTENSIONS
REFERENCE_ENCODINGS = ("utf-8-sig", "utf-8", "cp949", "latin-1")
REFERENCE_SKIP_DIRS = {".git", "__pycache__", ".pytest_cache", "node_modules", "logs", "outputs", "output", "tmp", "temp"}

SOURCE_EVIDENCE_COLUMNS = [
    "evidence_id",
    "Paper_ID",
    "Item_ID",
    "evidence_summary",
    "evidence_sentence_ids",
    "evidence_sentence_texts",
    "evidence_text_exact",
    "evidence_text_normalized",
    "evidence_source_type",
    "evidence_source_path",
    "source_pdf",
    "source_page",
    "source_image",
    "evidence_excel",
    "excel_file",
    "excel_sheet",
    "block_id",
    "block_csv_path",
    "excel_cell_or_range",
    "confidence",
    "manual_required",
    "reason",
    "pdf_page_index",
    "pdf_text_quote",
    "pdf_char_start",
    "pdf_char_end",
    "pdf_bbox_x0",
    "pdf_bbox_y0",
    "pdf_bbox_x1",
    "pdf_bbox_y1",
    "image_bbox_x0",
    "image_bbox_y0",
    "image_bbox_x1",
    "image_bbox_y1",
]

FIGURE_EVIDENCE_MAP_COLUMNS = [
    "Paper_ID",
    "Item_ID",
    "evidence_id",
    "supported_columns",
    "supported_column_count",
    "supported_row_ids",
    "supported_formulation_ids",
    "support_scope",
    "evidence_sentence_ids",
    "evidence_summary",
    "confidence",
    "manual_required",
    "reason",
]

MARKDOWN_SENTENCE_INDEX_COLUMNS = [
    "source_md_id",
    "sentence_id",
    "global_sentence_id",
    "source_md_path",
    "source_md_relpath",
    "source_page",
    "section_heading",
    "item_hint",
    "sentence_kind",
    "sentence_text",
    "sentence_text_normalized",
    "char_start",
    "char_end",
]

MARKDOWN_SENTENCE_INDEX_SKIP_DIRS = {
    "agent_workspace",
    "Exp_Excel_Blocks",
    "markdown_sentence_index",
    "separated_panels_gemini",
    "pdf_page_renders",
    "__pycache__",
    ".git",
}

ALLOWED_FIGURE_EVIDENCE_COLUMNS = [
    "Aqueous_buffer",
    "Dialysis_buffer",
    "Mixing_method",
    "Model",
    "Model_type",
    "Model_target",
    "Route_of_administration",
    "Cargo",
    "Cargo_type",
    "Dose_ug_nucleicacid",
    "Experiment_method",
    "Experiment_batching",
    "formulation_id",
    "Formulation_Name",
    "IL_name",
    "IL_molarratio",
    "HL_name",
    "HL_molarratio",
    "CHL_name",
    "CHL_molarratio",
    "PEG_name",
    "PEG_molarratio",
    "Fifth_component_name",
    "Fifth_component_molarratio",
    "IL_to_nucleicacid_massratio",
    "metric_type",
    "original_values",
    "aggregated_value",
    "unit",
    "replicate_type",
]
ALLOWED_FIGURE_EVIDENCE_COLUMN_SET = set(ALLOWED_FIGURE_EVIDENCE_COLUMNS)

EVIDENCE_SOURCE_TYPES = {
    "markdown",
    "pdf_caption",
    "pdf_text",
    "figure_image",
    "excel_block",
    "smiles_resolved",
    "compound_inventory",
    "inferred_from_context",
    "manual_review_placeholder",
    "caption_fallback",
    "image_caption_fallback",
    "figure_caption",
    "methods_global",
    "methods_item_specific",
    "excel_block_value",
    "source_data_excel",
}

LNPDB_LIKE_EXCLUDED_COLUMNS = {"evidence_text", "evidence_image", "evidence_excel"}
ADMINISTRATIVE_EVIDENCE_EXCLUDED_COLUMNS = {"row_id", "Paper_ID", "confidence", "manual_required", "reason"}
SCIENTIFIC_EVIDENCE_OPTIONAL_COLUMNS = {"visual_type", "source_type"}
DISALLOWED_IMAGE_SMILES_TERMS = (
    "decimer",
    "molscribe",
    "worker_mol",
    "structure image",
    "structure_image",
    "image structure",
    "image_structure",
    "structure crop",
    "structure_crop",
    "molecule crop",
    "molecular structure",
    "mol_annotator",
    "recognition.py",
    "segmentation.py",
    "fromimage",
    "pdf/image crop",
)
MANUAL_VERIFIED_SMILES_TERMS = ("human_curated", "manual_verified", "manually_verified", "manual curated", "manual verified")


def utc_now() -> str:
    return datetime.now(timezone.utc).replace(microsecond=0).isoformat()


def load_json(path: Path, default: Any) -> Any:
    if not path.exists():
        return default
    try:
        return json.loads(path.read_text(encoding="utf-8"))
    except Exception:
        return default


def write_json(path: Path, data: Any) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    path.write_text(json.dumps(data, ensure_ascii=False, indent=2) + "\n", encoding="utf-8")


def import_pandas_for_reference() -> tuple[Any | None, str]:
    try:
        import pandas as pd  # type: ignore

        return pd, ""
    except Exception as exc:
        return None, f"pandas/openpyxl support unavailable for spreadsheet reference reading: {exc}"


def safe_reference_text(value: Any) -> str:
    if value is None:
        return ""
    text = str(value).strip()
    if text.lower() in {"nan", "none", "nat"}:
        return ""
    return re.sub(r"\s+", " ", text)


def truncate_text(text: str, max_chars: int) -> str:
    if len(text) <= max_chars:
        return text
    return text[:max_chars] + "\n[truncated]"


def read_text_with_encoding_fallback(path: Path, max_chars: int = 20_000) -> tuple[str, str]:
    for encoding in REFERENCE_ENCODINGS:
        try:
            return truncate_text(path.read_text(encoding=encoding), max_chars), ""
        except UnicodeDecodeError:
            continue
        except Exception as exc:
            return "", f"Could not read text reference {path}: {exc}"
    try:
        return truncate_text(path.read_text(errors="replace"), max_chars), ""
    except Exception as exc:
        return "", f"Could not read text reference {path}: {exc}"


def should_skip_reference_path(path: Path) -> bool:
    if path.name.startswith("~$"):
        return True
    for part in path.parts:
        if part.startswith(".") or part.lower() in REFERENCE_SKIP_DIRS:
            return True
    return False


def reference_file_key(path: Path) -> str:
    try:
        return str(path.resolve()).lower()
    except Exception:
        return str(path.absolute()).lower()


def collect_reference_files_from_root(root: Path, warnings: list[str]) -> list[Path]:
    try:
        if root.is_file():
            if root.suffix.lower() in REFERENCE_EXTENSIONS and not root.name.startswith("~$"):
                return [root]
            warnings.append(f"Unsupported reference file skipped: {root}")
            return []
        if root.is_dir():
            files = [
                path
                for path in root.rglob("*")
                if path.is_file()
                and path.suffix.lower() in REFERENCE_EXTENSIONS
                and not should_skip_reference_path(path)
            ]
            return sorted(files, key=lambda p: str(p).lower())
        warnings.append(f"Reference path not found: {root}")
        return []
    except Exception as exc:
        warnings.append(f"Could not scan reference path {root}: {exc}")
        return []


def discover_reference_files(paper_folder: Path) -> dict[str, Any]:
    warnings: list[str] = []
    lnpdb_roots: list[Path] = []
    guide_roots: list[Path] = []
    for name in ("LNPDB_reference", "lnpdb_reference"):
        for suffix in REFERENCE_EXTENSIONS:
            lnpdb_roots.append(paper_folder / f"{name}{suffix}")
            lnpdb_roots.append(paper_folder / "reference" / f"{name}{suffix}")
            lnpdb_roots.append(WORKSPACE / "reference" / f"{name}{suffix}")
    for reference_dir in (paper_folder / "reference", WORKSPACE / "reference"):
        if reference_dir.exists():
            lnpdb_roots.append(reference_dir)
    env_lnpdb = os.environ.get("LNPDB_REFERENCE_ROOT", "").strip()
    if env_lnpdb:
        lnpdb_roots.append(Path(env_lnpdb))

    for rel in ("column_guides", "schema_guides", "value_guides", "reference"):
        guide_roots.append(paper_folder / rel)
    for rel in ("reference", "column_guides", "schema_guides", "value_guides"):
        guide_roots.append(WORKSPACE / rel)
    for env_name in ("LNPDB_COLUMN_GUIDE_ROOT", "LNPDB_SCHEMA_GUIDE_ROOT", "LNPDB_VALUE_GUIDE_ROOT"):
        value = os.environ.get(env_name, "").strip()
        if value:
            guide_roots.append(Path(value))

    def is_named_lnpdb_reference(path: Path) -> bool:
        return "lnpdb" in path.stem.lower()

    def files_for_roots(roots: list[Path], exclude_lnpdb_named: bool = False, only_lnpdb_named: bool = False) -> list[Path]:
        files: list[Path] = []
        seen: set[str] = set()
        for root in roots:
            if not root.exists():
                continue
            for path in collect_reference_files_from_root(root, warnings):
                if only_lnpdb_named and root.is_dir() and "lnpdb" not in path.stem.lower():
                    continue
                if exclude_lnpdb_named and is_named_lnpdb_reference(path):
                    continue
                key = reference_file_key(path)
                if key in seen:
                    continue
                seen.add(key)
                files.append(path)
        return sorted(files, key=lambda p: str(p).lower())

    return {
        "lnpdb_reference_files": files_for_roots(lnpdb_roots, only_lnpdb_named=True),
        "human_guide_files": files_for_roots(guide_roots, exclude_lnpdb_named=True),
        "warnings": warnings,
    }


def column_lookup(columns: list[str], target: str) -> str | None:
    target_key = re.sub(r"[^a-z0-9]+", "", target.lower())
    for column in columns:
        if re.sub(r"[^a-z0-9]+", "", str(column).lower()) == target_key:
            return str(column)
    return None


def add_reference_column_values(
    summaries: dict[str, dict[str, Any]],
    target_column: str,
    source: str,
    values: list[Any],
    max_top_values: int = 30,
    max_examples: int = 20,
) -> None:
    entry = summaries.setdefault(
        target_column,
        {
            "column_exists": True,
            "sources": [],
            "non_empty_count": 0,
            "unique_count": 0,
            "top_values": [],
            "example_values": [],
            "_counts": Counter(),
            "_examples_seen": set(),
            "_style_examples": [],
            "_style_examples_seen": set(),
        },
    )
    if source not in entry["sources"]:
        entry["sources"].append(source)
    for value in values:
        text = safe_reference_text(value)
        if not text or should_ignore_reference_value(target_column, text):
            continue
        entry["non_empty_count"] += 1
        entry["_counts"][text] += 1
        if len(entry["example_values"]) < max_examples and text not in entry["_examples_seen"]:
            entry["example_values"].append(text)
            entry["_examples_seen"].add(text)
        if target_column == "Experiment_method" and is_readout_specific_reference_example(text) and text.lower() not in entry["_style_examples_seen"]:
            entry["_style_examples"].append(text)
            entry["_style_examples_seen"].add(text.lower())


def should_ignore_reference_value(column: str, value: str) -> bool:
    text = value.strip()
    if not text:
        return True
    lowered = text.lower()
    if lowered in {"nan", "none", "null", "n/a"}:
        return False
    if column.lower().endswith("id") or column in {"formulation_id"}:
        return False
    if re.match(r"^[a-z]:[\\/]", lowered) or lowered.startswith(("/", "\\", "http://", "https://")):
        return True
    if any(token in lowered for token in ("\\", "/users/", "/home/", ".xlsx", ".csv", ".png", ".jpg", ".pdf")):
        return True
    if len(text) > 160:
        return True
    return False


def is_readout_specific_reference_example(value: str) -> bool:
    return bool(
        "_" in value
        or re.search(r"(CD8|CD4|Treg|IL-?12|IFN|TNF|OVA|FLuc|NLuc|tumou?r|spleen|T_cells?)", value, flags=re.I)
    )


def build_column_examples_from_summary(summary: dict[str, Any], column: str, limit: int = 20) -> list[str]:
    values: list[str] = []
    seen: set[str] = set()

    def add(value: Any) -> None:
        text = safe_reference_text(value)
        if not text or text.lower() in seen:
            return
        values.append(text)
        seen.add(text.lower())

    for item in summary.get("top_values", []):
        add(item.get("value", ""))
        if len(values) >= limit:
            break
    if column == "Experiment_method":
        # Preserve readout-specific rare examples instead of collapsing to assay prefixes.
        candidates = [item.get("value", "") for item in summary.get("top_values", [])] + list(summary.get("_style_examples", [])) + list(summary.get("example_values", []))
        specific = [value for value in candidates if isinstance(value, str) and is_readout_specific_reference_example(value)]
        for value in specific:
            add(value)
            if len(values) >= limit:
                break
    for value in summary.get("example_values", []):
        add(value)
        if len(values) >= limit:
            break
    return values[:limit]


def finalize_reference_column_summaries(summaries: dict[str, dict[str, Any]], max_top_values: int = 30) -> dict[str, dict[str, Any]]:
    finalized: dict[str, dict[str, Any]] = {}
    for column, entry in summaries.items():
        counts = entry.pop("_counts", Counter())
        entry.pop("_examples_seen", None)
        entry.pop("_style_examples_seen", None)
        top = sorted(counts.items(), key=lambda item: (-item[1], item[0].lower()))
        entry["unique_count"] = len(top)
        entry["top_values"] = [{"value": value, "count": count} for value, count in top[:max_top_values]]
        entry["column_examples"] = build_column_examples_from_summary(entry, column)
        entry.pop("_style_examples", None)
        finalized[column] = entry
    return finalized


def summarize_reference_dataframe(df: Any, source: str, summaries: dict[str, dict[str, Any]]) -> None:
    columns = [str(col) for col in getattr(df, "columns", [])]
    for target in STAGE_06_REFERENCE_COLUMNS:
        actual = column_lookup(columns, target)
        if actual is None:
            continue
        try:
            values = list(df[actual].tolist())
        except Exception:
            continue
        add_reference_column_values(summaries, target, source, values)


def read_reference_table_file(path: Path, warnings: list[str]) -> list[tuple[str, Any]]:
    suffix = path.suffix.lower()
    if suffix in {".csv", ".tsv"}:
        delimiter = "\t" if suffix == ".tsv" else ","
        last_error = ""
        for encoding in REFERENCE_ENCODINGS:
            try:
                with path.open("r", encoding=encoding, newline="") as handle:
                    rows = list(csv.DictReader(handle, delimiter=delimiter))
                return [(path.name, rows)]
            except UnicodeDecodeError:
                continue
            except Exception as exc:
                last_error = str(exc)
        warnings.append(f"Could not read tabular reference {path}: {last_error}")
        return []
    pd, warning = import_pandas_for_reference()
    if pd is None:
        warnings.append(f"{path}: {warning}")
        return []
    try:
        book = pd.ExcelFile(path)
    except Exception as exc:
        warnings.append(f"Could not open spreadsheet reference {path}: {exc}")
        return []
    sheets: list[tuple[str, Any]] = []
    for sheet_name in book.sheet_names:
        try:
            sheets.append((str(sheet_name), pd.read_excel(book, sheet_name=sheet_name)))
        except Exception as exc:
            warnings.append(f"Could not read spreadsheet sheet {path}::{sheet_name}: {exc}")
    return sheets


def summarize_reference_rows(rows: list[dict[str, Any]], source: str, summaries: dict[str, dict[str, Any]]) -> None:
    columns = list(rows[0].keys()) if rows else []
    for target in STAGE_06_REFERENCE_COLUMNS:
        actual = column_lookup(columns, target)
        if actual is None:
            continue
        add_reference_column_values(summaries, target, source, [row.get(actual, "") for row in rows])


def rows_to_markdown_preview(rows: list[dict[str, Any]], source: str, max_rows: int = 50, max_chars: int = 20_000) -> str:
    columns = list(rows[0].keys()) if rows else []
    lines = [f"Source: {source}", "Columns: " + ", ".join(columns)]
    if columns:
        preview_columns = columns[:40]
        lines.append("| " + " | ".join(preview_columns) + " |")
        lines.append("| " + " | ".join(["---"] * len(preview_columns)) + " |")
        for row in rows[:max_rows]:
            values = [safe_reference_text(row.get(col, "")).replace("|", "/")[:160] for col in preview_columns]
            if any(values):
                lines.append("| " + " | ".join(values) + " |")
    return truncate_text("\n".join(lines), max_chars)


def dataframe_to_markdown_preview(df: Any, source: str, max_rows: int = 50, max_chars: int = 20_000) -> str:
    rows = []
    try:
        records = df.to_dict(orient="records")
    except Exception:
        records = []
    for record in records:
        cleaned = {str(key): safe_reference_text(value) for key, value in record.items()}
        if any(cleaned.values()):
            rows.append(cleaned)
    return rows_to_markdown_preview(rows[:max_rows], source, max_rows=max_rows, max_chars=max_chars)


def read_reference_text_file(path: Path, max_chars: int = 20_000) -> tuple[str, str]:
    if path.suffix.lower() == ".json":
        text, warning = read_text_with_encoding_fallback(path, max_chars=max_chars)
        if warning:
            return "", warning
        try:
            parsed = json.loads(text)
            return truncate_text(json.dumps(parsed, ensure_ascii=False, indent=2), max_chars), ""
        except Exception:
            return text, ""
    return read_text_with_encoding_fallback(path, max_chars=max_chars)


def collect_reference_context_for_06(paper_folder: Path) -> dict[str, Any]:
    discovered = discover_reference_files(paper_folder)
    warnings = list(discovered.get("warnings", []))
    column_summaries: dict[str, dict[str, Any]] = {}
    lnpdb_sources: list[str] = []
    lnpdb_notes: list[dict[str, Any]] = []

    for path in discovered.get("lnpdb_reference_files", []):
        suffix = path.suffix.lower()
        if suffix in REFERENCE_TABLE_EXTENSIONS:
            for sheet_name, data in read_reference_table_file(path, warnings):
                source = f"{path}::{sheet_name}" if suffix in {".xlsx", ".xlsm"} else str(path)
                lnpdb_sources.append(str(path))
                if isinstance(data, list):
                    summarize_reference_rows(data, source, column_summaries)
                else:
                    summarize_reference_dataframe(data, source, column_summaries)
        elif suffix in REFERENCE_TEXT_EXTENSIONS:
            text, warning = read_reference_text_file(path)
            if warning:
                warnings.append(warning)
                continue
            lnpdb_sources.append(str(path))
            lnpdb_notes.append({"source_path": str(path), "content": text, "included_chars": len(text)})

    guide_sources: list[str] = []
    guide_blocks: list[dict[str, Any]] = []
    for path in discovered.get("human_guide_files", []):
        suffix = path.suffix.lower()
        if suffix in {".csv", ".tsv"}:
            for sheet_name, data in read_reference_table_file(path, warnings):
                source = str(path)
                guide_sources.append(str(path))
                if isinstance(data, list):
                    content = rows_to_markdown_preview(data, source)
                else:
                    content = dataframe_to_markdown_preview(data, source)
                guide_blocks.append({"source_path": source, "content": content, "included_chars": len(content)})
        elif suffix in {".xlsx", ".xlsm"}:
            for sheet_name, data in read_reference_table_file(path, warnings):
                source = f"{path}::{sheet_name}"
                guide_sources.append(str(path))
                content = dataframe_to_markdown_preview(data, source)
                guide_blocks.append({"source_path": source, "content": content, "included_chars": len(content)})
        elif suffix in REFERENCE_TEXT_EXTENSIONS:
            text, warning = read_reference_text_file(path)
            if warning:
                warnings.append(warning)
                continue
            guide_sources.append(str(path))
            guide_blocks.append({"source_path": str(path), "content": text, "included_chars": len(text)})

    if not lnpdb_sources and not guide_sources:
        warnings.append("No external LNPDB reference schema/value context was available; proceed using paper evidence only.")

    finalized_columns = finalize_reference_column_summaries(column_summaries)
    column_examples = {
        column: summary.get("column_examples", [])
        for column, summary in finalized_columns.items()
        if summary.get("column_examples")
    }
    return {
        "has_any_reference": bool(finalized_columns or lnpdb_notes or guide_blocks),
        "lnpdb_reference": {
            "available": bool(finalized_columns or lnpdb_notes),
            "source_paths": sorted(set(lnpdb_sources)),
            "columns": finalized_columns,
            "column_examples": column_examples,
            "free_text_notes": lnpdb_notes,
        },
        "human_column_guides": {
            "available": bool(guide_blocks),
            "source_paths": sorted(set(guide_sources)),
            "text_blocks": guide_blocks,
        },
        "warnings": warnings,
    }


def render_reference_context_for_prompt(reference_context: dict[str, Any]) -> str:
    lines: list[str] = []
    if not reference_context.get("has_any_reference"):
        no_context = "No external LNPDB reference schema/value context was available; proceed using paper evidence only."
        lines.extend(
            [
                "## Existing LNPDB Column/Value Examples",
                no_context,
                "",
                "## Column-Specific Existing LNPDB Examples",
                no_context,
                "",
                "## Human-Curated Column and Value Definitions",
                no_context,
                "",
                "## Reference Context Warnings",
            ]
        )
        for warning in reference_context.get("warnings", []):
            lines.append(f"- {warning}")
        lines.extend(
            [
                "",
                "Reference-context rules:",
                "- Existing LNPDB values are examples, not a closed vocabulary.",
                "- Human-curated definitions are higher priority than frequency examples.",
                "- Use reference examples to normalize values into concise scalar LNPDB-style values.",
                "- Do not copy full source prose into LNPDB fields.",
                "- Full source sentences/captions belong only in `evidence_text`.",
                "- If a concise normalized value cannot be determined, leave blank and set `manual_required=true` with a reason.",
            ]
        )
        return "\n".join(lines)

    lines.append("## Existing LNPDB Column/Value Examples")
    lines.append("Existing LNPDB values are examples, not a closed vocabulary.")
    lnpdb_ref = reference_context.get("lnpdb_reference", {})
    columns = lnpdb_ref.get("columns", {})
    if columns:
        for column, summary in columns.items():
            lines.append(f"- `{column}`: column_exists={summary.get('column_exists', False)}, non_empty_count={summary.get('non_empty_count', 0)}, unique_count={summary.get('unique_count', 0)}")
            if summary.get("sources"):
                lines.append(f"  - sources: {'; '.join(summary.get('sources', [])[:5])}")
            top_values = summary.get("top_values", [])[:10]
            if top_values:
                rendered_top = "; ".join(f"{item.get('value')} ({item.get('count')})" for item in top_values)
                lines.append(f"  - top values: {rendered_top}")
            examples = summary.get("example_values", [])[:8]
            if examples:
                lines.append(f"  - examples: {'; '.join(examples)}")
    else:
        lines.append("- No tabular LNPDB column/value examples were available.")
    for note in lnpdb_ref.get("free_text_notes", [])[:5]:
        lines.append(f"\nReference note from `{note.get('source_path')}`:\n{note.get('content', '')}")

    lines.append("\n## Column-Specific Existing LNPDB Examples")
    lines.append("Use these as style examples. They are not a closed vocabulary, but new values should follow the same concise scalar naming style.")
    column_examples = lnpdb_ref.get("column_examples", {})
    if column_examples:
        ordered_columns = [col for col in STAGE_06_REFERENCE_COLUMNS if col in column_examples]
        ordered_columns.extend(col for col in column_examples if col not in ordered_columns)
        for column in ordered_columns:
            examples = column_examples.get(column, [])
            if not examples:
                continue
            lines.append(f"\n### {column}")
            lines.append("Existing LNPDB examples:")
            for example in examples[:20]:
                lines.append(f"- {example}")
    else:
        lines.append("- No column-specific existing LNPDB examples were available.")

    lines.append("\n## Human-Curated Column and Value Definitions")
    lines.append("Human-curated definitions are higher priority than frequency examples.")
    guide = reference_context.get("human_column_guides", {})
    if guide.get("text_blocks"):
        for block in guide.get("text_blocks", [])[:8]:
            lines.append(f"\nGuide from `{block.get('source_path')}`:\n{block.get('content', '')}")
    else:
        lines.append("- No human-curated guide files were available.")

    lines.append("\n## Reference Context Warnings")
    warnings = reference_context.get("warnings", [])
    if warnings:
        for warning in warnings:
            lines.append(f"- {warning}")
    else:
        lines.append("- none")
    lines.extend(
        [
            "",
            "Reference-context rules:",
            "- Existing LNPDB values are examples, not a closed vocabulary.",
            "- Human-curated definitions are higher priority than frequency examples.",
            "- Use reference examples to normalize values into concise scalar LNPDB-style values.",
            "- Do not copy full source prose into LNPDB fields.",
            "- Full source sentences/captions belong only in `evidence_text`.",
            "- If a concise normalized value cannot be determined, leave blank and set `manual_required=true` with a reason.",
        ]
    )
    return "\n".join(lines)


def reference_context_debug_path(stage: str, paper_folder: Path) -> Path:
    return TASK_DIR / f"{stage.replace('_unified_lnpdb_extraction', '')}_reference_context_{safe_name(paper_folder.name)}.json"


def summarize_reference_context(reference_context: dict[str, Any]) -> dict[str, Any]:
    lnpdb_ref = reference_context.get("lnpdb_reference", {})
    guide = reference_context.get("human_column_guides", {})
    return {
        "has_any_reference": bool(reference_context.get("has_any_reference")),
        "lnpdb_reference_available": bool(lnpdb_ref.get("available")),
        "lnpdb_source_paths": len(lnpdb_ref.get("source_paths", [])),
        "lnpdb_columns_with_examples": len(lnpdb_ref.get("columns", {})),
        "lnpdb_column_examples": len(lnpdb_ref.get("column_examples", {})),
        "lnpdb_free_text_notes": len(lnpdb_ref.get("free_text_notes", [])),
        "human_column_guides_available": bool(guide.get("available")),
        "human_guide_source_paths": len(guide.get("source_paths", [])),
        "human_guide_text_blocks": len(guide.get("text_blocks", [])),
        "warnings": reference_context.get("warnings", []),
    }


def append_log(paper_folder: Path, event: dict[str, Any]) -> None:
    LOG_DIR.mkdir(parents=True, exist_ok=True)
    safe_name = paper_folder.name or "root"
    log_path = LOG_DIR / f"{safe_name}.jsonl"
    record = {"timestamp": utc_now(), "paper_folder": str(paper_folder), **event}
    with log_path.open("a", encoding="utf-8") as f:
        f.write(json.dumps(record, ensure_ascii=False) + "\n")


def update_state(paper_folder: Path, stage: str | None, status: str, detail: dict[str, Any] | None = None) -> None:
    state = load_json(
        STATE_PATH,
        {
            "schema_version": 1,
            "project": "LNPDB_Articles_AgentExtraction",
            "mode": "external_cli_agent_workspace",
            "stage_status": {},
        },
    )
    state["active_paper_folder"] = str(paper_folder)
    state["current_stage"] = stage
    state["last_updated"] = utc_now()
    state["last_event"] = {"stage": stage, "status": status, "detail": detail or {}}
    if stage:
        state.setdefault("stage_status", {})[stage] = {
            "status": status,
            "updated": state["last_updated"],
            "detail": detail or {},
        }
    state["manual_review_required"] = not (paper_folder / MANUAL_MARKER).exists()
    write_json(STATE_PATH, state)


def resolve_paper_folder(value: str) -> Path:
    path = Path(value).expanduser()
    if not path.is_absolute():
        path = PROJECT_ROOT / path
    return path.resolve()


def import_module_from_path(module_name: str, script_path: Path):
    if str(script_path.parent) not in sys.path:
        sys.path.insert(0, str(script_path.parent))
    if str(PROJECT_ROOT) not in sys.path:
        sys.path.insert(0, str(PROJECT_ROOT))
    spec = importlib.util.spec_from_file_location(module_name, script_path)
    if spec is None or spec.loader is None:
        raise ImportError(f"Cannot import {script_path}")
    module = importlib.util.module_from_spec(spec)
    spec.loader.exec_module(module)
    return module


def read_csv_rows(path: Path) -> list[dict[str, str]]:
    with path.open("r", encoding="utf-8-sig", newline="") as f:
        return list(csv.DictReader(f))


def non_empty_file(path: Path) -> bool:
    return path.is_file() and path.stat().st_size > 0


def has_manual_marker(paper_folder: Path) -> bool:
    return (paper_folder / MANUAL_MARKER).exists()


def iter_paper_files(paper_folder: Path):
    for path in paper_folder.rglob("*"):
        try:
            rel_parts = path.relative_to(paper_folder).parts
        except ValueError:
            rel_parts = path.parts
        if any(part in PROJECT_EXCLUDE_DIRS for part in rel_parts):
            continue
        if path.is_file():
            yield path


def rel_to_paper(path: Path, paper_folder: Path) -> str:
    try:
        return path.resolve().relative_to(paper_folder.resolve()).as_posix()
    except ValueError:
        return str(path)


def path_from_mapping_value(paper_folder: Path, value: Any) -> Path | None:
    if not isinstance(value, str) or not value.strip():
        return None
    path = Path(value.strip())
    if not path.is_absolute():
        path = paper_folder / path
    return path


def iter_total_figure_mapping_entries(data: Any):
    if not isinstance(data, dict):
        return
    for key, value in data.items():
        if not isinstance(value, dict):
            continue
        if isinstance(key, str) and key.startswith("_"):
            continue
        if any(field in value for field in ("source_image", "source_pdf", "source_page", "panels", "caption")):
            yield value
            continue
        for item_key, entry in value.items():
            if isinstance(item_key, str) and item_key.startswith("_"):
                continue
            if isinstance(entry, dict):
                yield entry


def render_pdf_page(paper_folder: Path, source_pdf: str, source_page: int, dpi: int = 220) -> str:
    try:
        import fitz
    except ImportError as exc:
        raise RuntimeError("PyMuPDF is required for PDF page render fallback. Install package 'pymupdf'.") from exc

    if source_page < 1:
        raise ValueError(f"source_page must be a 1-based positive integer, got {source_page!r}")

    pdf_path = Path(source_pdf)
    if not pdf_path.is_absolute():
        pdf_path = paper_folder / pdf_path
    if not pdf_path.is_file():
        raise FileNotFoundError(f"source_pdf does not exist: {pdf_path}")

    out_dir = paper_folder / "pdf_page_renders"
    out_dir.mkdir(parents=True, exist_ok=True)
    out_path = out_dir / f"{pdf_path.stem}_page_{source_page:03d}.png"
    if out_path.exists():
        return rel_to_paper(out_path, paper_folder)

    doc = fitz.open(str(pdf_path))
    try:
        if source_page > doc.page_count:
            raise ValueError(f"source_page {source_page} exceeds page count {doc.page_count} for {pdf_path}")
        page = doc[source_page - 1]
        pix = page.get_pixmap(dpi=dpi)
        pix.save(str(out_path))
    finally:
        doc.close()
    return rel_to_paper(out_path, paper_folder)


def is_local_path_value(value: Any) -> bool:
    if not isinstance(value, str) or not value.strip():
        return False
    lowered = value.strip().lower()
    return not re.match(r"^[a-z][a-z0-9+.-]*://", lowered) and not lowered.startswith("data:")


def missing_mapping_paths(paper_folder: Path, data: Any, fields: set[str]) -> list[str]:
    missing: list[str] = []
    for entry in iter_total_figure_mapping_entries(data):
        for field in fields:
            value = entry.get(field)
            if not is_local_path_value(value):
                continue
            path = path_from_mapping_value(paper_folder, value)
            if path and not path.exists():
                missing.append(f"{field}={value}")
    return missing


def invalid_source_quality_values(data: Any) -> list[str]:
    invalid: list[str] = []
    for entry in iter_total_figure_mapping_entries(data):
        value = entry.get("source_quality")
        if value is None or value == "":
            continue
        if str(value) not in VALID_SOURCE_QUALITIES:
            invalid.append(str(value))
    return invalid


def safe_name(value: str) -> str:
    cleaned = re.sub(r"[^A-Za-z0-9_.-]+", "_", value.strip())
    return cleaned.strip("._") or "paper"


def normalize_token(value: str) -> str:
    return re.sub(r"[^a-z0-9]+", "", value.lower())


def truthy_selection(value: str) -> bool:
    return value.strip().lower() in {"yes", "maybe", "y", "true", "1", "selected"}


def row_item_id(row: dict[str, str]) -> str:
    for key in ("item_id", "pdf_item_id", "item", "label", "base_id"):
        value = (row.get(key) or "").strip()
        if value:
            return value
    return "unknown_item"


def selected_ft_rows(rows: list[dict[str, str]]) -> list[dict[str, str]]:
    if not rows:
        return []
    columns = set(rows[0])
    if "manual_select" in columns:
        selected = [row for row in rows if truthy_selection(row.get("manual_select", ""))]
        if selected:
            return selected
    if "need_for_lnpdb" in columns:
        return [row for row in rows if truthy_selection(row.get("need_for_lnpdb", ""))]
    for fallback in ("is_lnpdb", "lnpdb_relevant", "selected"):
        if fallback in columns:
            selected = [row for row in rows if truthy_selection(row.get(fallback, ""))]
            if selected:
                return selected
    return []


def require_existing_file(path: Path, stage: str) -> None:
    if path.name == MANUAL_MARKER and path.exists():
        return
    if not non_empty_file(path):
        raise FileNotFoundError(f"{stage} requires a non-empty file: {path}")


def task_file_path(stage: str, paper_folder: Path) -> Path:
    TASK_DIR.mkdir(parents=True, exist_ok=True)
    return TASK_DIR / f"{stage}_{safe_name(paper_folder.name)}.md"


def write_task_file(stage: str, paper_folder: Path, content: str) -> Path:
    path = task_file_path(stage, paper_folder)
    path.write_text(content.rstrip() + "\n", encoding="utf-8")
    return path


def asset_list(paper_folder: Path, suffixes: set[str], limit: int = 200) -> list[str]:
    found = [rel_to_paper(p, paper_folder) for p in iter_paper_files(paper_folder) if p.suffix.lower() in suffixes]
    return found[:limit]


def render_bullet_list(items: list[str]) -> str:
    if not items:
        return "- none found"
    return "\n".join(f"- `{item}`" for item in items)


def render_legacy_context_section(stage: str) -> str:
    context_files = LEGACY_CONTEXT_BY_STAGE.get(stage, [])
    if not context_files:
        return ""
    lines = [
        "",
        "## Legacy context files",
        "",
        "These files are reference-only. Do not execute or import Gemini/API-dependent scripts. Use them only to understand prior deterministic logic, expected output shapes, and naming conventions. Current stage contract and `AGENT_INSTRUCTIONS.md` override legacy behavior.",
        "",
    ]
    for item in context_files:
        path = Path(item)
        status = "exists" if path.exists() else "missing"
        lines.append(f"- `{item}` ({status})")
    lines.extend(
        [
            "- `agent_workspace/legacy_context/README.md` (context policy)",
            "- `agent_workspace/legacy_context/LEGACY_CODE_INDEX.md` (stage-by-stage index)",
            "",
        ]
    )
    return "\n".join(lines)


def create_external_agent_task(stage: str, paper_folder: Path) -> dict[str, Any]:
    if stage == "03_figure_mapping":
        require_existing_file(paper_folder / MANUAL_MARKER, stage)
        require_existing_file(paper_folder / "fig_table_lnpdb_classified.csv", stage)
        require_existing_file(paper_folder / "fig_table_inventory.csv", stage)
        images = asset_list(paper_folder, {".png", ".jpg", ".jpeg", ".tif", ".tiff", ".bmp"})
        tables = asset_list(paper_folder, {".csv", ".xlsx", ".xls"})
        pdfs = asset_list(paper_folder, {".pdf"})
        content = f"""# External Agent Task: 03_figure_mapping

Target paper folder: `{paper_folder}`

## Stage Purpose
Map manually selected LNPDB-relevant figure/table items to source image, table, or PDF assets without using Gemini or any Python API key dependency.

## Required Input Files
- `{paper_folder / MANUAL_MARKER}`
- `{paper_folder / "fig_table_lnpdb_classified.csv"}`
- `{paper_folder / "fig_table_inventory.csv"}`

## Source Assets Found
Images:
{render_bullet_list(images)}

Tables:
{render_bullet_list(tables)}

PDFs:
{render_bullet_list(pdfs)}

## Expected Output Files
- `{paper_folder / "total_figure_mapping.json"}`

## Work Instructions
1. Read `fig_table_lnpdb_classified.csv`.
2. Use rows where `manual_select` is `yes` or `maybe` as selected FT items.
3. If `manual_select` is absent, fall back to `need_for_lnpdb` values `yes` or `maybe`.
4. For every selected item, inspect `item_id`, `base_id`, `caption`, and `reason`.
5. Search the paper folder for source image, table, and PDF assets.
6. Treat Marker-extracted `_page_x_Figure_y.jpeg` images as primary candidates only, not ground truth.
7. Map each selected figure/table item to the most likely source image/table path.
8. When possible, record `source_pdf` and 1-based `source_page` for each mapping entry.
9. Infer `source_page` from markdown image/caption page and order when explicit page metadata is unavailable.
10. If the source image is far from the caption, appears to include only part of the figure, or does not match the expected panel count, set `source_quality: "suspect_crop"`.
11. If no plausible image exists, set `source_quality: "missing_image"` and `manual_required: true`.
12. If image and caption appear mismatched, set `source_quality: "caption_image_mismatch"`.
13. If the source image is complete and caption-consistent, set `source_quality: "ok"`.
14. Create `total_figure_mapping.json` in the paper folder root.
15. Follow `agent_workspace/OUTPUT_SCHEMA.md` for the `total_figure_mapping.json` schema.
16. Store paths relative to the paper folder when possible.
17. If uncertain, record `confidence: "low"` or `confidence: "unmatched"` and a short `reason`; do not guess.

## Optional Mapping Fields
- `source_image`
- `source_pdf`
- `source_page`
- `source_quality`
- `fallback_render`
- `selected_source_for_paneling`
- `manual_required`
- `confidence`
- `reason`

## Validation Command
```bash
python Agent_Task_Runner.py validate --stage 03_figure_mapping --paper-folder "{paper_folder}"
```

## Constraints
- Do not run `0_mark_down_gen/03_figure_mapping.py`.
- Do not import or require `find_api.py`.
- Do not use Gemini, Vertex, `LLM_API.py`, or `LLM_Batch.py`.
- Do not hard-code API keys or credentials.
"""
    elif stage == "03_split_excel_blocks_batch":
        require_existing_file(paper_folder / MANUAL_MARKER, stage)
        exp_excel = paper_folder / "Exp_Excel"
        excel_files = [
            rel_to_paper(p, paper_folder)
            for p in iter_paper_files(exp_excel)
            if p.suffix.lower() in {".xlsx", ".xls", ".csv"}
        ] if exp_excel.exists() else []
        content = f"""# External Agent Task: 03_split_excel_blocks_batch

Target paper folder: `{paper_folder}`

## Stage Purpose
Split experimental Excel workbooks/sheets into API-free table blocks and classify block type by direct CLI agent judgment.

## Required Input Files
- `{paper_folder / MANUAL_MARKER}`
- Excel files under `{paper_folder / "Exp_Excel"}`

Excel files found:
{render_bullet_list(excel_files)}

## Expected Output Files
- `{paper_folder / "Exp_Excel_Blocks"}`
- `{paper_folder / "excel_block_inventory.csv"}`
- `{paper_folder / "three_core_result_all.json"}`

## Work Instructions
1. Inspect the `Exp_Excel` folder.
2. Read Excel workbooks and sheets.
3. Split sheets into candidate blocks using merged cells, blank rows/columns, borders, fills, headers, and numeric density.
4. Prefer API-free helper logic such as `0_mark_down_gen/sheet_block_splitter.py`, `0_mark_down_gen/03_split_excel_blocks.py` pure utilities, or a deterministic helper script.
5. Do not use Gemini or LLM judgment.
6. Save each block CSV under `Exp_Excel_Blocks/`.
7. Create `excel_block_inventory.csv` with required columns:
   - `excel_file`
   - `excel_sheet`
   - `block_id`
   - `group_id`
   - `element_id`
   - `block_csv_path`
   - `block_meta_path`
   - `block_type`
8. Classify `block_type` by direct inspection as one of:
   - `title_and_table`
   - `table_body`
   - `table_title`
   - `multi_table`
   - `note`
   - `other`
9. Create `three_core_result_all.json` with JSON reasoning for every workbook/sheet.
10. If useful, create or run an API-free helper such as `agent_workspace/tools/api_free_excel_block_splitter.py`; use pandas/openpyxl deterministic parsing only.

## Validation Command
```bash
python Agent_Task_Runner.py validate --stage 03_split_excel_blocks_batch --paper-folder "{paper_folder}"
```

## Constraints
- Do not run `0_mark_down_gen/03_split_excel_blocks_batch.py`.
- Do not import or require `find_api.py`.
- Do not use Gemini, Vertex, `LLM_API.py`, or `LLM_Batch.py`.
- Do not hard-code API keys or credentials.
"""
    elif stage == "04_figure_separate":
        require_existing_file(paper_folder / MANUAL_MARKER, stage)
        require_existing_file(paper_folder / "total_figure_mapping.json", stage)
        content = f"""# External Agent Task: 04_figure_separate

Target paper folder: `{paper_folder}`

## Stage Purpose
Separate mapped source images into panels or mark entries for manual review without using Gemini.

## Required Input Files
- `{paper_folder / MANUAL_MARKER}`
- `{paper_folder / "total_figure_mapping.json"}`

## Expected Output Files
- `{paper_folder / "separated_panels_gemini"}`
- updated `{paper_folder / "total_figure_mapping.json"}`

## Work Instructions
1. Read `total_figure_mapping.json`.
2. For each entry, first inspect `source_image`.
3. If `source_image` appears complete and consistent with the caption, use it as the paneling source and set `selected_source_for_paneling` to that path.
4. If `source_image` appears incomplete, wrongly cropped, missing panel labels, merged with unrelated content, or inconsistent with the caption, do not rely on it and do not force-crop it.
5. If `source_pdf` and 1-based `source_page` are available for a suspect entry, render the corresponding original PDF page using PyMuPDF.
6. Save rendered pages under `pdf_page_renders/`.
7. Add the rendered page path as `fallback_render`.
8. Set `selected_source_for_paneling = fallback_render`.
9. Set `source_quality = "pdf_page_render_fallback"`.
10. Decide whether panel cropping is needed from `selected_source_for_paneling`.
11. Use OpenCV/PIL helper code or write a small deterministic script only if crop boundaries are clear.
12. Save panel images under `separated_panels_gemini/`.
13. Add panel paths back into `total_figure_mapping.json`.
14. If panel boundaries remain uncertain, set `manual_required: true`, `confidence: "low"`, and a short `reason`; do not hallucinate panel crops.
15. If PyMuPDF is unavailable, record a dependency note in `reason` and keep `manual_required: true`.

## PyMuPDF Render Example
```python
import fitz
from pathlib import Path

pdf_path = Path("source.pdf")
source_page = 1
out_path = Path("pdf_page_renders/source_page_001.png")

doc = fitz.open(str(pdf_path))
page = doc[source_page - 1]
pix = page.get_pixmap(dpi=220)
pix.save(str(out_path))
doc.close()
```

## Validation Command
```bash
python Agent_Task_Runner.py validate --stage 04_figure_separate --paper-folder "{paper_folder}"
```

## Constraints
- Do not run `0_mark_down_gen/04_figure_saperate_gemini.py`.
- Do not import or require `find_api.py`.
- Do not use Gemini, Vertex, `LLM_API.py`, or `LLM_Batch.py`.
- Do not hard-code API keys or credentials.
"""
    elif stage == "04_ft_excel_matcher":
        require_existing_file(paper_folder / MANUAL_MARKER, stage)
        require_existing_file(paper_folder / "fig_table_lnpdb_classified.csv", stage)
        require_existing_file(paper_folder / "fig_table_inventory.csv", stage)
        require_existing_file(paper_folder / "excel_block_inventory.csv", stage)
        content = f"""# External Agent Task: 04_ft_excel_matcher

Target paper folder: `{paper_folder}`

## Stage Purpose
Match selected figure/table items to Excel blocks by direct CLI agent judgment without Gemini.

## Required Input Files
- `{paper_folder / MANUAL_MARKER}`
- `{paper_folder / "fig_table_lnpdb_classified.csv"}`
- `{paper_folder / "fig_table_inventory.csv"}`
- `{paper_folder / "excel_block_inventory.csv"}`

## Expected Output Files
- `{paper_folder / "excel_mapping.json"}`
- `{paper_folder / "excel_mapping_rows.csv"}`
- updated `{paper_folder / "fig_table_lnpdb_classified.csv"}` when possible

## Work Instructions
1. Read `fig_table_lnpdb_classified.csv`.
2. Read `fig_table_inventory.csv`.
3. Read `excel_block_inventory.csv`.
4. Match every selected FT item to candidate Excel blocks using caption, `item_id`, `base_id`, sheet name, block preview, `block_type`, and keywords.
5. Create `excel_mapping.json`.
6. Create `excel_mapping_rows.csv`.
7. When possible, update `fig_table_lnpdb_classified.csv` columns:
   - `excel_item_id`
   - `matched_blocks`
   - `matched_block_csv_path`
   - `matched_sheet`
   - `matched_sheet_file`
8. Follow `agent_workspace/OUTPUT_SCHEMA.md` for `excel_mapping.json` and `excel_mapping_rows.csv`.

## Validation Command
```bash
python Agent_Task_Runner.py validate --stage 04_ft_excel_matcher --paper-folder "{paper_folder}"
```

## Constraints
- Do not run `0_mark_down_gen/04_FT-Excel_matcher.py`.
- Do not import or require `find_api.py`.
- Do not use Gemini, Vertex, `LLM_API.py`, or `LLM_Batch.py`.
- Do not hard-code API keys or credentials.
"""
    elif stage == "05_smiles_structure_resolution":
        require_existing_file(paper_folder / MANUAL_MARKER, stage)
        require_existing_file(paper_folder / "total_figure_mapping.json", stage)
        markdowns = asset_list(paper_folder, {".md"})
        pdfs = asset_list(paper_folder, {".pdf"})
        images = asset_list(paper_folder, {".png", ".jpg", ".jpeg", ".tif", ".tiff", ".bmp"})
        content = f"""# External Agent Task: 05_smiles_structure_resolution

Target paper folder: `{paper_folder}`

## Stage Purpose
Resolve compound names and text/reference/manual-curated SMILES without Gemini/API dependencies. Molecule-structure-image-based SMILES extraction is disabled in the active workflow.

## Required Input Files
- `{paper_folder / MANUAL_MARKER}`
- markdown files and/or PDFs from the paper folder
- `{paper_folder / "total_figure_mapping.json"}` when available
- source images when available only for figure provenance; do not use them for SMILES

Markdown files:
{render_bullet_list(markdowns)}

PDFs:
{render_bullet_list(pdfs)}

Images available for provenance only, not SMILES extraction:
{render_bullet_list(images)}

## Optional Input Files
- existing LNPDB reference file if configured locally
- manually curated or manually verified SMILES files, if explicitly present
- local text/IUPAC extraction outputs that do not rely on structure images

## Expected Output Files
- `{paper_folder / "compound_inventory_standardized.csv"}`
- `{paper_folder / "smiles_resolved.csv"}`
- `{paper_folder / "smiles_resolution_qc.csv"}`

## Work Instructions
1. Collect lipid/component names, aliases, and IUPAC names from markdown, PDFs, captions, text tables, and curated/reference files.
2. Resolve SMILES only from allowed sources: exact LNPDB/reference name or alias match, curated/local known mapping, text/name/IUPAC-based deterministic lookup that does not rely on structure images, or manually curated/manual-verified SMILES files.
3. Create `compound_inventory_standardized.csv` with one row per compound/name candidate.
4. Create `smiles_resolved.csv` with at least `Name` or `compound_id`, and `SMILES` or `resolved_smiles`.
5. Create `smiles_resolution_qc.csv` with unresolved names, conflicts, ambiguous matches, and evidence notes.
6. Preserve provenance fields such as source file, item id, caption snippet, text table block, or curated reference path when available.
7. If a SMILES cannot be resolved from allowed text/reference/manual-curated sources, leave it blank and mark it for manual review.
8. Novel pILs such as `G0-SS-AA-C12`, `G0-6C-AA-C12`, `P2A-SS-AA-C10`, etc. must remain blank in SMILES unless an exact text/reference/manual-curated SMILES entry is present. Use reason: `Structure-image-based SMILES extraction is disabled; no exact text/reference SMILES was available.`

## Validation Command
```bash
python Agent_Task_Runner.py validate --stage 05_smiles_structure_resolution --paper-folder "{paper_folder}"
```

## Constraints
- Do not run Gemini/API-assisted SMILES scripts.
- Do not import or require `find_api.py`, `LLM_API.py`, or `LLM_Batch.py`.
- Do not use Gemini, Vertex, or hard-coded credentials.
- Do not run, import, or use DECIMER, MolScribe, `worker_mol.py`, structure-recognition `pipeline.py`, `recognition.py`, or `segmentation.py`.
- Do not scan figure images for chemical structures, crop molecular structures from figures, infer SMILES from PDF/image crops, or hallucinate SMILES from visible structures.
- Do not use image-derived SMILES helper outputs unless a row is explicitly marked `human_curated` or `manual_verified`.
"""
    elif stage == "06_unified_lnpdb_extraction":
        require_existing_file(paper_folder / MANUAL_MARKER, stage)
        require_existing_file(paper_folder / "fig_table_lnpdb_classified.csv", stage)
        require_existing_file(paper_folder / "total_figure_mapping.json", stage)
        require_existing_file(paper_folder / "excel_mapping.json", stage)
        require_existing_file(paper_folder / "excel_block_inventory.csv", stage)
        if not (paper_folder / "Exp_Excel_Blocks").is_dir():
            raise FileNotFoundError(f"{stage} requires a directory: {paper_folder / 'Exp_Excel_Blocks'}")
        markdowns = asset_list(paper_folder, {".md"})
        optional_inputs = [
            "separated_panels_gemini",
            "compound_inventory_standardized.csv",
            "text_extracted_iupac.csv",
            "smiles_resolved.csv",
        ]
        reference_context = collect_reference_context_for_06(paper_folder)
        reference_debug_path = reference_context_debug_path(stage, paper_folder)
        write_json(reference_debug_path, reference_context)
        rendered_reference_context = render_reference_context_for_prompt(reference_context)
        condition_column_guide = """- `Aqueous_buffer`: short buffer label only, e.g. citrate buffer, acetate buffer, PBS, HEPES, N/A. Optional pH/concentration is allowed as a compact scalar, e.g. `10 mM citrate buffer pH 3`. Do not include prose such as `for mRNA before mixing`.
- `Dialysis_buffer`: short buffer label only, e.g. PBS, HEPES, water, N/A. Optional compact condition is allowed. Do not include full method details such as `MWCO 20 kDa, 2 h`; put those details in `evidence_text`.
- `Mixing_method`: short method label only, e.g. pipette, handmixed, microfluidics, T-junction, vortexing, liquid_handler. Do not combine alternatives with semicolons. If different panels use different mixing methods, split rows or leave blank with `manual_required=true`.
- `Model`: must be one of `in_vitro`, `in_vivo`, `ex_vivo`, or `N/A`.
- `Model_type`: if `Model=in_vitro`, cell line only, e.g. MC38, HEK293T, A549. If `Model=in_vivo`, animal/model only, e.g. C57BL/6_mouse, BALB/c_mouse, Mouse_MC38_tumor. Do not put full phrases such as `MC38 cells or MC38 tumour-bearing mice`.
- `Model_target`: if `Model=in_vitro`, use `in_vitro` or a measured cell/target. If `Model=in_vivo`, use organ/tissue/tumor target only, e.g. tumor, spleen, liver, lung. Do not combine multiple panels into one target with slash or prose.
- `Route_of_administration`: if `Model=in_vitro`, use `in_vitro`. If `Model=in_vivo`, use intravenous, intratumoral, intramuscular, subcutaneous, inhalation, oral, etc. Do not write sentences.
- `Cargo`: nucleic acid class only, e.g. mRNA, siRNA, pDNA, sgRNA, saRNA. Do not put reporter or encoded protein here.
- `Cargo_type`: encoded/reporter payload only, e.g. FLuc, NLuc, IL-12, OVA, Cre, Cas9. Do not put `mRNA` here if `Cargo` already contains mRNA.
- `Dose_ug_nucleicacid`: numeric microgram dose only, e.g. 0.05, 0.5, 2.5. No sentence. No `per mouse` prose. Convert straightforward ng values to ug, e.g. 50 ng = 0.05. If unit/context cannot be normalized, leave blank and set `manual_required=true`.
- `Experiment_method`: concise method/readout label only, e.g. luminescence, IVIS, ELISA_IL-12, flow_cytometry_CD8_T_cells, qPCR_IFN-gamma, RNA-seq, western_blot. Follow existing LNPDB column-specific examples when available. Do not combine multiple panels as `flow cytometry/ELISA/RNA-seq`. If a figure group contains multiple panels with different methods or readouts, split rows by panel/block.
- `Experiment_batching`: must be individual, barcoded, pooled, grouped, or N/A. Prefer individual unless the paper explicitly uses barcoded/pooled screening."""
        formulation_column_guide = """- `formulation_id`: stable row-level formulation identifier from the paper or a concise derived ID.
- `Formulation_Name`: formulation label exactly enough to identify the group, normalized to a short scalar.
- `IL_name`: ionizable lipid name.
- `IL_SMILES`: always blank in the current active LNPDB-like output. Do not project SMILES from `smiles_resolved.csv`, references, curated mappings, PubChem, OPSIN, CIR, or images.
- `IL_molarratio`: ionizable lipid molar ratio or mol% as a concise scalar.
- `HL_name`: helper lipid name.
- `HL_SMILES`: always blank in the current active LNPDB-like output. Preserve `HL_name` and `HL_molarratio`.
- `HL_molarratio`: helper lipid molar ratio or mol% as a concise scalar.
- `CHL_name`: cholesterol or cholesterol-like component name.
- `CHL_SMILES`: always blank in the current active LNPDB-like output. Preserve `CHL_name` and `CHL_molarratio`.
- `CHL_molarratio`: cholesterol component molar ratio or mol% as a concise scalar.
- `PEG_name`: PEG-lipid or PEG component name.
- `PEG_SMILES`: always blank in the current active LNPDB-like output. Preserve `PEG_name` and `PEG_molarratio`.
- `PEG_molarratio`: PEG component molar ratio or mol% as a concise scalar.
- `Fifth_component_name`: additional non-core formulation component name, if any.
- `Fifth_component_SMILES`: always blank in the current active LNPDB-like output. Preserve `Fifth_component_name` and `Fifth_component_molarratio`.
- `Fifth_component_molarratio`: additional component molar ratio or mol% as a concise scalar.
- `IL_to_nucleicacid_massratio`: ionizable lipid to nucleic acid mass ratio, e.g. N/P or wt/wt, when explicitly available."""
        content = f"""# External Agent Task: 06_unified_lnpdb_extraction

Target paper folder: `{paper_folder}`

## Stage Purpose
Create one unified long table at figure/table item level that combines experimental conditions, formulation composition, and provenance. Experimental numeric assay/readout value extraction is disabled for this stage and deferred to a future value-extraction stage.

## Required Input Files
- `{paper_folder / MANUAL_MARKER}`
- `{paper_folder / "fig_table_lnpdb_classified.csv"}`
- `{paper_folder / "total_figure_mapping.json"}`
- `{paper_folder / "excel_mapping.json"}`
- `{paper_folder / "excel_block_inventory.csv"}`
- `{paper_folder / "Exp_Excel_Blocks"}`
- markdown files:
{render_bullet_list(markdowns)}

## Optional Input Files
{render_bullet_list([item for item in optional_inputs if (paper_folder / item).exists()])}
- `smiles_resolved.csv` may exist for Stage 05 QC only. Do not project any SMILES value into unified extraction outputs.

## Expected Output Files
- `{paper_folder / "unified_extraction.csv"}`
- `{paper_folder / "unified_extraction.json"}`
- `{paper_folder / "unified_extraction_review_flags.csv"}`

## Required Output Columns
Use the columns documented in `agent_workspace/OUTPUT_SCHEMA.md` for `unified_extraction.csv`. Include all experimental condition, formulation composition, experimental value, evidence, confidence, and manual review fields. Populate `metric_type`, `original_values`, `aggregated_value`, `unit`, and `replicate_type` only from reliable mapped Excel/source-data blocks. Leave them blank when no reliable Excel/source-data mapping exists.

## LNPDB Experimental-Condition Column Guide
{condition_column_guide}

## LNPDB Formulation Column Guide
{formulation_column_guide}

## Scalar Normalization Rules
- LNPDB fields must contain concise normalized scalar values.
- SMILES fields must remain blank in `unified_extraction.csv`; do not infer/generate/project SMILES from `smiles_resolved.csv`, DB/reference/curated files, PubChem, OPSIN, CIR, figure images, or molecular structure crops.
- Prefer column-specific examples extracted from the existing LNPDB reference over generic examples.
- If existing LNPDB examples show assay+readout in `Experiment_method`, preserve that style.
- `Model_target` remains tissue/organ/site where applicable, while `Experiment_method` may include the measured readout/cell population when that is the established LNPDB style.
- Example: `Model_target=spleen` and `Experiment_method=flow_cytometry_CD8_T_cells` is valid.
- Do not reduce readout-specific methods to `flow_cytometry` when panel identity depends on the measured cell population.
- Do not copy full source sentences, captions, paragraphs, or methods prose into LNPDB fields.
- Full source sentences and captions may be stored only in `evidence_text`.
- If a concise normalized value cannot be determined, leave the field blank or mark `manual_required=true` with a reason.
- Do not use `variable` or `various` as a value unless the paper explicitly uses it as a label and no better scalar value exists.

{rendered_reference_context}

## Forbidden in LNPDB Condition Fields
- Full sentences or caption fragments.
- Values containing `or` when it merges multiple experimental contexts.
- Semicolon-separated mixed contexts.
- Panel-combined values such as `in vitro treatment; intratumoural injection for panel c`.
- Multi-method bundles such as `flow cytometry/ELISA/RNA-seq`.
- Any value that belongs in `evidence_text` rather than a scalar field.

If one caption describes multiple panels with different conditions, create separate rows per panel/item/block context. Do not merge panel b and panel c conditions into one row.

## Scalar Condition Examples
Bad:
- `Model = MC38 cells or MC38 tumour-bearing mice`
- `Route_of_administration = in vitro treatment; intratumoural injection for panel c`
- `Dose_ug_nucleicacid = 50 ng/well for in vitro screening; 2.5 ug per mouse for in vivo panel c`
- `Experiment_method = luciferase bioluminescence assay / IVIS imaging`

Good for Figure 2B:
- `Model = in_vitro`
- `Model_type = MC38`
- `Route_of_administration = in_vitro`
- `Dose_ug_nucleicacid = 0.05`
- `Experiment_method = luminescence`

Good for Figure 2C:
- `Model = in_vivo`
- `Model_type = Mouse_MC38_tumor`
- `Model_target = tumor`
- `Route_of_administration = intratumoral`
- `Dose_ug_nucleicacid = 2.5`
- `Experiment_method = IVIS`

QS_2026 Figure 4G-M style guidance:
- figure 4g: `Model_target=spleen`; `Experiment_method=flow_cytometry_CD8_T_cells`
- figure 4h: `Model_target=spleen`; `Experiment_method=flow_cytometry_CD4_T_cells`
- figure 4i: `Model_target=spleen`; `Experiment_method=flow_cytometry_CD8_effector_memory_T_cells`
- figure 4j: `Model_target=spleen`; `Experiment_method=flow_cytometry_CD8_central_memory_T_cells`
- figure 4k: `Model_target=spleen`; `Experiment_method=flow_cytometry_CD4_effector_memory_T_cells`
- figure 4l: `Model_target=spleen`; `Experiment_method=flow_cytometry_CD4_central_memory_T_cells`
- figure 4m: `Model_target=spleen`; `Experiment_method=flow_cytometry_regulatory_T_cells`

## Work Instructions
1. For every selected figure/table item, extract experimental conditions and formulation composition together into one unified long table.
2. Do not split the task into separate independent LLM calls for conditions and formulation.
3. Use all available condition/formulation context: markdown captions, PDF-derived images for labels/provenance only, separated panels for labels/provenance only, Excel block CSVs, `excel_mapping.json`, and `total_figure_mapping.json`. `smiles_resolved.csv` is Stage 05 QC context only and must not populate output SMILES fields.
4. Use Excel blocks only for sheet/block identity, labels, headers, formulation names, group labels, condition context, and provenance.
5. Extract experimental numeric assay/readout values only from mapped Excel/source-data blocks (`Exp_Excel_Blocks/`, source-data Excel files, `excel_mapping.json`, `excel_mapping_rows.csv`, and referenced `block_csv_path` files).
6. Populate `metric_type`, `original_values`, `aggregated_value`, `unit`, and `replicate_type` only when the value is tied to Excel/source-data provenance. Do not use captions alone for values.
7. Use figure/PDF images for labels, axes, legend, group interpretation, panel identity, and visual context.
8. Use markdown for caption, methods context, dose, model, route, and formulation descriptions.
9. If condition or formulation values are uncertain, do not hallucinate. Leave blank or set `manual_required=true` with a reason.
10. If multiple formulations, panels, methods, models, routes, or dose contexts exist in one figure/table, produce separate rows per formulation/condition/panel context.
11. Use long format.
12. Record `evidence_text`, `evidence_excel`, and `evidence_image` for every nontrivial extracted condition or formulation value.
13. Preserve provenance fields: `evidence_text`, `evidence_excel`, `evidence_image`, `confidence`, `manual_required`, and `reason`.
14. Create `unified_extraction_review_flags.csv` for missing metadata, low confidence, condition/formulation mismatch, missing figure evidence, and any manual review need. Blank output SMILES fields are expected and are not a manual-review issue by themselves.
15. Also write `unified_extraction.json` with records and source summary.
16. Always leave `IL_SMILES`, `HL_SMILES`, `CHL_SMILES`, `PEG_SMILES`, and `Fifth_component_SMILES` blank in Stage 06 outputs. Do not attempt image-based fallback, crop molecular structures, infer visible structures, use DB/reference/curated/PubChem/OPSIN/CIR sources, or generate/project any SMILES in 06. Do not mark a row manual-required solely because these SMILES fields are blank.
17. When Excel has replicate columns, keep exact replicate/source values pipe-separated in `original_values`; set `aggregated_value` from an explicit mean/value column when present, or compute an arithmetic mean only when replicates are unambiguous and note that in `reason`.

## Validation Command
```bash
python Agent_Task_Runner.py validate --stage 06_unified_lnpdb_extraction --paper-folder "{paper_folder}"
```

## Constraints
- Do not use Gemini/API/find_api/LLM_API/LLM_Batch.
- Do not run legacy scripts from `1_Extract_Exp_Figs`, `3_Extract_Formula_by_Figs`, or `4_Extract_Exp_Vals`.
- Do not run or use DECIMER, MolScribe, `worker_mol.py`, structure-recognition `pipeline.py`, `recognition.py`, `segmentation.py`, molecule image crops, or image-derived SMILES outputs.
- Do not use graph image digitization, pixel/axis extraction, bar-height estimation, heatmap color estimation, or visual numeric estimation from figure images.
- Do not hard-code API keys or credentials.
"""
    elif stage == "07_finalize_unified_table":
        require_existing_file(paper_folder / MANUAL_MARKER, stage)
        require_existing_file(paper_folder / "unified_extraction.csv", stage)
        content = f"""# External Agent Task: 07_finalize_unified_table

Target paper folder: `{paper_folder}`

## Stage Purpose
Finalize `unified_extraction.csv` into reviewed final and LNPDB-like tables, with figure/item-level source evidence tables and a QC report.

## Required Input Files
- `{paper_folder / MANUAL_MARKER}`
- `{paper_folder / "unified_extraction.csv"}`
- `{paper_folder / "unified_extraction_review_flags.csv"}`

## Expected Output Files
- `{paper_folder / "markdown_sentence_index" / "markdown_sentence_index_all.csv"}`
- `{paper_folder / "markdown_sentence_index" / "markdown_sentence_index_manifest.json"}`
- `{paper_folder / "unified_extraction_final.csv"}`
- `{paper_folder / "unified_extraction_lnpdb_like.csv"}`
- `{paper_folder / "unified_extraction_source_evidence.csv"}`
- `{paper_folder / "unified_extraction_figure_evidence_map.csv"}`
- `{paper_folder / "unified_extraction_qc_report.json"}`

## Work Instructions
1. Read `unified_extraction.csv`.
2. Preserve source rows and provenance.
3. Normalize booleans and blank values without inventing missing scientific data.
4. Add deterministic stable `row_id` values in the form `<Paper_ID>_R000001`.
5. Build `markdown_sentence_index/` from source markdown files, excluding markdown table regions and generated/output folders.
6. Create numbered sentence IDs like `QS_2026:S000145` and use them as the primary source-text evidence anchors.
7. Create `unified_extraction_final.csv` preserving all source/provenance columns.
8. Create `unified_extraction_lnpdb_like.csv` as the LNPDB-like value table with `row_id`; do not repeat large evidence text fields there.
9. Create `unified_extraction_source_evidence.csv` with one row per unique evidence phrase/source object, deterministic `evidence_id`, compact `evidence_summary`, and `evidence_sentence_ids` when indexed markdown support is found.
10. Create `unified_extraction_figure_evidence_map.csv` linking each evidence sentence/source object to pipe-separated LNPDB scientific condition/formulation columns it supports for each figure/item.
11. Do not create noisy per-cell evidence requirements for administrative/provenance columns.
12. Create `unified_extraction_qc_report.json` with value/evidence/map/sentence-index counts and warnings.

## Validation Command
```bash
python Agent_Task_Runner.py validate --stage 07_finalize_unified_table --paper-folder "{paper_folder}"
```

## Constraints
- Do not use Gemini/API/find_api/LLM_API/LLM_Batch.
- Do not hallucinate missing values during finalization.
"""
    else:
        raise ValueError(f"No external agent task template for stage: {stage}")

    content = content.rstrip() + render_legacy_context_section(stage)
    path = write_task_file(stage, paper_folder, content)
    return {
        "status": "external_agent_required",
        "stage": stage,
        "task_file": str(path),
        "message": "Legacy Gemini script was not executed. Ask Codex/Claude CLI agent to complete this task file.",
    }


def write_csv_dicts(path: Path, rows: list[dict[str, Any]], fieldnames: list[str]) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    with path.open("w", encoding="utf-8-sig", newline="") as f:
        writer = csv.DictWriter(f, fieldnames=fieldnames, extrasaction="ignore")
        writer.writeheader()
        for row in rows:
            writer.writerow({key: row.get(key, "") for key in fieldnames})


def write_csv_matrix(path: Path, rows: list[list[Any]]) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    with path.open("w", encoding="utf-8-sig", newline="") as f:
        writer = csv.writer(f)
        writer.writerows(rows)


def is_blank_cell(value: Any) -> bool:
    return value is None or str(value).strip() == ""


def read_csv_matrix(path: Path) -> list[list[Any]]:
    for encoding in ("utf-8-sig", "cp949", "latin-1"):
        try:
            with path.open("r", encoding=encoding, newline="") as f:
                return [row for row in csv.reader(f)]
        except UnicodeDecodeError:
            continue
    with path.open("r", encoding="utf-8-sig", errors="replace", newline="") as f:
        return [row for row in csv.reader(f)]


def contiguous_groups(indices: list[int]) -> list[list[int]]:
    if not indices:
        return []
    groups = [[indices[0]]]
    for index in indices[1:]:
        if index == groups[-1][-1] + 1:
            groups[-1].append(index)
        else:
            groups.append([index])
    return groups


def trim_matrix(matrix: list[list[Any]]) -> list[list[Any]]:
    while matrix and all(is_blank_cell(cell) for cell in matrix[-1]):
        matrix.pop()
    if not matrix:
        return []
    max_width = max(len(row) for row in matrix)
    padded = [row + [""] * (max_width - len(row)) for row in matrix]
    while padded and padded[0] and all(is_blank_cell(row[-1]) for row in padded):
        padded = [row[:-1] for row in padded]
    return padded


def split_matrix_blocks(matrix: list[list[Any]]) -> list[dict[str, Any]]:
    matrix = trim_matrix([list(row) for row in matrix])
    if not matrix:
        return []
    max_width = max(len(row) for row in matrix)
    padded = [row + [""] * (max_width - len(row)) for row in matrix]
    non_empty_rows = [i for i, row in enumerate(padded) if any(not is_blank_cell(cell) for cell in row)]
    blocks: list[dict[str, Any]] = []
    for row_group in contiguous_groups(non_empty_rows):
        row_slice = padded[row_group[0] : row_group[-1] + 1]
        non_empty_cols = [
            j for j in range(max_width) if any(not is_blank_cell(row[j]) for row in row_slice)
        ]
        for col_group in contiguous_groups(non_empty_cols):
            cells = [row[col_group[0] : col_group[-1] + 1] for row in row_slice]
            cells = trim_matrix(cells)
            if cells:
                blocks.append(
                    {
                        "row_start": row_group[0] + 1,
                        "row_end": row_group[-1] + 1,
                        "col_start": col_group[0] + 1,
                        "col_end": col_group[-1] + 1,
                        "cells": cells,
                    }
                )
    return blocks


def numeric_cell(value: Any) -> bool:
    text = str(value).strip().replace(",", "").replace("%", "")
    if not text:
        return False
    try:
        float(text)
        return True
    except ValueError:
        return False


def infer_block_type(cells: list[list[Any]]) -> str:
    non_empty = [cell for row in cells for cell in row if not is_blank_cell(cell)]
    numeric_count = sum(1 for cell in non_empty if numeric_cell(cell))
    text_count = len(non_empty) - numeric_count
    row_count = len(cells)
    col_count = max((len(row) for row in cells), default=0)
    first_row_text = sum(1 for cell in (cells[0] if cells else []) if not is_blank_cell(cell) and not numeric_cell(cell))
    if row_count <= 2 and numeric_count == 0:
        return "table_title" if text_count <= 6 else "note"
    if row_count >= 12 and col_count >= 8 and numeric_count >= 10:
        return "multi_table"
    if numeric_count >= 3 and first_row_text > 0:
        return "title_and_table"
    if numeric_count >= max(3, text_count) and row_count >= 2:
        return "table_body"
    if row_count <= 3 and text_count > numeric_count:
        return "note"
    return "other"


def load_excel_like_sheets(path: Path) -> list[tuple[str, list[list[Any]], str | None]]:
    suffix = path.suffix.lower()
    if suffix == ".csv":
        return [("csv", read_csv_matrix(path), None)]
    if suffix == ".xlsx":
        try:
            from openpyxl import load_workbook
        except ImportError as exc:
            return [("workbook", [], f"openpyxl is required for {path.name}: {exc}")]
        workbook = load_workbook(path, read_only=False, data_only=True)
        sheets = []
        for sheet in workbook.worksheets:
            matrix = [
                [sheet.cell(row=i, column=j).value for j in range(1, sheet.max_column + 1)]
                for i in range(1, sheet.max_row + 1)
            ]
            sheets.append((sheet.title, matrix, None))
        return sheets
    if suffix == ".xls":
        try:
            import pandas as pd
        except ImportError as exc:
            return [("workbook", [], f"pandas/xlrd is required for {path.name}: {exc}")]
        try:
            sheet_map = pd.read_excel(path, sheet_name=None, header=None)
        except Exception as exc:
            return [("workbook", [], f"failed to read {path.name}: {exc}")]
        return [(name, frame.fillna("").values.tolist(), None) for name, frame in sheet_map.items()]
    return []


def run_heuristic_figure_mapping(paper_folder: Path) -> dict[str, Any]:
    classified = paper_folder / "fig_table_lnpdb_classified.csv"
    require_existing_file(classified, "03_figure_mapping")
    rows = read_csv_rows(classified)
    selected = selected_ft_rows(rows)
    image_files = [
        p for p in iter_paper_files(paper_folder)
        if p.suffix.lower() in {".png", ".jpg", ".jpeg", ".tif", ".tiff", ".bmp"}
    ]
    table_files = [
        p for p in iter_paper_files(paper_folder)
        if p.suffix.lower() in {".csv", ".xlsx", ".xls"}
        and p.name not in {
            "fig_table_inventory.csv",
            "fig_table_lnpdb_classified.csv",
            "excel_block_inventory.csv",
            "excel_mapping_rows.csv",
        }
    ]
    mapping: dict[str, Any] = {
        paper_folder.name: {
            "_metadata": {
                "created_by": "Agent_Task_Runner heuristic mode",
                "accuracy_note": "Low-accuracy temporary mapping based on filename substring matching only.",
                "created_at": utc_now(),
            }
        }
    }
    for row in selected:
        item_id = row_item_id(row)
        base_id = (row.get("base_id") or "").strip()
        tokens = [normalize_token(v) for v in (item_id, base_id) if normalize_token(v)]

        def score(path: Path) -> int:
            name = normalize_token(path.stem)
            return max((len(token) for token in tokens if token and (token in name or name in token)), default=0)

        image_match = max(image_files, key=score, default=None)
        table_match = max(table_files, key=score, default=None)
        image_score = score(image_match) if image_match else 0
        table_score = score(table_match) if table_match else 0
        matched = image_score > 0 or table_score > 0
        image_found = bool(image_match and image_score > 0)
        mapping[paper_folder.name][item_id] = {
            "item_id": item_id,
            "base_id": base_id,
            "caption": row.get("caption", ""),
            "source_image": rel_to_paper(image_match, paper_folder) if image_found else None,
            "source_table": rel_to_paper(table_match, paper_folder) if table_match and table_score > 0 else None,
            "source_quality": "ok" if image_found else "missing_image",
            "manual_required": not image_found,
            "confidence": "low" if matched else "unmatched",
            "reason": "Filename matched item_id/base_id in heuristic mode." if matched else "No filename matched item_id/base_id in heuristic mode.",
        }
    write_json(paper_folder / "total_figure_mapping.json", mapping)
    return {"selected_items": len(selected), "output": str(paper_folder / "total_figure_mapping.json")}


def run_heuristic_split_excel_blocks(paper_folder: Path) -> dict[str, Any]:
    exp_excel = paper_folder / "Exp_Excel"
    block_root = paper_folder / "Exp_Excel_Blocks"
    block_root.mkdir(parents=True, exist_ok=True)
    excel_files = []
    if exp_excel.exists():
        excel_files = [
            p for p in exp_excel.rglob("*")
            if p.is_file() and p.suffix.lower() in {".xlsx", ".xls", ".csv"}
        ]
    inventory_rows: list[dict[str, Any]] = []
    summary: list[dict[str, Any]] = []
    fieldnames = [
        "excel_file",
        "excel_sheet",
        "block_id",
        "group_id",
        "element_id",
        "block_csv_path",
        "block_meta_path",
        "block_type",
    ]
    for excel_file in excel_files:
        sheet_results = []
        for sheet_name, matrix, error in load_excel_like_sheets(excel_file):
            if error:
                sheet_results.append({"excel_sheet": sheet_name, "error": error, "blocks": 0})
                continue
            blocks = split_matrix_blocks(matrix)
            sheet_results.append({"excel_sheet": sheet_name, "blocks": len(blocks), "method": "blank_row_column_components"})
            for index, block in enumerate(blocks, start=1):
                block_id = f"block_{len(inventory_rows) + 1:04d}"
                rel_dir = Path("Exp_Excel_Blocks") / safe_name(excel_file.stem) / safe_name(sheet_name)
                block_csv = paper_folder / rel_dir / f"{block_id}.csv"
                block_meta = paper_folder / rel_dir / f"{block_id}.json"
                block_type = infer_block_type(block["cells"])
                write_csv_matrix(block_csv, block["cells"])
                write_json(
                    block_meta,
                    {
                        "excel_file": rel_to_paper(excel_file, paper_folder),
                        "excel_sheet": sheet_name,
                        "block_id": block_id,
                        "bounds": {
                            "row_start": block["row_start"],
                            "row_end": block["row_end"],
                            "col_start": block["col_start"],
                            "col_end": block["col_end"],
                        },
                        "block_type": block_type,
                        "created_by": "Agent_Task_Runner heuristic mode",
                    },
                )
                inventory_rows.append(
                    {
                        "excel_file": rel_to_paper(excel_file, paper_folder),
                        "excel_sheet": sheet_name,
                        "block_id": block_id,
                        "group_id": f"{safe_name(excel_file.stem)}_{safe_name(sheet_name)}",
                        "element_id": f"{safe_name(excel_file.stem)}_{safe_name(sheet_name)}_{index:03d}",
                        "block_csv_path": rel_to_paper(block_csv, paper_folder),
                        "block_meta_path": rel_to_paper(block_meta, paper_folder),
                        "block_type": block_type,
                    }
                )
        summary.append(
            {
                "excel_file": rel_to_paper(excel_file, paper_folder),
                "result": sheet_results,
                "created_by": "Agent_Task_Runner heuristic mode",
            }
        )
    write_csv_dicts(paper_folder / "excel_block_inventory.csv", inventory_rows, fieldnames)
    write_json(paper_folder / "three_core_result_all.json", summary)
    return {"excel_files": len(excel_files), "blocks": len(inventory_rows)}


def run_heuristic_figure_separate(paper_folder: Path) -> dict[str, Any]:
    mapping_path = paper_folder / "total_figure_mapping.json"
    require_existing_file(mapping_path, "04_figure_separate")
    data = load_json(mapping_path, {})
    panel_dir = paper_folder / "separated_panels_gemini"
    panel_dir.mkdir(parents=True, exist_ok=True)
    updated = 0
    fallback_render_count = 0
    fallback_render_errors: list[str] = []
    for entry in iter_total_figure_mapping_entries(data):
        source_quality = str(entry.get("source_quality") or "").strip()
        if source_quality in FALLBACK_SOURCE_QUALITIES and entry.get("source_pdf") and entry.get("source_page"):
            try:
                source_page = int(entry["source_page"])
                fallback_render = render_pdf_page(paper_folder, str(entry["source_pdf"]), source_page)
            except Exception as exc:
                message = f"PDF page render fallback failed: {exc}"
                entry["manual_required"] = True
                entry["confidence"] = "low"
                entry["reason"] = f"{entry.get('reason', '').strip()} {message}".strip()
                fallback_render_errors.append(message)
            else:
                entry["fallback_render"] = fallback_render
                entry["selected_source_for_paneling"] = fallback_render
                entry["source_quality"] = "pdf_page_render_fallback"
                entry["manual_required"] = True
                entry["confidence"] = "low"
                entry["reason"] = (
                    f"{entry.get('reason', '').strip()} "
                    "Heuristic mode rendered the source PDF page because the Marker image was not reliable; "
                    "panel separation still requires manual review."
                ).strip()
                fallback_render_count += 1
        if entry.get("source_image") or entry.get("fallback_render"):
            entry.setdefault("panels", {})
            entry["panel_separation"] = "not_performed"
            if not entry.get("fallback_render"):
                entry["confidence"] = "not_separated"
                entry["reason"] = "Panel separation not performed in heuristic mode."
            updated += 1
    write_json(
        panel_dir / "manifest.json",
        {
            "created_by": "Agent_Task_Runner heuristic mode",
            "status": "panel separation not performed in heuristic mode",
            "updated_mapping_entries": updated,
            "fallback_render_count": fallback_render_count,
            "fallback_render_errors": fallback_render_errors,
            "created_at": utc_now(),
        },
    )
    write_json(mapping_path, data)
    return {
        "panel_dir": str(panel_dir),
        "updated_mapping_entries": updated,
        "fallback_render_count": fallback_render_count,
        "fallback_render_errors": fallback_render_errors,
    }


def keyword_set(*values: str) -> set[str]:
    stopwords = {"figure", "fig", "table", "the", "and", "with", "from", "that", "this", "for", "lnpdb"}
    words: set[str] = set()
    for value in values:
        for word in re.findall(r"[A-Za-z0-9]+", value.lower()):
            if len(word) >= 3 and word not in stopwords:
                words.add(word)
    return words


def block_preview_text(paper_folder: Path, block_csv_path: str, max_rows: int = 20) -> str:
    if not block_csv_path:
        return ""
    path = paper_folder / block_csv_path
    if not path.exists():
        return ""
    rows = read_csv_matrix(path)[:max_rows]
    return " ".join(str(cell) for row in rows for cell in row if str(cell).strip())


def selected_item_rows_by_id(paper_folder: Path) -> dict[str, dict[str, str]]:
    classified = paper_folder / "fig_table_lnpdb_classified.csv"
    if not classified.exists():
        return {}
    rows = selected_ft_rows(read_csv_rows(classified))
    return {row_item_id(row): row for row in rows}


def mapping_entries_by_item(paper_folder: Path) -> dict[str, dict[str, Any]]:
    data = load_json(paper_folder / "total_figure_mapping.json", {})
    entries: dict[str, dict[str, Any]] = {}
    for entry in iter_total_figure_mapping_entries(data):
        item_id = str(entry.get("item_id") or entry.get("pdf_item_id") or entry.get("Item_ID") or "").strip()
        if item_id:
            entries[item_id] = entry
    return entries


def excel_matches_by_item(paper_folder: Path) -> dict[str, list[dict[str, Any]]]:
    data = load_json(paper_folder / "excel_mapping.json", {})
    matches: dict[str, list[dict[str, Any]]] = {}
    if isinstance(data, dict):
        for item_id, value in data.items():
            if isinstance(value, list):
                matches[str(item_id)] = [entry for entry in value if isinstance(entry, dict)]
            elif isinstance(value, dict):
                matches[str(item_id)] = [value]
    return matches


def normalize_bool_text(value: Any) -> str:
    if isinstance(value, bool):
        return "true" if value else "false"
    text = str(value or "").strip().lower()
    if text in {"1", "true", "yes", "y"}:
        return "true"
    if text in {"0", "false", "no", "n"}:
        return "false"
    return "true" if text else ""


def row_text_for_smiles_source(row: dict[str, Any]) -> str:
    return " ".join(
        str(row.get(col, ""))
        for col in (
            "source_type",
            "source_path",
            "source_image",
            "resolution_method",
            "method",
            "provenance",
            "evidence_text",
            "reason",
            "notes",
        )
    ).casefold()


def is_manual_verified_smiles_row(row: dict[str, Any]) -> bool:
    text = row_text_for_smiles_source(row)
    return any(term in text for term in MANUAL_VERIFIED_SMILES_TERMS)


def is_disallowed_image_smiles_row(row: dict[str, Any]) -> bool:
    text = row_text_for_smiles_source(row)
    return any(term in text for term in DISALLOWED_IMAGE_SMILES_TERMS) and not is_manual_verified_smiles_row(row)


def row_resolved_smiles_value(row: dict[str, Any]) -> str:
    return str(row.get("resolved_smiles") or row.get("Resolved_SMILES") or row.get("SMILES") or row.get("smiles") or "").strip()


def numeric_string(value: Any) -> str:
    text = str(value or "").strip()
    if not text:
        return ""
    normalized = text.replace(",", "")
    if normalized.endswith("%"):
        normalized = normalized[:-1].strip()
    if re.fullmatch(r"[-+]?(?:\d+(?:\.\d*)?|\.\d+)(?:[eE][-+]?\d+)?", normalized):
        return normalized
    return ""


def cell_has_numeric_value(value: Any) -> bool:
    text = str(value or "").strip()
    return bool(text and re.search(r"[-+]?\d", text))


def excel_context_labels(matrix: list[list[Any]], limit: int = 12) -> list[str]:
    labels: list[str] = []
    seen: set[str] = set()
    for matrix_row in matrix:
        for cell in matrix_row:
            text = str(cell or "").strip()
            if not text:
                continue
            normalized = text.replace(",", "").replace("%", "").strip()
            if re.fullmatch(r"[-+]?(?:\d+(?:\.\d*)?|\.\d+)(?:[eE][-+]?\d+)?", normalized):
                continue
            compact = re.sub(r"\s+", " ", text)
            key = compact.lower()
            if key in seen:
                continue
            labels.append(compact)
            seen.add(key)
            if len(labels) >= limit:
                return labels
    return labels


def nearest_row_label(row: list[Any], col_index: int) -> str:
    for idx in range(min(col_index, len(row)) - 1, -1, -1):
        text = str(row[idx] or "").strip()
        if text and not cell_has_numeric_value(text):
            return text
    return ""


def nearest_column_label(matrix: list[list[Any]], row_index: int, col_index: int) -> str:
    for idx in range(row_index - 1, -1, -1):
        if col_index < len(matrix[idx]):
            text = str(matrix[idx][col_index] or "").strip()
            if text and not cell_has_numeric_value(text):
                return text
    return ""


def base_unified_row(paper_folder: Path, item_id: str, item_row: dict[str, str], mapping_entry: dict[str, Any] | None) -> dict[str, Any]:
    entry = mapping_entry or {}
    row = {column: "" for column in UNIFIED_EXTRACTION_COLUMNS}
    row.update(
        {
            "Paper_ID": paper_folder.name,
            "Item_ID": item_id,
            "visual_type": item_row.get("type") or item_row.get("visual_type") or item_row.get("label") or "",
            "source_image": entry.get("source_image", ""),
            "source_pdf": entry.get("source_pdf", ""),
            "source_page": entry.get("source_page", ""),
            "selected_source_for_paneling": entry.get("selected_source_for_paneling", ""),
            "evidence_text": item_row.get("caption", ""),
            "evidence_image": entry.get("selected_source_for_paneling") or entry.get("fallback_render") or entry.get("source_image", ""),
            "confidence": "low",
            "manual_required": "true",
            "reason": "Heuristic unified extraction; complex metadata requires manual or external-agent review.",
        }
    )
    return row


def normalize_fallback_group_label(label: str) -> str:
    label = re.sub(r"\bGO-SS-AA", "G0-SS-AA", str(label or ""), flags=re.I)
    label = re.sub(r"\bGO-6C-AA", "G0-6C-AA", label, flags=re.I)
    label = re.sub(r"\bFluc\b", "FLuc", label, flags=re.I)
    return re.sub(r"\s+", " ", label).strip(" ;,")


def fallback_group_labels_from_text(text: str) -> list[str]:
    patterns = [
        r"\bPBS\b",
        r"\bMC3(?:\s+(?:FLuc|IL-12))?\s+LNP\b",
        r"\bG0-SS-AA-C12(?:\s+(?:FLuc|IL-12))?\s+pLNP\b",
        r"\bG0-6C-AA-C12(?:\s+(?:FLuc|IL-12))?\s+LNP(?:\s+plus\s+free\s+IDO inhibitor)?\b",
        r"\bfree\s+IL-12\s+mRNA\b",
        r"\bfree\s+IDO inhibitor\b",
        r"\buntreated cells\b",
        r"\bcontrol group\b",
        r"\btreated group\b",
    ]
    labels: list[str] = []
    seen: set[str] = set()
    for pattern in patterns:
        for match in re.finditer(pattern, text or "", flags=re.I):
            label = normalize_fallback_group_label(match.group(0))
            key = label.lower()
            if label and key not in seen:
                seen.add(key)
                labels.append(label)
    return labels


def apply_generic_fallback_group(row: dict[str, Any], group_label: str) -> None:
    low = group_label.lower()
    row["condition_1_name"] = "fallback_source"
    row["condition_1_value"] = "caption/image group labels"
    row["condition_2_name"] = "formulation_or_treatment_group"
    row["condition_2_value"] = group_label
    row["Formulation_Name"] = "" if any(term in low for term in ("pbs", "untreated", "control group")) else group_label
    row["formulation_id"] = re.sub(r"[^A-Za-z0-9_]+", "_", row["Formulation_Name"]).strip("_") if row["Formulation_Name"] else ""
    if "il-12" in low:
        row["Cargo"] = "mRNA"
        row["Cargo_type"] = "IL-12"
    elif "fluc" in low:
        row["Cargo"] = "mRNA"
        row["Cargo_type"] = "FLuc"
    elif any(term in low for term in ("pbs", "untreated", "control group", "free ido")):
        row["Cargo"] = ""
        row["Cargo_type"] = ""
        row["Dose_ug_nucleicacid"] = ""
    if "g0-ss-aa-c12" in low or "g0-6c-aa-c12" in low:
        row["IL_name"] = "G0-SS-AA-C12" if "g0-ss-aa-c12" in low else "G0-6C-AA-C12"
        row["IL_molarratio"] = "50"
        row["HL_name"] = "DOPE"
        row["HL_molarratio"] = "10"
        row["CHL_name"] = "cholesterol"
        row["CHL_molarratio"] = "38.5"
        row["PEG_name"] = "DMG-PEG"
        row["PEG_molarratio"] = "1.5"
    elif "mc3" in low:
        row["IL_name"] = "DLin-MC3-DMA"
        row["IL_molarratio"] = "50"
        row["HL_name"] = "DSPC"
        row["HL_molarratio"] = "10"
        row["CHL_name"] = "cholesterol"
        row["CHL_molarratio"] = "38.5"
        row["PEG_name"] = "DMG-PEG"
        row["PEG_molarratio"] = "1.5"
    for field in DEFERRED_VALUE_COLUMNS_06:
        row[field] = ""


def generic_fallback_rows_for_item(paper_folder: Path, item_id: str, item_row: dict[str, str], mapping_entry: dict[str, Any] | None) -> list[dict[str, Any]]:
    mapping_entry = mapping_entry or {}
    evidence_text = " ".join(
        str(value or "")
        for value in (
            item_row.get("caption", ""),
            item_row.get("reason", ""),
            mapping_entry.get("reason", ""),
            mapping_entry.get("source_note", ""),
        )
    )
    labels = fallback_group_labels_from_text(evidence_text)
    rows: list[dict[str, Any]] = []
    for label in labels:
        row = base_unified_row(paper_folder, item_id, item_row, mapping_entry)
        row["source_type"] = "image_caption_fallback" if row.get("evidence_image") else "caption_fallback"
        row["evidence_text"] = evidence_text.strip()
        row["confidence"] = "medium"
        row["manual_required"] = "true"
        row["reason"] = "Excel block absent; fallback condition/formulation row generated from explicit caption/image group label. Experimental numeric assay/readout values were not extracted. Manual review required for completeness."
        apply_generic_fallback_group(row, label)
        rows.append(row)
    return rows


def run_heuristic_unified_lnpdb_extraction(paper_folder: Path) -> dict[str, Any]:
    selected_by_id = selected_item_rows_by_id(paper_folder)
    mapping_by_item = mapping_entries_by_item(paper_folder)
    matches_by_item = excel_matches_by_item(paper_folder)
    output_rows: list[dict[str, Any]] = []
    review_flags: list[dict[str, Any]] = []

    for item_id, item_row in selected_by_id.items():
        item_output_count = 0
        matches = matches_by_item.get(item_id, [])
        for match in matches:
            block_csv_path = str(match.get("block_csv_path") or "").strip()
            block_path = path_from_mapping_value(paper_folder, block_csv_path)
            if not block_csv_path or not block_path or not block_path.exists():
                review_flags.append(
                    {
                        "Paper_ID": paper_folder.name,
                        "Item_ID": item_id,
                        "block_id": match.get("block_id", ""),
                        "field": "block_csv_path",
                        "issue": "missing Excel block mapping",
                        "severity": "high",
                        "reason": f"Mapped block CSV is missing: {block_csv_path}",
                    }
                )
                continue
            matrix = read_csv_matrix(block_path)
            labels = excel_context_labels(matrix)
            out = base_unified_row(paper_folder, item_id, item_row, mapping_by_item.get(item_id))
            out.update(
                {
                    "source_type": "excel_block_context",
                    "excel_file": match.get("excel_file", ""),
                    "excel_sheet": match.get("excel_sheet", ""),
                    "block_id": match.get("block_id", ""),
                    "block_csv_path": block_csv_path,
                    "condition_1_name": "excel_context_labels" if labels else "",
                    "condition_1_value": "; ".join(labels),
                    "evidence_excel": block_csv_path,
                    "reason": "Heuristic 06 extraction records Excel block identity, labels, headers, group/formulation context, provenance, and Excel-backed values only when a reliable group-to-value mapping is available. Figure-image digitization is not used.",
                }
            )
            output_rows.append(out)
            item_output_count += 1
        if item_output_count == 0:
            fallback_rows = generic_fallback_rows_for_item(paper_folder, item_id, item_row, mapping_by_item.get(item_id))
            if fallback_rows:
                output_rows.extend(fallback_rows)
                review_flags.append(
                    {
                        "Paper_ID": paper_folder.name,
                        "Item_ID": item_id,
                        "block_id": "",
                        "field": "fallback_group_labels",
                        "issue": "manual fallback review",
                        "severity": "medium",
                        "reason": f"Excel block absent; generated {len(fallback_rows)} fallback condition/formulation rows from caption/image labels.",
                    }
                )
            else:
                out = base_unified_row(paper_folder, item_id, item_row, mapping_by_item.get(item_id))
                out["source_type"] = "manual_review_placeholder"
                out["reason"] = "No mapped Excel context/value block or supported fallback group labels could be converted heuristically; condition/formulation extraction and any Excel-backed value extraction require manual or external-agent review."
                output_rows.append(out)
                review_flags.append(
                    {
                        "Paper_ID": paper_folder.name,
                        "Item_ID": item_id,
                        "block_id": "",
                        "field": "row",
                        "issue": "condition/formulation extraction placeholder only",
                        "severity": "high",
                        "reason": out["reason"],
                    }
                )

    write_csv_dicts(paper_folder / "unified_extraction.csv", output_rows, UNIFIED_EXTRACTION_COLUMNS)
    write_json(
        paper_folder / "unified_extraction.json",
        {
            "created_by": "Agent_Task_Runner heuristic mode",
            "created_at": utc_now(),
            "records": output_rows,
            "source_summary": {
                "selected_items": len(selected_by_id),
                "excel_mapped_items": len(matches_by_item),
            },
        },
    )
    write_csv_dicts(paper_folder / "unified_extraction_review_flags.csv", review_flags, UNIFIED_REVIEW_FLAG_COLUMNS)
    return {"rows": len(output_rows), "review_flags": len(review_flags), "selected_items": len(selected_by_id)}


def ensure_stable_row_ids(rows: list[dict[str, str]], paper_id: str) -> list[dict[str, str]]:
    seen: set[str] = set()
    for idx, row in enumerate(rows, start=1):
        existing = str(row.get("row_id", "")).strip()
        if existing and existing not in seen:
            row["row_id"] = existing
            seen.add(existing)
            continue
        row_id = f"{paper_id}_R{idx:06d}"
        while row_id in seen:
            idx += 1
            row_id = f"{paper_id}_R{idx:06d}"
        row["row_id"] = row_id
        seen.add(row_id)
    return rows


def read_text_with_fallback(path: Path) -> str:
    for encoding in REFERENCE_ENCODINGS:
        try:
            return path.read_text(encoding=encoding)
        except UnicodeDecodeError:
            continue
    return path.read_text(encoding="utf-8", errors="replace")


def source_md_id_for_path(path: Path) -> str:
    stem = re.sub(r"[^A-Za-z0-9_]+", "_", path.stem).strip("_")
    return stem or "source_md"


def classify_source_document_type(path: Path) -> str:
    text = " ".join(part.casefold() for part in path.parts)
    if any(token in text for token in ("source data", "source_data", "sourcedata")):
        return "source_data"
    if "reporting summary" in text or "reporting_summary" in text:
        return "reporting_summary"
    if any(token in text for token in ("moesm", "esm", "supplementary", "supplemental", "supporting information")):
        return "supplementary_information"
    return "main_article"


def find_matching_pdf_for_markdown(md_path: Path) -> str:
    for candidate in (md_path.with_suffix(".pdf"), md_path.parent / f"{md_path.stem}.PDF"):
        if candidate.exists():
            return str(candidate)
    pdfs = sorted(md_path.parent.glob("*.pdf")) + sorted(md_path.parent.glob("*.PDF"))
    return str(pdfs[0]) if pdfs else ""


def extract_title_and_doi_from_text(text: str) -> tuple[str, str]:
    title = ""
    doi = ""
    for line in text.splitlines():
        clean = normalize_sentence_text(line)
        if not clean:
            continue
        if not title and len(clean.split()) >= 3 and not clean.lower().startswith(("abstract", "figure", "fig.")):
            title = truncate_text(clean, 220)
        doi_match = re.search(r"\b10\.\d{4,9}/[-._;()/:A-Z0-9]+\b", clean, flags=re.IGNORECASE)
        if doi_match:
            doi = doi_match.group(0)
        if title and doi:
            break
    return title, doi


def build_paper_source_context(paper_folder: Path, paper_id: str, sentence_index: dict[str, Any]) -> dict[str, Any]:
    documents: list[dict[str, Any]] = []
    seen: set[str] = set()
    for source in sentence_index.get("sources", []):
        relpath = source.get("source_md_relpath", "")
        if not relpath:
            continue
        md_path = paper_folder / relpath
        source_md_id = source.get("source_md_id", "") or source_md_id_for_path(md_path)
        if source_md_id in seen:
            continue
        seen.add(source_md_id)
        doc_type = classify_source_document_type(md_path)
        try:
            title, doi = extract_title_and_doi_from_text(read_text_with_fallback(md_path)[:8000])
        except Exception:
            title, doi = "", ""
        documents.append(
            {
                "source_doc_id": f"D{len(documents) + 1:03d}",
                "source_document_type": doc_type,
                "source_md_path": str(md_path),
                "source_pdf_path": find_matching_pdf_for_markdown(md_path),
                "source_md_id": source_md_id,
                "title": title,
                "doi": doi,
                "priority": {"main_article": 1, "supplementary_information": 2, "source_data": 3, "reporting_summary": 4}.get(doc_type, 9),
                "sentence_count": source.get("sentence_count", 0),
            }
        )
    documents.sort(key=lambda doc: (doc.get("priority", 9), doc.get("source_md_id", "")))
    return {
        "Paper_ID": paper_id,
        "paper_folder": str(paper_folder),
        "source_documents": documents,
        "source_document_types": dict(Counter(doc["source_document_type"] for doc in documents)),
    }


def is_markdown_table_line(line: str) -> bool:
    stripped = line.strip()
    if not stripped:
        return False
    if stripped.startswith("|") and stripped.endswith("|") and stripped.count("|") >= 2:
        return True
    if re.match(r"^\|?\s*:?-{3,}:?\s*(\|\s*:?-{3,}:?\s*)+\|?$", stripped):
        return True
    return False


def useful_markdown_line(line: str) -> bool:
    stripped = line.strip()
    if not stripped:
        return False
    if is_markdown_table_line(stripped):
        return False
    if re.match(r"^!\[[^\]]*\]\([^)]+\)\s*$", stripped):
        return False
    if len(stripped) > 2000 and sum(ch.isalpha() for ch in stripped) < len(stripped) * 0.25:
        return False
    return True


def normalize_sentence_text(text: str) -> str:
    text = re.sub(r"!\[[^\]]*\]\([^)]+\)", " ", str(text or ""))
    text = re.sub(r"\[[^\]]+\]\([^)]+\)", lambda m: m.group(0).split("]")[0].lstrip("["), text)
    text = re.sub(r"[`*_>#]+", " ", text)
    return re.sub(r"\s+", " ", text).strip()


def sentence_normalized_for_match(text: str) -> str:
    return normalize_sentence_text(text).casefold()


def infer_item_hint(text: str, fallback: str = "") -> str:
    patterns = [
        r"\b(?:Fig\.?|Figure)\s*S?\d+[A-Za-z]?(?:[-,]\s*[A-Za-z])?",
        r"\bSupplementary\s+(?:Fig\.?|Figure)\s*\d+[A-Za-z]?",
        r"\bfigure\s+\d+[a-z]\b",
    ]
    for pattern in patterns:
        match = re.search(pattern, text, flags=re.IGNORECASE)
        if match:
            return re.sub(r"\s+", " ", match.group(0)).strip()
    return fallback


def infer_sentence_kind(sentence: str, section_heading: str, source_path: Path) -> str:
    text = sentence.strip()
    lowered = text.lower()
    if text.startswith("#"):
        return "section"
    if re.match(r"^(fig\.?|figure|supplementary fig\.?|supplementary figure)\b", lowered):
        return "caption" if "supplementary" not in lowered else "supplementary_caption"
    if any(token in (section_heading or "").lower() for token in ("method", "materials", "experimental")):
        return "methods"
    if len(text.split()) <= 12 and text.istitle():
        return "title"
    if "supplement" in source_path.stem.lower() and "fig" in lowered:
        return "supplementary_caption"
    if len(text) > 30:
        return "body"
    return "unknown"


def split_sentence_like_chunks(text: str) -> list[tuple[int, int, str]]:
    protected = text
    replacements = {
        "Fig.": "Fig§",
        "fig.": "fig§",
        "Dr.": "Dr§",
        "e.g.": "e§g§",
        "i.e.": "i§e§",
        "et al.": "et al§",
    }
    for src, dst in replacements.items():
        protected = protected.replace(src, dst)
    chunks: list[tuple[int, int, str]] = []
    start = 0
    for match in re.finditer(r"(?<=[.!?])\s+(?=[A-Z0-9\[])|\n{2,}", protected):
        end = match.start()
        raw = text[start:end].strip()
        if raw:
            chunks.append((start + len(text[start:end]) - len(text[start:end].lstrip()), end, raw.replace("§", ".")))
        start = match.end()
    tail = text[start:].strip()
    if tail:
        chunks.append((start + len(text[start:]) - len(text[start:].lstrip()), len(text), tail.replace("§", ".")))
    merged: list[tuple[int, int, str]] = []
    for chunk_start, chunk_end, chunk_text in chunks:
        if merged and len(chunk_text) < 25:
            prev_start, _, prev_text = merged.pop()
            merged.append((prev_start, chunk_end, f"{prev_text} {chunk_text}".strip()))
        else:
            merged.append((chunk_start, chunk_end, chunk_text))
    return merged


def iter_source_markdown_files(paper_folder: Path) -> list[Path]:
    files: list[Path] = []
    for path in paper_folder.rglob("*.md"):
        rel_parts = path.relative_to(paper_folder).parts
        if any(part in MARKDOWN_SENTENCE_INDEX_SKIP_DIRS or part.startswith(".") for part in rel_parts):
            continue
        files.append(path)
    return sorted(files)


def build_markdown_sentence_index(paper_folder: Path) -> dict[str, Any]:
    index_dir = paper_folder / "markdown_sentence_index"
    index_dir.mkdir(parents=True, exist_ok=True)
    markdown_files = iter_source_markdown_files(paper_folder)
    all_rows: list[dict[str, str]] = []
    manifest_sources: list[dict[str, Any]] = []
    used_source_ids: Counter[str] = Counter()

    for md_path in markdown_files:
        relpath = md_path.relative_to(paper_folder).as_posix()
        base_id = source_md_id_for_path(md_path)
        used_source_ids[base_id] += 1
        source_md_id = base_id if used_source_ids[base_id] == 1 else f"{base_id}_{used_source_ids[base_id]}"
        text = read_text_with_fallback(md_path)
        kept_segments: list[tuple[int, str, str, str]] = []
        section_heading = ""
        item_hint = ""
        cursor = 0
        current_lines: list[str] = []
        current_start = 0
        current_section = ""
        current_item_hint = ""

        def flush_current() -> None:
            nonlocal current_lines, current_start, current_section, current_item_hint
            paragraph = " ".join(line.strip() for line in current_lines if line.strip()).strip()
            if paragraph:
                kept_segments.append((current_start, paragraph, current_section, current_item_hint))
            current_lines = []

        for raw_line in text.splitlines(keepends=True):
            line_start = cursor
            cursor += len(raw_line)
            line = raw_line.rstrip("\r\n")
            stripped = line.strip()
            if stripped.startswith("#"):
                flush_current()
                section_heading = stripped.lstrip("#").strip()
                current_section = section_heading
                current_item_hint = infer_item_hint(stripped, item_hint)
                kept_segments.append((line_start, section_heading, section_heading, current_item_hint))
                continue
            if not useful_markdown_line(line):
                flush_current()
                continue
            hint = infer_item_hint(stripped, item_hint)
            if hint:
                item_hint = hint
            if not current_lines:
                current_start = line_start
                current_section = section_heading
                current_item_hint = item_hint
            current_lines.append(stripped)
        flush_current()

        rows: list[dict[str, str]] = []
        sentence_counter = 0
        for segment_start, paragraph, heading, hint in kept_segments:
            for chunk_start, chunk_end, sentence_text in split_sentence_like_chunks(paragraph):
                sentence = normalize_sentence_text(sentence_text)
                if len(sentence) < 8:
                    continue
                sentence_counter += 1
                sentence_id = f"S{sentence_counter:06d}"
                row = {
                    "source_md_id": source_md_id,
                    "sentence_id": sentence_id,
                    "global_sentence_id": f"{source_md_id}:{sentence_id}",
                    "source_md_path": str(md_path),
                    "source_md_relpath": relpath,
                    "source_page": "",
                    "section_heading": heading,
                    "item_hint": hint,
                    "sentence_kind": infer_sentence_kind(sentence, heading, md_path),
                    "sentence_text": sentence,
                    "sentence_text_normalized": sentence_normalized_for_match(sentence),
                    "char_start": str(segment_start + chunk_start),
                    "char_end": str(segment_start + chunk_end),
                }
                rows.append(row)
                all_rows.append(row)

        sentence_md = index_dir / f"{source_md_id}.sentences.md"
        sentence_csv = index_dir / f"{source_md_id}.sentences.csv"
        sentence_json = index_dir / f"{source_md_id}.sentences.json"
        sentence_md.write_text(
            "# Sentence index for " + md_path.name + "\n"
            + f"Source: {relpath}\n\n"
            + "\n".join(f"[{row['sentence_id']}] {row['sentence_text']}" for row in rows)
            + ("\n" if rows else ""),
            encoding="utf-8",
        )
        write_csv_dicts(sentence_csv, rows, MARKDOWN_SENTENCE_INDEX_COLUMNS)
        write_json(sentence_json, {"source_md_id": source_md_id, "source_md_relpath": relpath, "sentences": rows})
        manifest_sources.append(
            {
                "source_md_id": source_md_id,
                "source_md_relpath": relpath,
                "source_document_type": classify_source_document_type(md_path),
                "sentence_count": len(rows),
            }
        )

    all_csv = index_dir / "markdown_sentence_index_all.csv"
    manifest_json = index_dir / "markdown_sentence_index_manifest.json"
    write_csv_dicts(all_csv, all_rows, MARKDOWN_SENTENCE_INDEX_COLUMNS)
    write_json(
        manifest_json,
        {
            "created_at": utc_now(),
            "paper_folder": str(paper_folder),
            "markdown_file_count": len(markdown_files),
            "sentence_count": len(all_rows),
            "sources": manifest_sources,
            "table_regions_excluded": True,
        },
    )
    return {
        "index_dir": str(index_dir),
        "markdown_files": len(markdown_files),
        "sentence_rows": len(all_rows),
        "all_csv": str(all_csv),
        "manifest_json": str(manifest_json),
        "rows": all_rows,
        "sources": manifest_sources,
    }


def normalized_evidence_text(value: Any) -> str:
    return re.sub(r"\s+", " ", str(value or "").strip()).lower()


def compact_evidence_phrase(evidence_text: str, cell_value: str, column_name: str, max_chars: int = 320) -> str:
    text = re.sub(r"\s+", " ", str(evidence_text or "").strip())
    if not text:
        return ""
    value = str(cell_value or "").strip()
    candidates = [value]
    candidates.extend(part for part in re.split(r"[_\s/-]+", value) if len(part) >= 3)
    lowered = text.lower()
    hit_index = -1
    for candidate in candidates:
        candidate = candidate.strip()
        if not candidate:
            continue
        idx = lowered.find(candidate.lower())
        if idx >= 0:
            hit_index = idx
            break
    if hit_index < 0:
        keyword_map = {
            "Route_of_administration": ["intratum", "intraven", "subcutaneous", "injection", "administer"],
            "Dose_ug_nucleicacid": ["ug", "µg", "μg", "ng", "dose"],
            "Experiment_method": ["flow cytometry", "elisa", "ivis", "luminescence", "rna-seq", "qpcr", "western"],
            "Model_type": ["mc38", "mouse", "mice", "cell"],
            "Model_target": ["tumor", "tumour", "spleen", "liver", "lung"],
            "Cargo": ["mrna", "sirna", "pdna", "sgrna", "sarna"],
            "Cargo_type": ["fluc", "nluc", "il-12", "ova", "cre", "cas9"],
        }
        for keyword in keyword_map.get(column_name, []):
            idx = lowered.find(keyword)
            if idx >= 0:
                hit_index = idx
                break
    if hit_index >= 0:
        start = max(0, hit_index - max_chars // 2)
        end = min(len(text), start + max_chars)
        return text[start:end].strip(" ;,.")
    return truncate_text(text, max_chars)


def infer_evidence_source_type(row: dict[str, str], column_name: str) -> str:
    source_type = str(row.get("source_type", "")).strip()
    if source_type == "manual_review_placeholder":
        return "manual_review_placeholder"
    if source_type in {"caption_fallback", "image_caption_fallback"}:
        return source_type
    if row.get("evidence_excel") or row.get("block_csv_path"):
        if column_name in DEFERRED_VALUE_COLUMNS_06:
            return "excel_block_value"
        if column_name in {"Formulation_Name", "formulation_id", "condition_1_name", "condition_1_value", "condition_2_name", "condition_2_value", "condition_3_name", "condition_3_value", "condition_4_name", "condition_4_value"}:
            return "excel_block"
    if row.get("source_image") or row.get("evidence_image"):
        if not row.get("evidence_text"):
            return "figure_image"
    if row.get("source_pdf") or row.get("source_page"):
        return "pdf_caption"
    if row.get("evidence_text"):
        return "markdown"
    return "inferred_from_context"


def evidence_source_path(row: dict[str, str], evidence_type: str) -> str:
    if evidence_type in {"caption_fallback", "image_caption_fallback"}:
        return row.get("source_pdf", "") or row.get("source_image", "") or row.get("evidence_image", "")
    if evidence_type in {"excel_block", "excel_block_value", "source_data_excel"}:
        return row.get("block_csv_path", "") or row.get("evidence_excel", "")
    if evidence_type == "figure_image":
        return row.get("evidence_image", "") or row.get("source_image", "")
    if evidence_type in {"pdf_caption", "pdf_text"}:
        return row.get("source_pdf", "")
    if evidence_type in {"smiles_resolved", "compound_inventory"}:
        return evidence_type
    return row.get("source_pdf", "") or row.get("source_image", "") or row.get("block_csv_path", "")


def support_role_for_cell(row: dict[str, str], column_name: str, evidence_type: str) -> str:
    if normalize_bool_text(row.get("manual_required")) == "true" and not str(row.get(column_name, "")).strip():
        return "unresolved_manual_review"
    if evidence_type in {"inferred_from_context", "manual_review_placeholder"}:
        return "inferred" if evidence_type == "inferred_from_context" else "unresolved_manual_review"
    if column_name in {"condition_1_name", "condition_1_value", "condition_2_name", "condition_2_value", "condition_3_name", "condition_3_value", "condition_4_name", "condition_4_value"}:
        return "supporting_context"
    if column_name in SCALAR_CONDITION_COLUMNS_06 or column_name.endswith("_molarratio"):
        return "normalized"
    return "direct"


def cell_requires_evidence(column_name: str, value: Any) -> bool:
    if column_name in ADMINISTRATIVE_EVIDENCE_EXCLUDED_COLUMNS or column_name in LNPDB_LIKE_EXCLUDED_COLUMNS:
        return False
    if column_name in SCIENTIFIC_EVIDENCE_OPTIONAL_COLUMNS:
        return False
    return bool(str(value or "").strip())


def column_id(column_name: str) -> str:
    return "C_" + re.sub(r"[^A-Za-z0-9_]+", "_", column_name).strip("_")


def make_evidence_row(row: dict[str, str], column_name: str, cell_value: str) -> dict[str, str]:
    evidence_type = infer_evidence_source_type(row, column_name)
    text = ""
    normalized_from = ""
    if evidence_type in {"excel_block", "excel_block_value", "source_data_excel"}:
        text = str(row.get("evidence_excel") or row.get("block_csv_path") or row.get("excel_sheet") or "").strip()
        if evidence_type == "excel_block_value":
            text = f"{text}; metric_type={row.get('metric_type', '')}; original_values={row.get('original_values', '')}; aggregated_value={row.get('aggregated_value', '')}; unit={row.get('unit', '')}; replicate_type={row.get('replicate_type', '')}"
        normalized_from = text
    else:
        text = compact_evidence_phrase(row.get("evidence_text", ""), cell_value, column_name)
        normalized_from = text
    if not text:
        text = str(row.get("reason") or "No direct source phrase available; evidence requires manual review.").strip()
        evidence_type = "manual_review_placeholder" if normalize_bool_text(row.get("manual_required")) == "true" else "inferred_from_context"
    source_page = str(row.get("source_page", "")).strip()
    return {
        "evidence_id": "",
        "Paper_ID": row.get("Paper_ID", ""),
        "Item_ID": row.get("Item_ID", ""),
        "evidence_summary": "",
        "evidence_sentence_ids": "",
        "evidence_sentence_texts": "",
        "evidence_text_exact": truncate_text(text, 500),
        "evidence_text_normalized": truncate_text(normalized_evidence_text(text), 500),
        "evidence_source_type": evidence_type if evidence_type in EVIDENCE_SOURCE_TYPES else "inferred_from_context",
        "evidence_source_path": evidence_source_path(row, evidence_type),
        "source_pdf": row.get("source_pdf", ""),
        "source_page": source_page,
        "source_image": row.get("source_image", "") or row.get("evidence_image", ""),
        "evidence_excel": row.get("evidence_excel", ""),
        "excel_file": row.get("excel_file", ""),
        "excel_sheet": row.get("excel_sheet", ""),
        "block_id": row.get("block_id", ""),
        "block_csv_path": row.get("block_csv_path", ""),
        "excel_cell_or_range": "",
        "confidence": row.get("confidence", ""),
        "manual_required": normalize_bool_text(row.get("manual_required")),
        "reason": row.get("reason", ""),
        "pdf_page_index": source_page,
        "pdf_text_quote": truncate_text(text, 500) if evidence_type in {"pdf_caption", "pdf_text", "markdown"} else "",
        "pdf_char_start": "",
        "pdf_char_end": "",
        "pdf_bbox_x0": "",
        "pdf_bbox_y0": "",
        "pdf_bbox_x1": "",
        "pdf_bbox_y1": "",
        "image_bbox_x0": "",
        "image_bbox_y0": "",
        "image_bbox_x1": "",
        "image_bbox_y1": "",
    }


def evidence_summary_for_columns(row: dict[str, str], supported_columns: list[str], evidence_type: str) -> str:
    cols = set(supported_columns)
    method = row.get("Experiment_method", "")
    model = row.get("Model_type", "") or row.get("Model", "")
    target = row.get("Model_target", "")
    cargo = row.get("Cargo", "")
    cargo_type = row.get("Cargo_type", "")
    if evidence_type == "excel_block":
        formulation = row.get("Formulation_Name") or row.get("formulation_id") or "formulation"
        return f"Excel block supports {formulation} treatment/formulation identity."
    if evidence_type in {"excel_block_value", "source_data_excel"}:
        metric = row.get("metric_type") or "experimental value"
        return f"Mapped Excel/source-data block supports {metric} metric/value fields."
    if evidence_type in {"caption_fallback", "image_caption_fallback"}:
        item = row.get("Item_ID", "item")
        return f"Caption/image fallback supports treatment group labels for {item}."
    if evidence_type == "manual_review_placeholder":
        return "Manual review placeholder for missing or uncertain source evidence."
    if {"Model_target", "Experiment_method"} & cols and method:
        return f"Source supports {target or 'target'} context and {method} readout."
    if {"Model", "Model_type", "Route_of_administration", "Cargo", "Cargo_type", "Dose_ug_nucleicacid"} & cols:
        details = " ".join(part for part in (model, cargo, cargo_type) if part)
        return f"Source supports {details or 'model/treatment'} condition."
    if {"IL_name", "HL_name", "CHL_name", "PEG_name", "IL_molarratio", "HL_molarratio", "CHL_molarratio", "PEG_molarratio"} & cols:
        formulation = row.get("Formulation_Name") or row.get("formulation_id") or "formulation"
        return f"Source supports {formulation} formulation composition."
    return "Source supports listed LNPDB condition/formulation columns."


def keyword_values_for_sentence_match(row: dict[str, str], supported_columns: list[str], evidence_text: str) -> list[str]:
    values: list[str] = []
    for col in supported_columns:
        value = str(row.get(col, "")).strip()
        if value and value.upper() != "N/A":
            values.append(value)
            values.extend(part for part in re.split(r"[_\s;/,-]+", value) if len(part) >= 3)
    item = str(row.get("Item_ID", "")).strip()
    if item:
        values.append(item)
        values.extend(re.findall(r"\d+[A-Za-z]?", item))
    values.extend(re.findall(r"[A-Za-z0-9][A-Za-z0-9+\-]{2,}", str(evidence_text or "")))
    values.extend(
        [
            "MC38",
            "mice",
            "tumour",
            "tumor",
            "intratumoural",
            "intratumoral",
            "intravenous",
            "in vitro",
            "mRNA",
            "FLuc",
            "IL-12",
            "flow cytometry",
            "CD8",
            "CD4",
            "ELISA",
            "IVIS",
            "luminescence",
            "RNA-seq",
            "citrate",
            "PBS",
            "cholesterol",
            "DSPC",
            "DOPE",
            "DMG-PEG",
            "molar ratio",
        ]
    )
    unique: list[str] = []
    seen: set[str] = set()
    for value in values:
        norm = sentence_normalized_for_match(value)
        if len(norm) < 2 or norm in seen:
            continue
        seen.add(norm)
        unique.append(value)
    return unique


def score_sentence_for_evidence(
    sentence_row: dict[str, str],
    row: dict[str, str],
    supported_columns: list[str],
    keywords: list[str],
    item_id: str,
    figure_tokens: list[str],
) -> int:
    sentence = sentence_row.get("sentence_text_normalized") or sentence_normalized_for_match(sentence_row.get("sentence_text", ""))
    if not sentence:
        return 0
    score = 0
    item_hint = str(sentence_row.get("item_hint", "")).strip().lower()
    if item_id and (item_id in item_hint or item_id in sentence):
        score += 8
    if any(token and token in sentence for token in figure_tokens):
        score += 2
    for value in keywords:
        norm = sentence_normalized_for_match(value)
        if norm and norm in sentence:
            score += 3 if len(norm) >= 5 else 1
    if "Experiment_method" in supported_columns and any(token in sentence for token in ("flow cytometry", "elisa", "ivis", "luminescence", "rna-seq", "qpcr", "cd8", "cd4")):
        score += 4
    if any(col in supported_columns for col in ("IL_name", "HL_name", "CHL_name", "PEG_name", "IL_molarratio", "HL_molarratio", "CHL_molarratio", "PEG_molarratio")) and any(token in sentence for token in ("lipid", "lnp", "cholesterol", "dspc", "peg", "molar", "ratio", "formulat")):
        score += 4
    return score


def find_sentence_support(
    sentence_rows: list[dict[str, str]],
    row: dict[str, str],
    supported_columns: list[str],
    evidence_text: str,
    max_sentences: int = 3,
) -> tuple[str, str]:
    if not sentence_rows or not supported_columns:
        return "", ""
    scored: list[tuple[int, int, dict[str, str]]] = []
    keywords = keyword_values_for_sentence_match(row, supported_columns, evidence_text)
    item_id = str(row.get("Item_ID", "")).strip().lower()
    figure_tokens = re.findall(r"\d+[a-z]?", item_id)
    figure_numbers = [match for match in re.findall(r"\d+", item_id)]
    item_candidate_items = [
        (idx, sentence_row)
        for idx, sentence_row in enumerate(sentence_rows)
        if item_id
        and (
            item_id in str(sentence_row.get("item_hint", "")).strip().lower()
            or item_id in str(sentence_row.get("sentence_text_normalized", ""))
            or any(re.search(rf"\b(?:fig\.?|figure)\s*{re.escape(num)}\b", str(sentence_row.get("sentence_text_normalized", ""))) for num in figure_numbers)
            or any(re.search(rf"\b(?:fig\.?|figure)\s*{re.escape(num)}\b", str(sentence_row.get("item_hint", "")).strip().lower()) for num in figure_numbers)
        )
    ]
    candidate_items = item_candidate_items or [
        (idx, sentence_row)
        for idx, sentence_row in enumerate(sentence_rows)
        if not item_id
        or item_id in str(sentence_row.get("item_hint", "")).strip().lower()
        or any(token and token in str(sentence_row.get("sentence_text_normalized", "")) for token in figure_tokens)
        or any(sentence_normalized_for_match(value) in str(sentence_row.get("sentence_text_normalized", "")) for value in keywords[:20])
    ]
    if not candidate_items:
        candidate_items = list(enumerate(sentence_rows))
    for idx, sentence_row in candidate_items:
        score = score_sentence_for_evidence(sentence_row, row, supported_columns, keywords, item_id, figure_tokens)
        if score > 0:
            scored.append((score, idx, sentence_row))
    scored.sort(key=lambda item: (-item[0], item[1]))
    selected = [sentence for score, _, sentence in scored[:max_sentences] if score >= 6]
    if not selected:
        return "", ""
    ids = "|".join(sentence["global_sentence_id"] for sentence in selected if sentence.get("global_sentence_id"))
    texts = " | ".join(truncate_text(sentence.get("sentence_text", ""), 240) for sentence in selected)
    return ids, texts


def split_pipe_or_semicolon(value: Any) -> list[str]:
    text = str(value or "").strip()
    if not text:
        return []
    separator = "|" if "|" in text else ";"
    return [part.strip() for part in text.split(separator) if part.strip()]


def evidence_dedupe_key(evidence: dict[str, str]) -> tuple[str, ...]:
    return (
        evidence.get("Paper_ID", ""),
        evidence.get("Item_ID", ""),
        evidence.get("evidence_text_exact", ""),
        evidence.get("evidence_source_type", ""),
        evidence.get("source_pdf", ""),
        evidence.get("source_page", ""),
        evidence.get("source_image", ""),
        evidence.get("evidence_excel", ""),
        evidence.get("block_csv_path", ""),
    )


def scientific_nonempty_columns(row: dict[str, str]) -> list[str]:
    return [col for col in ALLOWED_FIGURE_EVIDENCE_COLUMNS if str(row.get(col, "")).strip()]


def supported_columns_for_evidence(row: dict[str, str], evidence_type: str, evidence_text: str) -> list[str]:
    nonempty = set(scientific_nonempty_columns(row))
    text = normalized_evidence_text(evidence_text)
    supported: set[str] = set()
    if evidence_type == "manual_review_placeholder":
        return sorted(nonempty)
    if evidence_type in {"excel_block_value", "source_data_excel"}:
        return [col for col in DEFERRED_VALUE_COLUMNS_06 if col in nonempty]
    if evidence_type == "excel_block":
        supported.update(col for col in nonempty if col in {
            "formulation_id",
            "Formulation_Name",
            "IL_name",
            "IL_molarratio",
            "HL_name",
            "HL_molarratio",
            "CHL_name",
            "CHL_molarratio",
            "PEG_name",
            "PEG_molarratio",
            "Fifth_component_name",
            "Fifth_component_molarratio",
        })
    if any(token in text for token in ("mc38", "mouse", "mice", "tumour", "tumor", "cell", "in vitro", "in vivo")):
        supported.update(col for col in ("Model", "Model_type") if col in nonempty)
    if any(token in text for token in ("intratum", "intraven", "subcutaneous", "intramuscular", "administer", "injection", "treated")):
        supported.update(col for col in ("Route_of_administration",) if col in nonempty)
    if any(token in text for token in ("mrna", "sirna", "pdna", "sgrna", "sarna", "fluc", "nluc", "il-12", "ova", "cre", "cas9")):
        supported.update(col for col in ("Cargo", "Cargo_type") if col in nonempty)
    # Dose cells are mapped by dedicated dose-evidence objects so broad caption
    # evidence does not claim dose support unless a dose sentence was linked.
    if any(token in text for token in ("spleen", "liver", "lung", "tumour", "tumor", "organ", "tissue")):
        supported.update(col for col in ("Model_target",) if col in nonempty)
    if any(token in text for token in ("flow cytometry", "elisa", "ivis", "luminescence", "rna-seq", "qpcr", "western", "cd8", "cd4", "treg", "cytokine")):
        supported.update(col for col in ("Experiment_method",) if col in nonempty)
    # Buffer and mixing-method cells are mapped by dedicated global methods
    # evidence objects so figure captions do not overclaim preparation support.
    if any(token in text for token in ("formulation", "lnp", "lipid", "cholesterol", "dspc", "peg", "molar", "ratio")):
        supported.update(col for col in nonempty if col in {
            "formulation_id",
            "Formulation_Name",
            "IL_name",
            "IL_molarratio",
            "HL_name",
            "HL_molarratio",
            "CHL_name",
            "CHL_molarratio",
            "PEG_name",
            "PEG_molarratio",
            "Fifth_component_name",
            "Fifth_component_molarratio",
            "IL_to_nucleicacid_massratio",
        })
    if not supported:
        # Fall back to grouping all non-empty scientific columns by source object; this keeps evidence item-level instead of per-cell.
        supported.update(nonempty - {"Aqueous_buffer", "Dialysis_buffer", "Mixing_method", "Dose_ug_nucleicacid"})
    return [col for col in ALLOWED_FIGURE_EVIDENCE_COLUMNS if col in supported]


def sentence_ids_and_texts(sentence_rows: list[dict[str, str]], rows: list[dict[str, str]]) -> tuple[str, str]:
    ids = "|".join(row.get("global_sentence_id", "") for row in rows if row.get("global_sentence_id"))
    texts = " | ".join(truncate_text(row.get("sentence_text", ""), 240) for row in rows)
    return ids, texts


def find_buffer_mixing_sentences(sentence_rows: list[dict[str, str]]) -> tuple[str, str]:
    wanted = [
        ("citrate buffer",),
        ("pipette mixing",),
        ("microfluidic chip",),
        ("dialysing against", "pbs"),
    ]
    selected: list[dict[str, str]] = []
    seen: set[str] = set()
    for terms in wanted:
        for row in sentence_rows:
            text = sentence_normalized_for_match(row.get("sentence_text", ""))
            if all(term in text for term in terms):
                gid = row.get("global_sentence_id", "")
                if gid and gid not in seen:
                    selected.append(row)
                    seen.add(gid)
                break
    return sentence_ids_and_texts(sentence_rows, selected)


def sentence_source_document_type(sentence_row: dict[str, str], paper_context: dict[str, Any]) -> str:
    source_md_id = sentence_row.get("source_md_id", "")
    for doc in paper_context.get("source_documents", []):
        if doc.get("source_md_id") == source_md_id:
            return doc.get("source_document_type", "")
    return ""


def build_global_methods_contexts(sentence_rows: list[dict[str, str]], paper_context: dict[str, Any]) -> list[dict[str, str]]:
    patterns: list[tuple[str, list[tuple[str, ...]], str]] = [
        (
            "lnp_preparation",
            [
                ("preparation of other lnps",),
                ("mrna was dissolved", "citrate buffer"),
                ("pipette mixing",),
                ("microfluidic chip",),
                ("dialysing against", "pbs"),
            ],
            "Global LNP preparation methods support aqueous buffer, dialysis buffer, and mixing method.",
        ),
        (
            "in_vivo_dose_2_5",
            [
                ("2.5", "per mouse"),
                ("2.5", "intratum"),
            ],
            "Global in vivo methods/caption evidence supports 2.5 ug nucleic-acid dose per mouse.",
        ),
        (
            "flow_cytometry_methods",
            [
                ("flow cytometry",),
                ("cd8", "flow"),
                ("cd4", "flow"),
            ],
            "Paper-level methods evidence supports flow-cytometry readout context.",
        ),
    ]
    contexts: list[dict[str, str]] = []
    for context_type, term_groups, summary in patterns:
        selected: list[dict[str, str]] = []
        seen: set[str] = set()
        for terms in term_groups:
            for row in sentence_rows:
                text = sentence_normalized_for_match(row.get("sentence_text", ""))
                if all(term in text for term in terms):
                    gid = row.get("global_sentence_id", "")
                    if gid and gid not in seen:
                        selected.append(row)
                        seen.add(gid)
                    break
        if not selected:
            continue
        ids, texts = sentence_ids_and_texts(sentence_rows, selected)
        doc_types = sorted({sentence_source_document_type(row, paper_context) for row in selected if sentence_source_document_type(row, paper_context)})
        source_md_ids = sorted({row.get("source_md_id", "") for row in selected if row.get("source_md_id")})
        contexts.append(
            {
                "global_context_id": f"GC{len(contexts) + 1:04d}",
                "context_type": context_type,
                "evidence_sentence_ids": ids,
                "evidence_sentence_texts": texts,
                "evidence_summary": summary,
                "source_document_type": "|".join(doc_types),
                "source_md_id": "|".join(source_md_ids),
                "priority": "1" if "main_article" in doc_types else "2",
            }
        )
    return contexts


def global_context_by_type(contexts: list[dict[str, str]], context_type: str) -> dict[str, str]:
    for context in contexts:
        if context.get("context_type") == context_type:
            return context
    return {}


def is_lnp_preparation_row(row: dict[str, str]) -> bool:
    haystack = " ".join(
        str(row.get(col, ""))
        for col in (
            "Formulation_Name",
            "formulation_id",
            "IL_name",
            "HL_name",
            "CHL_name",
            "PEG_name",
            "source_type",
        )
    ).casefold()
    if any(token in haystack for token in ("pbs", "untreated", "free ido", "free il-12 mrna", "control group")) and not any(
        token in haystack for token in ("lnp", "plnp", "mc3", "g0-ss", "g0-6c")
    ):
        return False
    if any(str(row.get(col, "")).strip() for col in ("IL_name", "HL_name", "CHL_name", "PEG_name")):
        return True
    return any(token in haystack for token in ("lnp", "plnp", "mc3", "g0-ss", "g0-6c"))


def find_dose_sentences(sentence_rows: list[dict[str, str]], row: dict[str, str]) -> tuple[str, str]:
    item_id = str(row.get("Item_ID", "")).strip().lower()
    figure_numbers = re.findall(r"\d+", item_id)
    scored: list[tuple[int, int, dict[str, str]]] = []
    for idx, sentence_row in enumerate(sentence_rows):
        text = sentence_normalized_for_match(sentence_row.get("sentence_text", ""))
        if "2.5" not in text:
            continue
        if not any(token in text for token in ("per mouse", "mouse", "dose", "intratum", "administer", "injection")):
            continue
        score = 10
        hint = sentence_normalized_for_match(sentence_row.get("item_hint", ""))
        for num in figure_numbers:
            if re.search(rf"\b(?:fig\.?|figure)\s*{re.escape(num)}\b", text) or re.search(rf"\b(?:fig\.?|figure)\s*{re.escape(num)}\b", hint):
                score += 12
        if "per mouse" in text:
            score += 5
        if "intratum" in text:
            score += 3
        scored.append((score, idx, sentence_row))
    scored.sort(key=lambda item: (-item[0], item[1]))
    selected = [row for _, _, row in scored[:2]]
    return sentence_ids_and_texts(sentence_rows, selected)


def make_column_specific_evidence(
    row: dict[str, str],
    evidence_type: str,
    summary: str,
    sentence_ids: str,
    sentence_texts: str,
) -> dict[str, str]:
    text = sentence_texts or summary
    return {
        "evidence_id": "",
        "Paper_ID": row.get("Paper_ID", ""),
        "Item_ID": row.get("Item_ID", ""),
        "evidence_summary": summary,
        "evidence_sentence_ids": sentence_ids,
        "evidence_sentence_texts": sentence_texts,
        "evidence_text_exact": truncate_text(text, 500),
        "evidence_text_normalized": truncate_text(normalized_evidence_text(text), 500),
        "evidence_source_type": evidence_type,
        "evidence_source_path": row.get("source_pdf", "") or row.get("source_image", ""),
        "source_pdf": row.get("source_pdf", ""),
        "source_page": row.get("source_page", ""),
        "source_image": row.get("source_image", "") or row.get("evidence_image", ""),
        "evidence_excel": "",
        "excel_file": "",
        "excel_sheet": "",
        "block_id": row.get("block_id", ""),
        "block_csv_path": "",
        "excel_cell_or_range": "",
        "confidence": row.get("confidence", ""),
        "manual_required": normalize_bool_text(row.get("manual_required")),
        "reason": summary,
        "pdf_page_index": row.get("source_page", ""),
        "pdf_text_quote": truncate_text(text, 500),
        "pdf_char_start": "",
        "pdf_char_end": "",
        "pdf_bbox_x0": "",
        "pdf_bbox_y0": "",
        "pdf_bbox_x1": "",
        "pdf_bbox_y1": "",
        "image_bbox_x0": "",
        "image_bbox_y0": "",
        "image_bbox_x1": "",
        "image_bbox_y1": "",
    }


def figure_evidence_reason(evidence_type: str, supported_columns: list[str]) -> str:
    cols = set(supported_columns)
    if evidence_type == "excel_block":
        return "Excel block labels support treatment/formulation group identity."
    if evidence_type in {"excel_block_value", "source_data_excel"}:
        return "Mapped Excel/source-data block supports experimental metric, original values, aggregate, unit, and replicate type."
    if evidence_type in {"caption_fallback", "image_caption_fallback"}:
        return "Caption/image fallback supports treatment/formulation group labels and condition context."
    if evidence_type == "methods_global":
        return "Paper-level global methods evidence supports listed condition/formulation columns across source documents in this paper package."
    if evidence_type == "methods_item_specific":
        return "Item-specific caption or methods evidence supports listed condition columns."
    if evidence_type == "manual_review_placeholder":
        return "Manual review placeholder records missing or uncertain source evidence."
    if {"Model_target", "Experiment_method"} & cols and "Experiment_method" in cols:
        return "Source phrase supports target tissue/site and method/readout context."
    if {"Model", "Model_type", "Route_of_administration", "Cargo", "Cargo_type", "Dose_ug_nucleicacid"} & cols:
        return "Source phrase supports model, treatment route, cargo, and dose context."
    if {"IL_name", "HL_name", "CHL_name", "PEG_name", "IL_molarratio", "HL_molarratio", "CHL_molarratio", "PEG_molarratio"} & cols:
        return "Source phrase supports formulation composition context."
    return "Source phrase supports listed LNPDB condition/formulation columns."


def support_scope_for_rows(row_ids: set[str], formulation_ids: set[str], item_row_count: int, evidence_type: str) -> str:
    if evidence_type == "manual_review_placeholder":
        return "manual_review_placeholder"
    if formulation_ids:
        return "formulation_level"
    if len(row_ids) >= item_row_count:
        return "item_level_all_rows"
    if len(row_ids) > 1:
        return "item_level_subset_rows"
    return "column_context"


def build_relational_evidence_tables(
    rows: list[dict[str, str]],
    lnpdb_like_fieldnames: list[str],
    sentence_rows: list[dict[str, str]] | None = None,
    paper_context: dict[str, Any] | None = None,
    global_methods_contexts: list[dict[str, str]] | None = None,
) -> tuple[list[dict[str, str]], list[dict[str, str]], dict[str, Any]]:
    evidence_by_key: dict[tuple[str, ...], dict[str, str]] = {}
    evidence_rows: list[dict[str, str]] = []
    grouped: dict[tuple[str, str, tuple[str, ...]], dict[str, Any]] = {}
    excluded_admin_count = 0
    broad_caption_overclaim_fixed = 0
    non_lnp_rows_excluded_from_lnp_preparation_backfill = 0
    rows_by_item = Counter(row.get("Item_ID", "") for row in rows)

    def add_evidence_group(row: dict[str, str], evidence: dict[str, str], supported_columns: list[str]) -> None:
        nonlocal broad_caption_overclaim_fixed
        if not supported_columns:
            return
        key = evidence_dedupe_key(evidence)
        if key not in evidence_by_key:
            evidence["evidence_id"] = f"{row.get('Paper_ID', '')}_E{len(evidence_rows) + 1:06d}"
            evidence["reason"] = evidence.get("reason") or figure_evidence_reason(evidence.get("evidence_source_type", ""), supported_columns)
            evidence_by_key[key] = evidence
            evidence_rows.append(evidence)
        evidence_row = evidence_by_key[key]
        if {"Aqueous_buffer", "Dialysis_buffer", "Mixing_method", "Dose_ug_nucleicacid"} & set(supported_columns):
            broad_caption_overclaim_fixed += 1
        group_key = (row.get("Item_ID", ""), evidence_row["evidence_id"], tuple(supported_columns))
        group = grouped.setdefault(
            group_key,
            {
                "Paper_ID": row.get("Paper_ID", ""),
                "Item_ID": row.get("Item_ID", ""),
                "evidence_id": evidence_row["evidence_id"],
                "supported_columns": set(supported_columns),
                "supported_row_ids": set(),
                "supported_formulation_ids": set(),
                "confidence": row.get("confidence", ""),
                "manual_required": normalize_bool_text(row.get("manual_required")),
                "evidence_source_type": evidence_row.get("evidence_source_type", ""),
                "evidence_sentence_ids": evidence_row.get("evidence_sentence_ids", ""),
                "evidence_summary": evidence_row.get("evidence_summary", ""),
            },
        )
        group["supported_row_ids"].add(row.get("row_id", ""))
        if row.get("formulation_id"):
            group["supported_formulation_ids"].add(row.get("formulation_id", ""))
        if normalize_bool_text(row.get("manual_required")) == "true":
            group["manual_required"] = "true"
        if str(row.get("confidence", "")).lower() in {"low", ""}:
            group["confidence"] = row.get("confidence", "")

    buffer_context = global_context_by_type(global_methods_contexts or [], "lnp_preparation")
    buffer_sentence_ids = buffer_context.get("evidence_sentence_ids", "")
    buffer_sentence_texts = buffer_context.get("evidence_sentence_texts", "")
    if not buffer_sentence_ids:
        buffer_sentence_ids, buffer_sentence_texts = find_buffer_mixing_sentences(sentence_rows or [])
    for row in rows:
        excluded_admin_count += sum(1 for col in ADMINISTRATIVE_EVIDENCE_EXCLUDED_COLUMNS if str(row.get(col, "")).strip())
        row_scientific_cols = scientific_nonempty_columns(row)
        if not row_scientific_cols:
            continue
        seed_columns = ["Experiment_method", "Formulation_Name"]
        if row_has_value_columns(row):
            seed_columns.append("metric_type")
        if not any(row.get(col, "") for col in seed_columns):
            seed_columns = [row_scientific_cols[0]]
        for seed_col in seed_columns:
            if not row.get(seed_col) and seed_col != row_scientific_cols[0]:
                continue
            evidence = make_evidence_row(row, seed_col, str(row.get(seed_col, "")))
            supported_columns = supported_columns_for_evidence(row, evidence.get("evidence_source_type", ""), evidence.get("evidence_text_exact", ""))
            if not supported_columns:
                continue
            summary = evidence_summary_for_columns(row, supported_columns, evidence.get("evidence_source_type", ""))
            sentence_ids, sentence_texts = find_sentence_support(sentence_rows or [], row, supported_columns, evidence.get("evidence_text_exact", ""))
            if summary:
                evidence["evidence_summary"] = summary
            if sentence_ids:
                evidence["evidence_sentence_ids"] = sentence_ids
                evidence["evidence_sentence_texts"] = sentence_texts
            add_evidence_group(row, evidence, supported_columns)

        buffer_cols = [col for col in ("Aqueous_buffer", "Dialysis_buffer", "Mixing_method") if str(row.get(col, "")).strip()]
        if buffer_cols:
            if is_lnp_preparation_row(row):
                summary = "Paper-level Methods evidence supports LNP aqueous buffer, dialysis buffer, and mixing method."
                if buffer_context.get("source_document_type"):
                    summary = f"Global methods evidence from {buffer_context.get('source_document_type')} supports LNP aqueous buffer, dialysis buffer, and mixing method for this paper-package row."
                evidence = make_column_specific_evidence(
                    row,
                    "methods_global",
                    summary,
                    buffer_sentence_ids,
                    buffer_sentence_texts,
                )
                add_evidence_group(row, evidence, buffer_cols)
            else:
                non_lnp_rows_excluded_from_lnp_preparation_backfill += 1

        if str(row.get("Dose_ug_nucleicacid", "")).strip() == "2.5":
            dose_sentence_ids, dose_sentence_texts = find_dose_sentences(sentence_rows or [], row)
            dose_context = global_context_by_type(global_methods_contexts or [], "in_vivo_dose_2_5")
            if not dose_sentence_ids and dose_context:
                dose_sentence_ids = dose_context.get("evidence_sentence_ids", "")
                dose_sentence_texts = dose_context.get("evidence_sentence_texts", "")
            evidence = make_column_specific_evidence(
                row,
                "methods_global" if dose_context and dose_context.get("evidence_sentence_ids") == dose_sentence_ids else "methods_item_specific",
                "Caption or paper-level methods evidence supports 2.5 ug nucleic-acid dose per mouse.",
                dose_sentence_ids,
                dose_sentence_texts,
            )
            add_evidence_group(row, evidence, ["Dose_ug_nucleicacid"])
    map_rows: list[dict[str, str]] = []
    for group in grouped.values():
        supported_columns = [col for col in ALLOWED_FIGURE_EVIDENCE_COLUMNS if col in group["supported_columns"]]
        row_ids = sorted(v for v in group["supported_row_ids"] if v)
        formulation_ids = sorted(v for v in group["supported_formulation_ids"] if v)
        scope = support_scope_for_rows(set(row_ids), set(formulation_ids), rows_by_item.get(group["Item_ID"], 0), group.get("evidence_source_type", ""))
        map_rows.append(
            {
                "Paper_ID": group["Paper_ID"],
                "Item_ID": group["Item_ID"],
                "evidence_id": group["evidence_id"],
                "supported_columns": "|".join(supported_columns),
                "supported_column_count": str(len(supported_columns)),
                "supported_row_ids": "|".join(row_ids),
                "supported_formulation_ids": "|".join(formulation_ids),
                "support_scope": scope,
                "evidence_sentence_ids": group.get("evidence_sentence_ids", ""),
                "evidence_summary": group.get("evidence_summary", ""),
                "confidence": group.get("confidence", ""),
                "manual_required": group.get("manual_required", ""),
                "reason": figure_evidence_reason(group.get("evidence_source_type", ""), supported_columns),
            }
        )
    stats = {
        "excluded_administrative_cell_count": excluded_admin_count,
        "broad_caption_overclaim_fixed": broad_caption_overclaim_fixed,
        "non_lnp_rows_excluded_from_lnp_preparation_backfill": non_lnp_rows_excluded_from_lnp_preparation_backfill,
    }
    return evidence_rows, map_rows, stats


def lnpdb_like_fieldnames_from_rows(rows: list[dict[str, str]]) -> list[str]:
    source_fields = list(rows[0].keys()) if rows else UNIFIED_EXTRACTION_COLUMNS
    fields = ["row_id"]
    for field in source_fields:
        if field == "row_id" or field in LNPDB_LIKE_EXCLUDED_COLUMNS:
            continue
        fields.append(field)
    return fields


def count_nonempty_scientific_cells(rows: list[dict[str, str]], fieldnames: list[str]) -> int:
    count = 0
    for row in rows:
        for field in fieldnames:
            if cell_requires_evidence(field, row.get(field, "")):
                count += 1
    return count


def run_heuristic_finalize_unified_table(paper_folder: Path) -> dict[str, Any]:
    source = paper_folder / "unified_extraction.csv"
    require_existing_file(source, "07_finalize_unified_table")
    rows = read_csv_rows(source)
    paper_id = paper_folder.name
    for row in rows:
        row.setdefault("Paper_ID", paper_id)
        row["manual_required"] = normalize_bool_text(row.get("manual_required", ""))
        row.setdefault("confidence", "")
        force_blank_output_smiles(row)
    ensure_stable_row_ids(rows, paper_id)
    source_fieldnames = list(rows[0].keys()) if rows else ["row_id", *UNIFIED_EXTRACTION_COLUMNS]
    if "row_id" not in source_fieldnames:
        source_fieldnames = ["row_id", *source_fieldnames]
    lnpdb_like_fieldnames = lnpdb_like_fieldnames_from_rows(rows)
    sentence_index = build_markdown_sentence_index(paper_folder)
    sentence_rows = sentence_index.get("rows", [])
    paper_context = build_paper_source_context(paper_folder, paper_id, sentence_index)
    write_json(paper_folder / "paper_source_context.json", paper_context)
    global_methods_contexts = build_global_methods_contexts(sentence_rows, paper_context)
    evidence_rows, map_rows, evidence_stats = build_relational_evidence_tables(
        rows,
        lnpdb_like_fieldnames,
        sentence_rows,
        paper_context,
        global_methods_contexts,
    )

    write_csv_dicts(paper_folder / "unified_extraction_final.csv", rows, source_fieldnames)
    write_csv_dicts(paper_folder / "unified_extraction_lnpdb_like.csv", rows, lnpdb_like_fieldnames)
    write_csv_dicts(paper_folder / "unified_extraction_source_evidence.csv", evidence_rows, SOURCE_EVIDENCE_COLUMNS)
    write_csv_dicts(paper_folder / "unified_extraction_figure_evidence_map.csv", map_rows, FIGURE_EVIDENCE_MAP_COLUMNS)

    manual_count = sum(1 for row in rows if normalize_bool_text(row.get("manual_required")) == "true")
    low_confidence_count = sum(1 for row in rows if str(row.get("confidence", "")).strip().lower() in {"", "low", "unmatched"})
    missing_item_id = sum(1 for row in rows if not str(row.get("Item_ID", "")).strip())
    selected_items = set(selected_item_rows_by_id(paper_folder))
    excel_items = {row.get("Item_ID", "") for row in rows if row.get("source_type") == "excel_block_context"}
    fallback_items = {row.get("Item_ID", "") for row in rows if row.get("source_type") in {"image_caption_fallback", "caption_fallback"}}
    placeholder_items = {row.get("Item_ID", "") for row in rows if row.get("source_type") == "manual_review_placeholder"}
    selected_items_without_excel_blocks = selected_items - excel_items
    evidence_source_type_counts = Counter(row.get("evidence_source_type", "") for row in evidence_rows)
    evidence_usage = Counter(row.get("evidence_id", "") for row in map_rows)
    smiles_rows = read_csv_rows(paper_folder / "smiles_resolved.csv") if (paper_folder / "smiles_resolved.csv").exists() else []
    image_structure_smiles_rows_used = sum(1 for row in smiles_rows if row_resolved_smiles_value(row) and is_disallowed_image_smiles_row(row))
    smiles_from_reference_or_curated_count = sum(
        1
        for row in smiles_rows
        if row_resolved_smiles_value(row)
        and any(token in row_text_for_smiles_source(row) for token in ("reference", "curated", "manual_verified", "human_curated", "api_free_curated"))
        and not is_disallowed_image_smiles_row(row)
    )
    smiles_from_text_iupac_count = sum(
        1
        for row in smiles_rows
        if row_resolved_smiles_value(row)
        and any(token in row_text_for_smiles_source(row) for token in ("iupac", "opsin", "text_lookup", "text/name", "text_name"))
        and not is_disallowed_image_smiles_row(row)
    )
    smiles_unresolved_due_to_disabled_structure_image_count = sum(
        1
        for row in smiles_rows
        if not row_resolved_smiles_value(row)
        and "structure-image-based smiles extraction is disabled" in row_text_for_smiles_source(row)
    )
    excel_value_rows = [row for row in rows if row_has_value_columns(row)]
    excel_value_row_ids = {row.get("row_id", "") for row in excel_value_rows}
    excel_value_rows_without_excel_provenance = sum(1 for row in excel_value_rows if not value_row_has_excel_provenance(row))
    excel_value_rows_numeric_parse_failures = sum(
        1
        for row in excel_value_rows
        if not is_pipe_numeric_like_text(row.get("original_values", ""))
        or not is_numeric_like_text(str(row.get("aggregated_value", "")).replace(",", "").replace("%", ""))
    )
    excel_value_rows_computed_mean = sum(1 for row in excel_value_rows if "|" in str(row.get("original_values", "")) and str(row.get("aggregated_value", "")).strip())
    image_digitized_value_rows_used = sum(1 for row in excel_value_rows if value_row_uses_image_digitization(row))
    nonempty_output_smiles_cells = count_nonempty_output_smiles_cells(rows)
    row_id_counts = Counter(row.get("row_id", "") for row in rows)
    duplicate_row_ids = sorted(row_id for row_id, count in row_id_counts.items() if row_id and count > 1)
    empty_evidence_text_rows = sum(1 for row in evidence_rows if not str(row.get("evidence_text_exact", "")).strip())
    scientific_cell_count = count_nonempty_scientific_cells(rows, lnpdb_like_fieldnames)
    items_with_scientific_columns = {row.get("Item_ID", "") for row in rows if scientific_nonempty_columns(row)}
    items_with_evidence = {row.get("Item_ID", "") for row in map_rows if row.get("supported_columns")}
    items_missing_evidence = sorted(item for item in items_with_scientific_columns if item and item not in items_with_evidence)
    invalid_supported_columns = 0
    for map_row in map_rows:
        for col in split_pipe_or_semicolon(map_row.get("supported_columns", "")):
            if col not in ALLOWED_FIGURE_EVIDENCE_COLUMN_SET:
                invalid_supported_columns += 1
    sentence_id_set = {row.get("global_sentence_id", "") for row in sentence_rows if row.get("global_sentence_id")}
    source_evidence_with_sentence_ids = sum(1 for row in evidence_rows if str(row.get("evidence_sentence_ids", "")).strip())
    figure_map_with_sentence_ids = sum(1 for row in map_rows if str(row.get("evidence_sentence_ids", "")).strip())
    cross_document_evidence_rows = sum(1 for row in evidence_rows if row.get("evidence_source_type") == "methods_global")
    missing_source_sentence_ids = sum(
        1
        for row in evidence_rows
        for sentence_id in split_pipe_or_semicolon(row.get("evidence_sentence_ids", ""))
        if sentence_id not in sentence_id_set
    )
    missing_map_sentence_ids = sum(
        1
        for row in map_rows
        for sentence_id in split_pipe_or_semicolon(row.get("evidence_sentence_ids", ""))
        if sentence_id not in sentence_id_set
    )
    buffer_mixing_row_ids = {
        row.get("row_id", "")
        for row in rows
        if any(str(row.get(col, "")).strip() for col in ("Aqueous_buffer", "Dialysis_buffer", "Mixing_method"))
    }
    lnp_buffer_mixing_row_ids = {
        row.get("row_id", "")
        for row in rows
        if row.get("row_id", "") in buffer_mixing_row_ids and is_lnp_preparation_row(row)
    }
    dose_row_ids = {row.get("row_id", "") for row in rows if str(row.get("Dose_ug_nucleicacid", "")).strip() == "2.5"}
    buffer_mixing_with_evidence: set[str] = set()
    dose_with_evidence: set[str] = set()
    for map_row in map_rows:
        supported = set(split_pipe_or_semicolon(map_row.get("supported_columns", "")))
        row_ids = set(split_pipe_or_semicolon(map_row.get("supported_row_ids", "")))
        has_sentence_ids = bool(split_pipe_or_semicolon(map_row.get("evidence_sentence_ids", "")))
        if has_sentence_ids and supported & {"Aqueous_buffer", "Dialysis_buffer", "Mixing_method"}:
            buffer_mixing_with_evidence.update(row_ids)
        if has_sentence_ids and "Dose_ug_nucleicacid" in supported:
            dose_with_evidence.update(row_ids)
    support_scope_counts = Counter(row.get("support_scope", "") for row in map_rows)
    warnings = []
    if items_missing_evidence:
        warnings.append(f"figure/items with non-empty scientific cells but no evidence mapping: {len(items_missing_evidence)}")
    if empty_evidence_text_rows:
        warnings.append(f"evidence rows with empty evidence_text_exact: {empty_evidence_text_rows}")
    if duplicate_row_ids:
        warnings.append(f"duplicate row_id values: {len(duplicate_row_ids)}")
    write_json(
        paper_folder / "unified_extraction_qc_report.json",
        {
            "created_by": "Agent_Task_Runner heuristic mode",
            "created_at": utc_now(),
            "rows": len(rows),
            "lnpdb_like_rows": len(rows),
            "lnpdb_like_columns": len(lnpdb_like_fieldnames),
            "lnpdb_like_nonempty_cells": scientific_cell_count,
            "image_caption_fallback_rows": sum(1 for row in rows if row.get("source_type") == "image_caption_fallback"),
            "caption_fallback_rows": sum(1 for row in rows if row.get("source_type") == "caption_fallback"),
            "manual_review_placeholder_rows": sum(1 for row in rows if row.get("source_type") == "manual_review_placeholder"),
            "selected_items_without_excel_blocks": len(selected_items_without_excel_blocks),
            "selected_items_expanded_by_fallback": len(fallback_items),
            "selected_items_still_placeholder": len(placeholder_items - fallback_items),
            "markdown_sentence_index_files": sentence_index.get("markdown_files", 0),
            "markdown_sentence_index_rows": sentence_index.get("sentence_rows", 0),
            "source_evidence_rows": len(evidence_rows),
            "smiles_output_columns_forced_blank": True,
            "nonempty_output_smiles_cells": nonempty_output_smiles_cells,
            "image_structure_smiles_disabled": True,
            "image_structure_smiles_rows_used": image_structure_smiles_rows_used,
            "smiles_from_reference_or_curated_count": smiles_from_reference_or_curated_count,
            "smiles_from_text_iupac_count": smiles_from_text_iupac_count,
            "smiles_unresolved_due_to_disabled_structure_image_count": smiles_unresolved_due_to_disabled_structure_image_count,
            "excel_value_rows_populated": len(excel_value_rows),
            "excel_value_items_populated": len({row.get("Item_ID", "") for row in excel_value_rows if row.get("Item_ID")}),
            "excel_value_rows_without_excel_provenance": excel_value_rows_without_excel_provenance,
            "excel_value_rows_numeric_parse_failures": excel_value_rows_numeric_parse_failures,
            "excel_value_rows_computed_mean": excel_value_rows_computed_mean,
            "excel_value_rows_left_blank_no_excel_mapping": sum(1 for row in rows if not row_has_value_columns(row) and not str(row.get("block_csv_path", "")).strip()),
            "image_digitized_value_rows_used": image_digitized_value_rows_used,
            "source_evidence_with_sentence_ids": source_evidence_with_sentence_ids,
            "source_evidence_without_sentence_ids": len(evidence_rows) - source_evidence_with_sentence_ids,
            "figure_evidence_map_rows": len(map_rows),
            "figure_evidence_map_with_sentence_ids": figure_map_with_sentence_ids,
            "fallback_rows_with_sentence_ids": sum(1 for row in evidence_rows if row.get("evidence_source_type") in {"caption_fallback", "image_caption_fallback"} and row.get("evidence_sentence_ids")),
            "fallback_rows_missing_sentence_ids": sum(1 for row in evidence_rows if row.get("evidence_source_type") in {"caption_fallback", "image_caption_fallback"} and not row.get("evidence_sentence_ids")),
            "paper_source_documents_count": len(paper_context.get("source_documents", [])),
            "paper_source_document_types": paper_context.get("source_document_types", {}),
            "global_methods_context_count": len(global_methods_contexts),
            "cross_document_evidence_rows": cross_document_evidence_rows,
            "buffer_mixing_rows_with_evidence": len(buffer_mixing_row_ids & buffer_mixing_with_evidence),
            "buffer_mixing_rows_missing_evidence": len(buffer_mixing_row_ids - buffer_mixing_with_evidence),
            "buffer_mixing_rows_with_global_methods_evidence": len(lnp_buffer_mixing_row_ids & buffer_mixing_with_evidence),
            "buffer_mixing_rows_missing_global_methods_evidence": len(lnp_buffer_mixing_row_ids - buffer_mixing_with_evidence),
            "dose_rows_with_evidence": len(dose_row_ids & dose_with_evidence),
            "dose_rows_with_item_or_global_evidence": len(dose_row_ids & dose_with_evidence),
            "dose_rows_missing_evidence": len(dose_row_ids - dose_with_evidence),
            "broad_caption_overclaim_fixed": evidence_stats.get("broad_caption_overclaim_fixed", 0),
            "non_lnp_rows_excluded_from_lnp_preparation_backfill": evidence_stats.get("non_lnp_rows_excluded_from_lnp_preparation_backfill", 0),
            "figure_items_with_evidence": len(items_with_evidence),
            "figure_items_missing_evidence": len(items_missing_evidence),
            "excluded_administrative_cell_count": evidence_stats.get("excluded_administrative_cell_count", 0),
            "evidence_rows_with_pdf_page": sum(1 for row in evidence_rows if str(row.get("source_page", "")).strip() or str(row.get("pdf_page_index", "")).strip()),
            "evidence_rows_with_excel_block": sum(1 for row in evidence_rows if str(row.get("block_csv_path", "")).strip() or str(row.get("evidence_excel", "")).strip()),
            "evidence_rows_with_bbox": sum(1 for row in evidence_rows if any(str(row.get(col, "")).strip() for col in ("pdf_bbox_x0", "image_bbox_x0"))),
            "evidence_source_type_counts": dict(evidence_source_type_counts),
            "evidence_supporting_multiple_columns": sum(1 for row in map_rows if int(row.get("supported_column_count") or 0) > 1),
            "invalid_supported_columns_count": invalid_supported_columns,
            "missing_source_sentence_ids": missing_source_sentence_ids,
            "missing_figure_map_sentence_ids": missing_map_sentence_ids,
            "evidence_supporting_multiple_items": sum(1 for count in evidence_usage.values() if count > 1),
            "support_scope_counts": dict(support_scope_counts),
            "manual_required_evidence_rows": sum(1 for row in evidence_rows if normalize_bool_text(row.get("manual_required")) == "true"),
            "manual_required_rows": manual_count,
            "low_confidence_rows": low_confidence_count,
            "missing_item_id_rows": missing_item_id,
            "duplicate_row_ids": duplicate_row_ids,
            "figure_items_missing_evidence_ids": items_missing_evidence,
            "warnings": warnings,
            "source": "unified_extraction.csv",
        },
    )
    return {
        "rows": len(rows),
        "manual_required_rows": manual_count,
        "low_confidence_rows": low_confidence_count,
        "markdown_sentence_index_rows": sentence_index.get("sentence_rows", 0),
        "source_evidence_rows": len(evidence_rows),
        "figure_evidence_map_rows": len(map_rows),
    }


def run_heuristic_ft_excel_matcher(paper_folder: Path) -> dict[str, Any]:
    classified = paper_folder / "fig_table_lnpdb_classified.csv"
    inventory = paper_folder / "excel_block_inventory.csv"
    require_existing_file(classified, "04_ft_excel_matcher")
    require_existing_file(inventory, "04_ft_excel_matcher")
    classified_rows = read_csv_rows(classified)
    block_rows = read_csv_rows(inventory)
    selected = selected_ft_rows(classified_rows)
    block_keywords: list[tuple[dict[str, str], set[str], str]] = []
    for block in block_rows:
        preview = block_preview_text(paper_folder, block.get("block_csv_path", ""))
        words = keyword_set(
            block.get("excel_file", ""),
            block.get("excel_sheet", ""),
            block.get("block_id", ""),
            block.get("block_type", ""),
            preview,
        )
        block_keywords.append((block, words, preview))

    mapping: dict[str, list[dict[str, Any]]] = {}
    row_outputs: list[dict[str, Any]] = []
    best_by_item: dict[str, dict[str, Any]] = {}
    for row in selected:
        item_id = row_item_id(row)
        words = keyword_set(item_id, row.get("base_id", ""), row.get("caption", ""), row.get("reason", ""))
        best_block: dict[str, str] | None = None
        best_score = 0
        for block, words_for_block, _preview in block_keywords:
            score = len(words & words_for_block)
            normalized_targets = [normalize_token(item_id), normalize_token(row.get("base_id", ""))]
            block_text = normalize_token(" ".join([block.get("excel_file", ""), block.get("excel_sheet", ""), block.get("block_id", "")]))
            score += sum(2 for token in normalized_targets if token and token in block_text)
            if score > best_score:
                best_score = score
                best_block = block
        if best_block and best_score > 0:
            match = {
                "pdf_item_id": item_id,
                "excel_item_id": best_block.get("element_id") or best_block.get("block_id", ""),
                "excel_file": best_block.get("excel_file", ""),
                "excel_sheet": best_block.get("excel_sheet", ""),
                "block_id": best_block.get("block_id", ""),
                "block_csv_path": best_block.get("block_csv_path", ""),
                "confidence": "low",
                "reason": f"Heuristic keyword overlap score={best_score}.",
            }
            mapping[item_id] = [match]
            row_outputs.append(match)
            best_by_item[item_id] = match
        else:
            mapping[item_id] = []

    extra_cols = ["excel_item_id", "matched_blocks", "matched_block_csv_path", "matched_sheet", "matched_sheet_file"]
    existing_fields = list(classified_rows[0].keys()) if classified_rows else []
    fieldnames = existing_fields + [col for col in extra_cols if col not in existing_fields]
    for row in classified_rows:
        item_id = row_item_id(row)
        match = best_by_item.get(item_id)
        if match:
            row["excel_item_id"] = match["excel_item_id"]
            row["matched_blocks"] = match["block_id"]
            row["matched_block_csv_path"] = match["block_csv_path"]
            row["matched_sheet"] = match["excel_sheet"]
            row["matched_sheet_file"] = match["excel_file"]
    write_json(paper_folder / "excel_mapping.json", mapping)
    write_csv_dicts(
        paper_folder / "excel_mapping_rows.csv",
        row_outputs,
        ["pdf_item_id", "excel_item_id", "excel_file", "excel_sheet", "block_id", "block_csv_path", "confidence", "reason"],
    )
    if classified_rows:
        write_csv_dicts(classified, classified_rows, fieldnames)
    return {"selected_items": len(selected), "matched_items": len(row_outputs)}


def run_heuristic_stage(stage: str, paper_folder: Path) -> Any:
    if stage == "03_figure_mapping":
        return run_heuristic_figure_mapping(paper_folder)
    if stage == "03_split_excel_blocks_batch":
        return run_heuristic_split_excel_blocks(paper_folder)
    if stage == "04_figure_separate":
        return run_heuristic_figure_separate(paper_folder)
    if stage == "04_ft_excel_matcher":
        return run_heuristic_ft_excel_matcher(paper_folder)
    if stage == "06_unified_lnpdb_extraction":
        return run_heuristic_unified_lnpdb_extraction(paper_folder)
    if stage == "07_finalize_unified_table":
        return run_heuristic_finalize_unified_table(paper_folder)
    raise ValueError(f"No heuristic implementation for stage: {stage}")


def find_markdown_files(paper_folder: Path) -> list[Path]:
    return [p for p in iter_paper_files(paper_folder) if p.suffix.lower() == ".md" and p.stat().st_size > 0]


DEFERRED_VALUE_COLUMNS_06 = ["metric_type", "original_values", "aggregated_value", "unit", "replicate_type"]
SCALAR_CONDITION_COLUMNS_06 = [
    "Aqueous_buffer",
    "Dialysis_buffer",
    "Mixing_method",
    "Model",
    "Model_type",
    "Model_target",
    "Route_of_administration",
    "Cargo",
    "Cargo_type",
    "Dose_ug_nucleicacid",
    "Experiment_method",
    "Experiment_batching",
]
VALID_MODEL_VALUES_06 = {"", "n/a", "in_vitro", "in_vivo", "ex_vivo"}
VALID_BATCHING_VALUES_06 = {"", "n/a", "individual", "barcoded", "pooled", "grouped"}
PROSE_PHRASES_06 = (" for ", " after ", " before ", " where ", " when ", " panel ")
OR_FORBIDDEN_COLUMNS_06 = {
    "Model",
    "Model_type",
    "Route_of_administration",
    "Cargo",
    "Cargo_type",
    "Experiment_method",
    "Dose_ug_nucleicacid",
}
CARGO_PAYLOAD_TERMS_06 = ("fluc", "nluc", "il-12", "il12", "ova", "cre", "cas9")
FIGURE_4GM_IDS_06 = {f"figure 4{letter}" for letter in "ghijklm"}


def is_na_or_blank(value: Any) -> bool:
    text = str(value or "").strip()
    return text == "" or text.lower() == "n/a"


def is_numeric_like_text(value: Any) -> bool:
    text = str(value or "").strip()
    if is_na_or_blank(text):
        return True
    return bool(re.fullmatch(r"[-+]?(?:\d+(?:\.\d*)?|\.\d+)(?:[eE][-+]?\d+)?", text))


def is_pipe_numeric_like_text(value: Any) -> bool:
    text = str(value or "").strip()
    if is_na_or_blank(text):
        return True
    parts = [part.strip() for part in text.split("|") if part.strip()]
    return bool(parts) and all(is_numeric_like_text(part.replace(",", "").replace("%", "")) for part in parts)


def force_blank_output_smiles(row: dict[str, Any]) -> None:
    for col in OUTPUT_SMILES_COLUMNS:
        row[col] = ""


def count_nonempty_output_smiles_cells(rows: list[dict[str, Any]]) -> int:
    return sum(1 for row in rows for col in OUTPUT_SMILES_COLUMNS if str(row.get(col, "")).strip())


def row_has_value_columns(row: dict[str, str]) -> bool:
    return any(str(row.get(col, "")).strip() for col in DEFERRED_VALUE_COLUMNS_06)


def value_row_has_excel_provenance(row: dict[str, str]) -> bool:
    return bool(str(row.get("evidence_excel", "")).strip() or str(row.get("block_csv_path", "")).strip() or str(row.get("excel_file", "")).strip())


def value_row_uses_image_digitization(row: dict[str, str]) -> bool:
    text = " ".join(str(row.get(col, "")) for col in ("source_type", "reason", "evidence_text", "evidence_image", "source_image")).lower()
    if any(phrase in text for phrase in ("digitization was not used", "image digitization was not used", "figure-image digitization was not used")):
        return False
    return any(token in text for token in ("image_digitized", "digitized_value", "pixel_extraction", "axis-based", "bar height", "heatmap color", "image-derived value", "visual estimation"))


def metric_type_is_concise(value: Any) -> bool:
    text = str(value or "").strip()
    if not text:
        return True
    if len(text) > 80:
        return False
    if any(token in f" {text.lower()} " for token in (" for ", " after ", " before ", " where ", " when ", " panel ")):
        return False
    if ";" in text:
        return False
    return bool(re.fullmatch(r"[A-Za-z0-9_%./+-]+(?:_[A-Za-z0-9%./+-]+)*", text))


def normalize_text_for_validation(value: Any) -> str:
    return re.sub(r"\s+", " ", str(value or "").strip().lower())


def has_prose_like_condition_value(column: str, value: Any) -> bool:
    text = str(value or "").strip()
    if not text or text.lower() == "n/a":
        return False
    lowered = f" {text.lower()} "
    if ";" in text:
        return True
    if any(phrase in lowered for phrase in PROSE_PHRASES_06):
        return True
    if column in OR_FORBIDDEN_COLUMNS_06 and " or " in lowered:
        return True
    if column == "Experiment_method" and "/" in text:
        return True
    return False


def count_stage_06_validation_issues(rows: list[dict[str, str]]) -> dict[str, int]:
    value_rows = 0
    value_rows_without_excel_provenance = 0
    value_rows_numeric_parse_failures = 0
    value_metric_prose_rows = 0
    image_digitized_value_rows_used = 0
    nonempty_output_smiles_cells = count_nonempty_output_smiles_cells(rows)
    prose_like = 0
    invalid_model = 0
    invalid_batching = 0
    invalid_dose = 0
    collapsed_figure_4gm_contexts: set[tuple[str, str]] = set()
    for row in rows:
        if row_has_value_columns(row):
            value_rows += 1
            if not value_row_has_excel_provenance(row):
                value_rows_without_excel_provenance += 1
            if value_row_uses_image_digitization(row):
                image_digitized_value_rows_used += 1
            if not metric_type_is_concise(row.get("metric_type", "")):
                value_metric_prose_rows += 1
            if not is_pipe_numeric_like_text(row.get("original_values", "")) or not is_numeric_like_text(str(row.get("aggregated_value", "")).replace(",", "").replace("%", "")):
                value_rows_numeric_parse_failures += 1

        model_value = str(row.get("Model", "")).strip().lower()
        if model_value not in VALID_MODEL_VALUES_06:
            invalid_model += 1

        batching_value = str(row.get("Experiment_batching", "")).strip().lower()
        if batching_value not in VALID_BATCHING_VALUES_06:
            invalid_batching += 1

        dose = str(row.get("Dose_ug_nucleicacid", "")).strip()
        dose_lower = f" {dose.lower()} "
        if not is_numeric_like_text(dose) or any(word in dose_lower for word in (" ng", " ug", " per ", " mouse", " well", " for ", " after ", " before ")):
            invalid_dose += 1

        row_has_prose = any(has_prose_like_condition_value(col, row.get(col, "")) for col in SCALAR_CONDITION_COLUMNS_06)
        cargo = str(row.get("Cargo", "")).strip().lower()
        if cargo and any(term in cargo for term in CARGO_PAYLOAD_TERMS_06):
            row_has_prose = True
        if row_has_prose:
            prose_like += 1
        item_id = normalize_text_for_validation(row.get("Item_ID", ""))
        if (
            item_id in FIGURE_4GM_IDS_06
            and str(row.get("Model_target", "")).strip().lower() == "spleen"
            and str(row.get("Experiment_method", "")).strip() == "flow_cytometry"
        ):
            collapsed_figure_4gm_contexts.add((item_id, str(row.get("block_id", "")).strip()))
    return {
        "excel_value_rows_populated": value_rows,
        "excel_value_rows_without_excel_provenance": value_rows_without_excel_provenance,
        "excel_value_rows_numeric_parse_failures": value_rows_numeric_parse_failures,
        "value_metric_prose_rows": value_metric_prose_rows,
        "image_digitized_value_rows_used": image_digitized_value_rows_used,
        "nonempty_output_smiles_cells": nonempty_output_smiles_cells,
        "prose_like_condition_rows": prose_like,
        "invalid_model_rows": invalid_model,
        "invalid_batching_rows": invalid_batching,
        "invalid_dose_rows": invalid_dose,
        "lost_readout_specificity_rows": len(collapsed_figure_4gm_contexts) if len(collapsed_figure_4gm_contexts) >= 2 else 0,
    }


def validate_stage(stage: str, paper_folder: Path) -> tuple[bool, list[str]]:
    messages: list[str] = []

    if stage in AGENT_STAGES and not has_manual_marker(paper_folder):
        return False, [f"missing required manual review marker: {paper_folder / MANUAL_MARKER}"]

    if stage == "00_marker":
        md_files = find_markdown_files(paper_folder)
        ok = bool(md_files)
        messages.append(f"non-empty markdown files: {len(md_files)}")
        return ok, messages

    if stage == "01_make_ft_csv":
        path = paper_folder / "fig_table_inventory.csv"
        if not non_empty_file(path):
            return False, [f"missing or empty: {path}"]
        rows = read_csv_rows(path)
        cols = set(rows[0].keys()) if rows else set()
        id_cols = {"item_id", "pdf_item_id", "item"} & cols
        return bool(rows and id_cols), [f"rows={len(rows)}", f"id_columns={sorted(id_cols)}"]

    if stage == "02_ft_selector":
        path = paper_folder / "fig_table_lnpdb_classified.csv"
        if not non_empty_file(path):
            return False, [f"missing or empty: {path}"]
        rows = read_csv_rows(path)
        return bool(rows), [f"rows={len(rows)}"]

    if stage == "02b_manual_review":
        marker = paper_folder / MANUAL_MARKER
        classified = paper_folder / "fig_table_lnpdb_classified.csv"
        reviewed = paper_folder / "fig_table_lnpdb_classified_manual_reviewed.csv"
        has_review_file = reviewed.exists()
        has_manual_col = False
        if classified.exists():
            rows = read_csv_rows(classified)
            has_manual_col = bool(rows and "manual_select" in rows[0])
        ok = marker.exists() and (has_review_file or has_manual_col)
        return ok, [f"marker={marker.exists()}", f"reviewed_copy={has_review_file}", f"manual_select_column={has_manual_col}"]

    if stage == "03_figure_mapping":
        path = paper_folder / "total_figure_mapping.json"
        if not non_empty_file(path):
            return False, [f"missing or empty: {path}"]
        data = json.loads(path.read_text(encoding="utf-8"))
        invalid_qualities = invalid_source_quality_values(data)
        fallback_missing = missing_mapping_paths(
            paper_folder,
            data,
            {"fallback_render", "selected_source_for_paneling"},
        )
        return isinstance(data, dict) and bool(data) and not invalid_qualities, [
            f"top_level_keys={len(data) if isinstance(data, dict) else 'not_object'}",
            f"invalid_source_quality_values={len(invalid_qualities)}",
            f"missing_fallback_or_selected_paths={len(fallback_missing)}",
        ]

    if stage == "03_split_excel_blocks":
        script = PROJECT_ROOT / STAGES[stage]["script"]
        exp_excel = paper_folder / "Exp_Excel"
        excel_files = []
        if exp_excel.exists():
            excel_files = [p for p in iter_paper_files(exp_excel) if p.suffix.lower() in {".xlsx", ".xls", ".csv"}]
        return script.exists(), [f"utility_script_exists={script.exists()}", f"excel_inputs={len(excel_files)}"]

    if stage == "03_split_excel_blocks_batch":
        inv = paper_folder / "excel_block_inventory.csv"
        summary = paper_folder / "three_core_result_all.json"
        block_dir = paper_folder / "Exp_Excel_Blocks"
        if not non_empty_file(inv):
            return False, [f"missing or empty: {inv}"]
        rows = read_csv_rows(inv)
        missing_paths = []
        for row in rows:
            rel = (row.get("block_csv_path") or "").strip()
            if rel and not (paper_folder / rel).exists():
                missing_paths.append(rel)
        ok = bool(rows) and block_dir.exists() and not missing_paths
        return ok, [f"rows={len(rows)}", f"summary_exists={summary.exists()}", f"block_dir_exists={block_dir.exists()}", f"missing_block_paths={len(missing_paths)}"]

    if stage == "04_figure_separate":
        mapping = paper_folder / "total_figure_mapping.json"
        if not non_empty_file(mapping):
            return False, [f"missing or empty: {mapping}"]
        data = json.loads(mapping.read_text(encoding="utf-8"))
        panel_dirs = list(paper_folder.rglob("separated_panels_gemini"))
        invalid_qualities = invalid_source_quality_values(data)
        missing_paths = missing_mapping_paths(
            paper_folder,
            data,
            {"fallback_render", "selected_source_for_paneling"},
        )
        ok = isinstance(data, dict) and not invalid_qualities and not missing_paths
        return ok, [
            f"mapping_keys={len(data) if isinstance(data, dict) else 'not_object'}",
            f"panel_dirs={len(panel_dirs)}",
            f"invalid_source_quality_values={len(invalid_qualities)}",
            f"missing_fallback_or_selected_paths={len(missing_paths)}",
        ]

    if stage == "04_ft_excel_matcher":
        mapping = paper_folder / "excel_mapping.json"
        rows_csv = paper_folder / "excel_mapping_rows.csv"
        if not non_empty_file(mapping):
            return False, [f"missing or empty: {mapping}"]
        data = json.loads(mapping.read_text(encoding="utf-8"))
        ok = isinstance(data, dict) and rows_csv.exists()
        return ok, [f"mapping_keys={len(data) if isinstance(data, dict) else 'not_object'}", f"rows_csv_exists={rows_csv.exists()}"]

    if stage == "05_smiles_structure_resolution":
        path = paper_folder / "smiles_resolved.csv"
        if not non_empty_file(path):
            return False, [f"missing or empty: {path}"]
        rows = read_csv_rows(path)
        cols = set(rows[0].keys()) if rows else set()
        name_cols = {"Name", "name", "compound_id", "Compound_ID"} & cols
        smiles_cols = {"SMILES", "smiles", "resolved_smiles", "Resolved_SMILES"} & cols
        image_structure_smiles_rows_used = sum(
            1 for row in rows if row_resolved_smiles_value(row) and is_disallowed_image_smiles_row(row)
        )
        novel_pil_with_smiles = sum(
            1
            for row in rows
            if row_resolved_smiles_value(row)
            and re.search(r"\b(?:[A-Z0-9]+-)?(?:SS|6C)-AA-C(?:10|12|14)\b", str(row.get("Name") or row.get("standardized_name") or ""), flags=re.I)
            and not is_manual_verified_smiles_row(row)
            and "reference" not in row_text_for_smiles_source(row)
            and "curated" not in row_text_for_smiles_source(row)
        )
        ok = bool(rows and name_cols and smiles_cols) and image_structure_smiles_rows_used == 0 and novel_pil_with_smiles == 0
        return ok, [
            f"rows={len(rows)}",
            f"name_columns={sorted(name_cols)}",
            f"smiles_columns={sorted(smiles_cols)}",
            "image_structure_smiles_disabled=true",
            f"image_structure_smiles_rows_used={image_structure_smiles_rows_used}",
            f"novel_pil_smiles_without_reference_or_manual={novel_pil_with_smiles}",
        ]

    if stage == "06_unified_lnpdb_extraction":
        path = paper_folder / "unified_extraction.csv"
        flags = paper_folder / "unified_extraction_review_flags.csv"
        json_path = paper_folder / "unified_extraction.json"
        if not non_empty_file(path):
            return False, [f"missing or empty: {path}"]
        rows = read_csv_rows(path)
        cols = set(rows[0].keys()) if rows else set()
        missing_cols = [col for col in UNIFIED_EXTRACTION_COLUMNS if col not in cols]
        scalar_issue_counts = count_stage_06_validation_issues(rows)
        selected_count = len(selected_item_rows_by_id(paper_folder))
        empty_item_ids = sum(1 for row in rows if not str(row.get("Item_ID", "")).strip())
        ok_json = True
        if json_path.exists():
            try:
                json.loads(json_path.read_text(encoding="utf-8"))
            except Exception:
                ok_json = False
        ok = (
            bool(rows)
            and not missing_cols
            and empty_item_ids == 0
            and "confidence" in cols
            and "manual_required" in cols
            and flags.exists()
            and ok_json
            and (selected_count == 0 or len(rows) >= 1)
            and all(
                scalar_issue_counts[key] == 0
                for key in (
                    "excel_value_rows_without_excel_provenance",
                    "excel_value_rows_numeric_parse_failures",
                    "value_metric_prose_rows",
                    "image_digitized_value_rows_used",
                    "nonempty_output_smiles_cells",
                    "prose_like_condition_rows",
                    "invalid_model_rows",
                    "invalid_batching_rows",
                    "invalid_dose_rows",
                    "lost_readout_specificity_rows",
                )
            )
        )
        return ok, [
            f"rows={len(rows)}",
            f"selected_items={selected_count}",
            f"missing_required_columns={len(missing_cols)}",
            f"empty_item_ids={empty_item_ids}",
            f"excel_value_rows_populated={scalar_issue_counts['excel_value_rows_populated']}",
            f"excel_value_rows_without_excel_provenance={scalar_issue_counts['excel_value_rows_without_excel_provenance']}",
            f"excel_value_rows_numeric_parse_failures={scalar_issue_counts['excel_value_rows_numeric_parse_failures']}",
            f"value_metric_prose_rows={scalar_issue_counts['value_metric_prose_rows']}",
            f"image_digitized_value_rows_used={scalar_issue_counts['image_digitized_value_rows_used']}",
            "smiles_output_columns_forced_blank=true",
            f"nonempty_output_smiles_cells={scalar_issue_counts['nonempty_output_smiles_cells']}",
            f"prose_like_condition_rows={scalar_issue_counts['prose_like_condition_rows']}",
            f"invalid_model_rows={scalar_issue_counts['invalid_model_rows']}",
            f"invalid_batching_rows={scalar_issue_counts['invalid_batching_rows']}",
            f"invalid_dose_rows={scalar_issue_counts['invalid_dose_rows']}",
            f"lost_readout_specificity_rows={scalar_issue_counts['lost_readout_specificity_rows']}",
            f"review_flags_exists={flags.exists()}",
            f"json_parses={ok_json}",
        ]

    if stage == "07_finalize_unified_table":
        final_csv = paper_folder / "unified_extraction_final.csv"
        lnpdb_csv = paper_folder / "unified_extraction_lnpdb_like.csv"
        evidence_csv = paper_folder / "unified_extraction_source_evidence.csv"
        figure_map_csv = paper_folder / "unified_extraction_figure_evidence_map.csv"
        qc_json = paper_folder / "unified_extraction_qc_report.json"
        sentence_index_csv = paper_folder / "markdown_sentence_index" / "markdown_sentence_index_all.csv"
        if not non_empty_file(final_csv):
            return False, [f"missing or empty: {final_csv}"]
        if not non_empty_file(lnpdb_csv):
            return False, [f"missing or empty: {lnpdb_csv}"]
        if not non_empty_file(evidence_csv):
            return False, [f"missing or empty: {evidence_csv}"]
        if not non_empty_file(figure_map_csv):
            return False, [f"missing or empty: {figure_map_csv}"]
        if not non_empty_file(qc_json):
            return False, [f"missing or empty: {qc_json}"]
        try:
            qc = json.loads(qc_json.read_text(encoding="utf-8"))
        except Exception as exc:
            return False, [f"qc_report_parse_error={exc}"]
        final_rows = read_csv_rows(final_csv)
        lnpdb_rows = read_csv_rows(lnpdb_csv)
        evidence_rows = read_csv_rows(evidence_csv)
        figure_map_rows = read_csv_rows(figure_map_csv)
        sentence_rows = read_csv_rows(sentence_index_csv) if sentence_index_csv.exists() else []
        final_cols = set(final_rows[0].keys()) if final_rows else set()
        lnpdb_cols = set(lnpdb_rows[0].keys()) if lnpdb_rows else set()
        evidence_cols = set(evidence_rows[0].keys()) if evidence_rows else set()
        figure_map_cols = set(figure_map_rows[0].keys()) if figure_map_rows else set()
        lnpdb_row_ids = [str(row.get("row_id", "")).strip() for row in lnpdb_rows]
        evidence_ids = [str(row.get("evidence_id", "")).strip() for row in evidence_rows]
        lnpdb_row_id_set = set(row_id for row_id in lnpdb_row_ids if row_id)
        evidence_id_set = set(evidence_id for evidence_id in evidence_ids if evidence_id)
        sentence_ids = [str(row.get("global_sentence_id", "")).strip() for row in sentence_rows]
        sentence_id_set = set(sentence_id for sentence_id in sentence_ids if sentence_id)
        markdown_files_exist = bool(iter_source_markdown_files(paper_folder))
        duplicate_sentence_ids = len([sentence_id for sentence_id, count in Counter(sentence_ids).items() if sentence_id and count > 1])
        duplicate_row_ids = len([row_id for row_id, count in Counter(lnpdb_row_ids).items() if row_id and count > 1])
        duplicate_evidence_ids = len([evidence_id for evidence_id, count in Counter(evidence_ids).items() if evidence_id and count > 1])
        empty_row_ids = sum(1 for row_id in lnpdb_row_ids if not row_id)
        empty_evidence_ids = sum(1 for evidence_id in evidence_ids if not evidence_id)
        required_map_cols = {"Paper_ID", "Item_ID", "evidence_id", "supported_columns"}
        missing_map_evidence = sum(1 for row in figure_map_rows if str(row.get("evidence_id", "")).strip() not in evidence_id_set)
        lnpdb_items = {str(row.get("Item_ID", "")).strip() for row in lnpdb_rows if str(row.get("Item_ID", "")).strip()}
        missing_map_item_ids = sum(1 for row in figure_map_rows if str(row.get("Item_ID", "")).strip() not in lnpdb_items)
        empty_supported_columns = sum(1 for row in figure_map_rows if not str(row.get("supported_columns", "")).strip())
        invalid_supported_columns = 0
        for row in figure_map_rows:
            for col in split_pipe_or_semicolon(row.get("supported_columns", "")):
                if col not in ALLOWED_FIGURE_EVIDENCE_COLUMN_SET:
                    invalid_supported_columns += 1
        missing_source_sentence_ids = sum(
            1
            for row in evidence_rows
            for sentence_id in split_pipe_or_semicolon(row.get("evidence_sentence_ids", ""))
            if sentence_id not in sentence_id_set
        )
        missing_map_sentence_ids = sum(
            1
            for row in figure_map_rows
            for sentence_id in split_pipe_or_semicolon(row.get("evidence_sentence_ids", ""))
            if sentence_id not in sentence_id_set
        )
        image_structure_smiles_rows_used = int(qc.get("image_structure_smiles_rows_used") or 0)
        nonempty_output_smiles_cells = count_nonempty_output_smiles_cells(final_rows) + count_nonempty_output_smiles_cells(lnpdb_rows)
        qc_nonempty_output_smiles_cells = int(qc.get("nonempty_output_smiles_cells") or 0)
        excel_value_rows_without_excel_provenance = int(qc.get("excel_value_rows_without_excel_provenance") or 0)
        excel_value_rows_numeric_parse_failures = int(qc.get("excel_value_rows_numeric_parse_failures") or 0)
        image_digitized_value_rows_used = int(qc.get("image_digitized_value_rows_used") or 0)
        items_with_scientific_columns = {
            str(row.get("Item_ID", "")).strip()
            for row in lnpdb_rows
            if str(row.get("Item_ID", "")).strip()
            and str(row.get("source_type", "")).strip() != "manual_review_placeholder"
            and scientific_nonempty_columns(row)
        }
        mapped_items = {
            str(row.get("Item_ID", "")).strip()
            for row in figure_map_rows
            if str(row.get("Item_ID", "")).strip() and str(row.get("supported_columns", "")).strip()
        }
        missing_figure_item_evidence = len(items_with_scientific_columns - mapped_items)
        lnp_buffer_row_ids = {
            str(row.get("row_id", "")).strip()
            for row in lnpdb_rows
            if str(row.get("row_id", "")).strip()
            and is_lnp_preparation_row(row)
            and any(str(row.get(col, "")).strip() for col in ("Aqueous_buffer", "Dialysis_buffer", "Mixing_method"))
        }
        lnp_buffer_row_ids_with_evidence: set[str] = set()
        for row in figure_map_rows:
            supported = set(split_pipe_or_semicolon(row.get("supported_columns", "")))
            if not (supported & {"Aqueous_buffer", "Dialysis_buffer", "Mixing_method"}):
                continue
            if not split_pipe_or_semicolon(row.get("evidence_sentence_ids", "")):
                continue
            lnp_buffer_row_ids_with_evidence.update(split_pipe_or_semicolon(row.get("supported_row_ids", "")))
        lnp_buffer_rows_missing_global_methods_evidence = len(lnp_buffer_row_ids - lnp_buffer_row_ids_with_evidence)
        ok = (
            isinstance(qc, dict)
            and bool(final_rows)
            and bool(lnpdb_rows)
            and bool(evidence_rows)
            and bool(figure_map_rows)
            and "row_id" in final_cols
            and "row_id" in lnpdb_cols
            and empty_row_ids == 0
            and duplicate_row_ids == 0
            and {"evidence_id", "evidence_text_exact"} <= evidence_cols
            and empty_evidence_ids == 0
            and duplicate_evidence_ids == 0
            and required_map_cols <= figure_map_cols
            and missing_map_evidence == 0
            and missing_map_item_ids == 0
            and empty_supported_columns == 0
            and invalid_supported_columns == 0
            and missing_figure_item_evidence == 0
            and (not markdown_files_exist or bool(sentence_rows))
            and duplicate_sentence_ids == 0
            and missing_source_sentence_ids == 0
            and missing_map_sentence_ids == 0
            and lnp_buffer_rows_missing_global_methods_evidence == 0
            and image_structure_smiles_rows_used == 0
            and nonempty_output_smiles_cells == 0
            and qc_nonempty_output_smiles_cells == 0
            and excel_value_rows_without_excel_provenance == 0
            and excel_value_rows_numeric_parse_failures == 0
            and image_digitized_value_rows_used == 0
        )
        return ok, [
            f"final_rows={len(final_rows)}",
            f"lnpdb_like_rows={len(lnpdb_rows)}",
            f"source_evidence_rows={len(evidence_rows)}",
            f"figure_evidence_map_rows={len(figure_map_rows)}",
            "qc_report_parses=true",
            f"row_id_in_final={'row_id' in final_cols}",
            f"row_id_in_lnpdb_like={'row_id' in lnpdb_cols}",
            f"empty_row_ids={empty_row_ids}",
            f"duplicate_row_ids={duplicate_row_ids}",
            f"empty_evidence_ids={empty_evidence_ids}",
            f"duplicate_evidence_ids={duplicate_evidence_ids}",
            f"missing_map_evidence_ids={missing_map_evidence}",
            f"missing_map_item_ids={missing_map_item_ids}",
            f"empty_supported_columns={empty_supported_columns}",
            f"invalid_supported_columns_count={invalid_supported_columns}",
            f"missing_figure_item_evidence={missing_figure_item_evidence}",
            f"markdown_sentence_index_rows={len(sentence_rows)}",
            f"duplicate_global_sentence_ids={duplicate_sentence_ids}",
            f"missing_source_sentence_ids={missing_source_sentence_ids}",
            f"missing_figure_map_sentence_ids={missing_map_sentence_ids}",
            f"lnp_buffer_rows_missing_global_methods_evidence={lnp_buffer_rows_missing_global_methods_evidence}",
            "image_structure_smiles_disabled=true",
            f"image_structure_smiles_rows_used={image_structure_smiles_rows_used}",
            "smiles_output_columns_forced_blank=true",
            f"nonempty_output_smiles_cells={nonempty_output_smiles_cells}",
            f"excel_value_rows_without_excel_provenance={excel_value_rows_without_excel_provenance}",
            f"excel_value_rows_numeric_parse_failures={excel_value_rows_numeric_parse_failures}",
            f"image_digitized_value_rows_used={image_digitized_value_rows_used}",
        ]

    raise ValueError(f"Unknown stage: {stage}")


def observe(paper_folder: Path) -> dict[str, Any]:
    paper_files = list(iter_paper_files(paper_folder)) if paper_folder.exists() else []
    files = {
        "pdf": len([p for p in paper_files if p.suffix.lower() == ".pdf"]),
        "markdown": len(find_markdown_files(paper_folder)) if paper_folder.exists() else 0,
        "excel": len([p for p in paper_files if p.suffix.lower() in {".xlsx", ".csv"}]),
    }
    artifacts = {
        "manual_marker": has_manual_marker(paper_folder),
        "fig_table_inventory.csv": (paper_folder / "fig_table_inventory.csv").exists(),
        "fig_table_lnpdb_classified.csv": (paper_folder / "fig_table_lnpdb_classified.csv").exists(),
        "total_figure_mapping.json": (paper_folder / "total_figure_mapping.json").exists(),
        "excel_block_inventory.csv": (paper_folder / "excel_block_inventory.csv").exists(),
        "excel_mapping.json": (paper_folder / "excel_mapping.json").exists(),
        "smiles_resolved.csv": (paper_folder / "smiles_resolved.csv").exists(),
        "unified_extraction.csv": (paper_folder / "unified_extraction.csv").exists(),
        "unified_extraction_final.csv": (paper_folder / "unified_extraction_final.csv").exists(),
    }
    result = {"exists": paper_folder.exists(), "files": files, "artifacts": artifacts}
    append_log(paper_folder, {"action": "observe", "result": result})
    update_state(paper_folder, None, "observed", result)
    return result


def next_stage(paper_folder: Path) -> dict[str, Any]:
    for stage in STAGE_ORDER:
        ok, messages = validate_stage(stage, paper_folder)
        if not ok:
            if stage in AGENT_STAGES and not has_manual_marker(paper_folder):
                return {"next_stage": "02b_manual_review", "blocked": True, "reason": "manual review marker is required before active agent stages"}
            return {"next_stage": stage, "blocked": STAGES.get(stage, {}).get("manual", False), "reason": "; ".join(messages)}
    return {"next_stage": None, "blocked": False, "reason": "all known stages validate"}


def backup_outputs(stage: str, paper_folder: Path) -> list[str]:
    backups: list[str] = []
    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    for output in STAGES[stage].get("outputs", []):
        path = paper_folder / output
        if "*" in output:
            continue
        if path.exists():
            backup = path.with_name(f"{path.name}.bak_{timestamp}")
            if path.is_dir():
                shutil.copytree(path, backup)
            else:
                shutil.copy2(path, backup)
            backups.append(str(backup))
    return backups


def run_legacy_stage(stage: str, paper_folder: Path) -> Any:
    script = PROJECT_ROOT / STAGES[stage]["script"]

    if stage == "00_marker":
        module = import_module_from_path("stage_00_marker", script)
        return module.process_all_pdfs(paper_folder)

    if stage == "01_make_ft_csv":
        module = import_module_from_path("stage_01_make_ft_csv", script)
        return module.process_single_folder(
            target_folder=paper_folder,
            model_name=getattr(module, "MODEL_NAME", "gemini-3.1-pro-preview"),
            api_mode=getattr(module, "API_MODE", "vertex"),
            api_json_name=getattr(module, "API_JSON_NAME", "vertex.json"),
            api_txt_name=getattr(module, "API_TXT_NAME", "gemini_api.txt"),
            project=getattr(module, "PROJECT_ID", None),
            location=getattr(module, "LOCATION", "global"),
            max_input_tokens=getattr(module, "MAX_INPUT_TOKENS", 200000),
            token_count_only=getattr(module, "TOKEN_COUNT_ONLY", False),
        )

    if stage == "02_ft_selector":
        module = import_module_from_path("stage_02_ft_selector", script)
        return module.classify_fig_table_csv_for_lnpdb(
            target_folder=paper_folder,
            inventory_csv_name="fig_table_inventory.csv",
            output_csv_name="fig_table_lnpdb_classified.csv",
            model_name=getattr(module, "MODEL_NAME", "gemini-3.1-pro-preview"),
            api_mode=getattr(module, "API_MODE", "vertex"),
            api_json_name=getattr(module, "API_JSON_NAME", "vertex.json"),
            api_txt_name=getattr(module, "API_TXT_NAME", "gemini_api.txt"),
            project=getattr(module, "PROJECT_ID", None),
            location=getattr(module, "LOCATION", "global"),
            count_only=getattr(module, "COUNT_ONLY_MODE", False),
        )

    if stage == "02b_manual_review":
        raise RuntimeError("Manual review must be run by a human with Streamlit.")

    if stage == "03_figure_mapping":
        module = import_module_from_path("stage_03_figure_mapping", script)
        api_mode = getattr(module, "MAIN_API_MODE", "vertex")
        api_json_name = getattr(module, "MAIN_API_JSON_NAME", "vertex.json")
        project_id = getattr(module, "PROJECT_ID", None)
        if api_mode == "vertex" and hasattr(module, "find_api_key_file"):
            api_key_path = module.find_api_key_file(api_json_name)
            cred_data = json.loads(Path(api_key_path).read_text(encoding="utf-8"))
            project_id = cred_data.get("project_id") or project_id
        return module.run_mapping_main(
            root_dir=paper_folder,
            model_name=getattr(module, "MAIN_MODEL_NAME", "gemini-3.1-pro-preview"),
            api_mode=api_mode,
            api_json_name=api_json_name,
            api_txt_name=getattr(module, "MAIN_API_TXT_NAME", "gemini_api.txt"),
            project_id=project_id,
            location=getattr(module, "MAIN_LOCATION", "global"),
            token_count_only=getattr(module, "MAIN_TOKEN_COUNT_ONLY", False),
            max_input_tokens=getattr(module, "MAIN_MAX_INPUT_TOKENS", 160000),
            classified_csv_path=None,
            exclude_excel_covered=False,
        )

    if stage == "03_split_excel_blocks":
        module = import_module_from_path("stage_03_split_excel_blocks", script)
        return {"imported": bool(module)}

    if stage == "03_split_excel_blocks_batch":
        module = import_module_from_path("stage_03_split_excel_blocks_batch", script)
        api_key_path = module.find_api_key_file(getattr(module, "API_JSON_NAME", "vertex.json"))
        cred_data = json.loads(Path(api_key_path).read_text(encoding="utf-8"))
        project_id = cred_data.get("project_id")
        if not project_id:
            raise ValueError(f"project_id missing in credentials: {api_key_path}")
        client = module.get_vertexai_client(api_key_path, project=project_id)
        return module.process_excel_block_splitter(
            paper_folder,
            client,
            getattr(module, "MODEL_NAME", "gemini-3.1-pro-preview"),
            gcs_bucket=getattr(module, "DEFAULT_GCS_BATCH_BUCKET"),
        )

    if stage == "04_figure_separate":
        module = import_module_from_path("stage_04_figure_separate", script)
        api_key_path = module.find_api_key_file(getattr(module, "API_JSON_NAME", "vertex.json"))
        cred_data = json.loads(Path(api_key_path).read_text(encoding="utf-8"))
        project_id = cred_data.get("project_id")
        if not project_id:
            raise ValueError(f"project_id missing in credentials: {api_key_path}")
        client = module.get_vertexai_client(api_key_path, project=project_id)
        model_name = getattr(module, "BATCH_MODEL_NAME", getattr(module, "MODEL_NAME", "gemini-3.1-pro-preview"))
        return module.run_batch_vlm_separation(
            paper_folder,
            model_name,
            client,
            use_batch_mode=getattr(module, "USE_BATCH_MODE", True),
        )

    if stage == "04_ft_excel_matcher":
        module = import_module_from_path("stage_04_ft_excel_matcher", script)
        api_key_path = module.find_api_key_file(getattr(module, "API_JSON_NAME", "vertex.json"))
        cred_data = json.loads(Path(api_key_path).read_text(encoding="utf-8"))
        project_id = cred_data.get("project_id")
        if not project_id:
            raise ValueError(f"project_id missing in credentials: {api_key_path}")
        client = module.get_vertexai_client(api_key_path, project=project_id)
        return module.process_excel_matcher(paper_folder, client, getattr(module, "MODEL_NAME", "gemini-3.1-pro-preview"))

    if stage in LEGACY_STAGE_SCRIPT_GROUPS:
        scripts = "\n".join(f"- {script_path}" for script_path in LEGACY_STAGE_SCRIPT_GROUPS[stage])
        raise RuntimeError(
            f"{stage} replaces multiple legacy Gemini/API-assisted scripts and has no single safe legacy runner. "
            f"Legacy scripts are preserved for manual legacy mode only:\n{scripts}"
        )

    raise ValueError(f"Unknown stage: {stage}")


def tail_text(value: str, max_chars: int = 4000) -> str:
    if not value:
        return ""
    return value[-max_chars:]


def safe_paper_name(paper_folder: Path) -> str:
    return safe_name(paper_folder.name)


def ensure_cli_full_log_dir(project_root: Path, paper_folder: Path) -> Path:
    path = project_root / "agent_workspace" / "cli_full_logs" / safe_paper_name(paper_folder)
    path.mkdir(parents=True, exist_ok=True)
    return path


def make_cli_full_log_path(project_root: Path, paper_folder: Path, stage: str, attempt: int | None) -> Path:
    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    attempt_number = attempt if attempt is not None else 1
    filename = f"{timestamp}__{safe_name(stage)}__attempt{attempt_number}__external_cli.txt"
    return ensure_cli_full_log_dir(project_root, paper_folder) / filename


def append_text(path: Path, text: str) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    with path.open("a", encoding="utf-8", errors="replace") as f:
        f.write(text)


def write_external_cli_log_header(
    log_path: Path,
    *,
    timestamp_start: str,
    paper_folder: Path,
    stage: str,
    attempt: int | None,
    command_template: str,
    command_display: str,
    command_args: list[str],
    task_file: Path,
    prompt_text: str,
) -> None:
    header = [
        "===== EXTERNAL CLI RUN LOG =====",
        f"timestamp_start: {timestamp_start}",
        f"paper_folder: {paper_folder}",
        f"stage: {stage}",
        f"attempt: {attempt if attempt is not None else ''}",
        f"project_root: {PROJECT_ROOT}",
        f"cwd: {PROJECT_ROOT}",
        f"external_agent_command_template: {command_template}",
        f"final_command: {command_display}",
        f"final_command_argv: {json.dumps(command_args, ensure_ascii=False)}",
        f"task_file: {task_file}",
        "mode: external_agent",
        "",
        "===== BEGIN PROMPT SENT TO EXTERNAL CLI =====",
        prompt_text,
        "===== END PROMPT SENT TO EXTERNAL CLI =====",
        "",
    ]
    log_path.write_text("\n".join(header), encoding="utf-8", errors="replace")


def write_external_cli_log_footer(log_path: Path, *, timestamp_end: str, returncode: Any, elapsed_seconds: float) -> None:
    append_text(
        log_path,
        "\n===== END EXTERNAL CLI OUTPUT =====\n"
        f"timestamp_end: {timestamp_end}\n"
        f"returncode: {returncode}\n"
        f"elapsed_seconds: {elapsed_seconds:.3f}\n",
    )


def run_subprocess_streaming(command_args: list[str], stdin_text: str | None = None, log_path: Path | None = None) -> dict[str, Any]:
    process = subprocess.Popen(
        command_args,
        cwd=PROJECT_ROOT,
        stdin=subprocess.PIPE if stdin_text is not None else None,
        stdout=subprocess.PIPE,
        stderr=subprocess.PIPE,
        text=True,
        encoding="utf-8",
        errors="replace",
    )
    stdout_parts: list[str] = []
    stderr_parts: list[str] = []
    log_lock = threading.Lock()

    def reader(pipe, output, parts: list[str]) -> None:
        try:
            for line in iter(pipe.readline, ""):
                parts.append(line)
                if log_path is not None:
                    with log_lock:
                        append_text(log_path, line)
                output.write(line)
                output.flush()
        finally:
            pipe.close()

    stdout_thread = threading.Thread(target=reader, args=(process.stdout, sys.stdout, stdout_parts), daemon=True)
    stderr_thread = threading.Thread(target=reader, args=(process.stderr, sys.stderr, stderr_parts), daemon=True)
    stdout_thread.start()
    stderr_thread.start()

    if stdin_text is not None and process.stdin is not None:
        try:
            process.stdin.write(stdin_text)
            process.stdin.close()
        except (BrokenPipeError, OSError, ValueError):
            pass

    returncode = process.wait()
    stdout_thread.join()
    stderr_thread.join()
    return {
        "returncode": returncode,
        "stdout_tail": tail_text("".join(stdout_parts)),
        "stderr_tail": tail_text("".join(stderr_parts)),
    }


def validation_result_dict(ok: bool, messages: list[str]) -> dict[str, Any]:
    return {"ok": ok, "messages": messages}


def append_validation_failure_feedback(task_file: Path, validation_result: dict[str, Any], attempt: int) -> None:
    messages = validation_result.get("messages", [])
    lines = [
        "",
        "## Validation Failure Feedback",
        "",
        f"Attempt: {attempt}",
        "",
        "Validation did not pass. Re-read the task instructions and fix the output files without using Gemini/API/find_api/LLM_API/LLM_Batch.",
        "",
        "Validation messages:",
    ]
    lines.extend(f"- {message}" for message in messages)
    lines.extend(
        [
            "",
            "Retry instructions:",
            "- Inspect the current outputs and the validation messages.",
            "- Modify or create only the required output CSV/JSON files.",
            "- Preserve provenance and mark uncertain values with manual_required=true.",
            "- Run the validation command again after changes.",
            "",
        ]
    )
    with task_file.open("a", encoding="utf-8") as f:
        f.write("\n".join(lines))


def build_agent_prompt(stage: str, paper_folder: Path, task_file: Path, validation_result: dict[str, Any] | None = None) -> str:
    task_text = task_file.read_text(encoding="utf-8")
    validation_text = ""
    if validation_result:
        validation_text = (
            "\n\nPrevious validation result:\n"
            + json.dumps(validation_result, ensure_ascii=False, indent=2)
            + "\nFix the validation failures and rerun validation.\n"
        )
    return f"""You are an external CLI coding agent working in the LNPDB_Articles_AgentExtraction repository.

Stage: {stage}
Target paper folder: {paper_folder}
Task file: {task_file}

Read and complete the task markdown below. Use the target paper folder from both this prompt and the task file.

Hard constraints:
- Do not use Gemini/API/find_api/LLM_API/LLM_Batch.
- Do not run legacy Gemini scripts.
- Follow agent_workspace/OUTPUT_SCHEMA.md.
- Create or modify the required output files so Agent_Task_Runner.py validation passes.
- If exact fields are missing, leave them blank and set manual_required=true.
- Record evidence/provenance for nontrivial extracted data.
- Do not delete original PDF or Excel files.
- If overwrite is needed, create a backup or follow the runner backup policy.
- Stage-specific task markdown instructions have priority.
- If you fail, record the cause and what you changed.
{validation_text}
Task markdown:

{task_text}
"""


def split_agent_command(command: str) -> list[str]:
    try:
        parts = shlex.split(command, posix=False)
    except ValueError:
        parts = shlex.split(command)
    cleaned: list[str] = []
    for part in parts:
        if len(part) >= 2 and part[0] == part[-1] and part[0] in {"'", '"'}:
            cleaned.append(part[1:-1])
        else:
            cleaned.append(part)
    return cleaned


def build_agent_command_args(
    command_template: str,
    prompt_text: str,
    task_file: Path,
    paper_folder: Path,
    stage: str,
) -> tuple[list[str], str, bool]:
    parts = split_agent_command(command_template)
    stdin_mode = "{prompt_stdin}" in command_template or "-" in parts
    replacements = {
        "{prompt_text}": "" if stdin_mode else prompt_text,
        "{prompt_stdin}": "-",
        "{prompt_file}": str(task_file),
        "{stage}": stage,
        "{paper_folder}": str(paper_folder),
        "{project_root}": str(PROJECT_ROOT),
    }
    command_args: list[str] = []
    display_parts: list[str] = []
    for part in parts:
        arg = part
        display = part
        for placeholder, value in replacements.items():
            arg = arg.replace(placeholder, value)
            if placeholder == "{prompt_text}":
                display = display.replace(placeholder, "<prompt_text>" if not stdin_mode else "")
            elif placeholder == "{prompt_stdin}":
                display = display.replace(placeholder, "-")
            else:
                display = display.replace(placeholder, value)
        if arg != "":
            command_args.append(arg)
        if display != "":
            display_parts.append(display)
    return command_args, " ".join(display_parts), stdin_mode


def run_external_cli_agent(
    agent: str,
    task_file: Path,
    paper_folder: Path,
    stage: str,
    command_template: str | None = None,
    dry_run: bool = False,
    validation_result: dict[str, Any] | None = None,
    stream_output: bool = False,
    attempt: int | None = None,
) -> dict[str, Any]:
    if agent not in {"codex", "claude", "custom"}:
        raise ValueError(f"Unsupported agent: {agent}")
    if agent == "custom":
        if not command_template:
            raise ValueError("--agent-command is required when --agent custom")
        template = command_template
    else:
        template = command_template or DEFAULT_AGENT_COMMAND_TEMPLATES[agent]

    prompt_text = build_agent_prompt(stage, paper_folder, task_file, validation_result=validation_result)
    command_args, command_display, stdin_mode = build_agent_command_args(template, prompt_text, task_file, paper_folder, stage)
    cli_full_log_path = make_cli_full_log_path(PROJECT_ROOT, paper_folder, stage, attempt) if not dry_run else None
    result: dict[str, Any] = {
        "agent": agent,
        "stage": stage,
        "task_file": str(task_file),
        "prompt_file": str(task_file),
        "prompt_length": len(prompt_text),
        "command": command_display,
        "stdin_mode": stdin_mode,
        "dry_run": dry_run,
        "stream_output": stream_output,
        "cli_full_log_path": str(cli_full_log_path) if cli_full_log_path else "",
    }
    if dry_run:
        result.update({"returncode": None, "stdout_tail": "", "stderr_tail": "", "skipped": True})
        return result

    timestamp_start = datetime.now().isoformat()
    started = time.monotonic()
    if cli_full_log_path is not None:
        write_external_cli_log_header(
            cli_full_log_path,
            timestamp_start=timestamp_start,
            paper_folder=paper_folder,
            stage=stage,
            attempt=attempt,
            command_template=template,
            command_display=command_display,
            command_args=command_args,
            task_file=task_file,
            prompt_text=prompt_text,
        )
        append_text(cli_full_log_path, "===== BEGIN EXTERNAL CLI OUTPUT =====\n")

    try:
        if stream_output:
            completed_result = run_subprocess_streaming(command_args, prompt_text if stdin_mode else None, log_path=cli_full_log_path)
            result.update(completed_result)
            if cli_full_log_path is not None:
                write_external_cli_log_footer(
                    cli_full_log_path,
                    timestamp_end=datetime.now().isoformat(),
                    returncode=result.get("returncode"),
                    elapsed_seconds=time.monotonic() - started,
                )
            return result
        completed = subprocess.run(
            command_args,
            cwd=PROJECT_ROOT,
            input=prompt_text if stdin_mode else None,
            text=True,
            encoding="utf-8",
            errors="replace",
            capture_output=True,
        )
    except FileNotFoundError as exc:
        result.update({"returncode": None, "stdout_tail": "", "stderr_tail": str(exc), "error": str(exc)})
        if cli_full_log_path is not None:
            append_text(cli_full_log_path, str(exc) + "\n")
            write_external_cli_log_footer(
                cli_full_log_path,
                timestamp_end=datetime.now().isoformat(),
                returncode=None,
                elapsed_seconds=time.monotonic() - started,
            )
        return result

    combined_output = (completed.stdout or "") + (completed.stderr or "")
    if cli_full_log_path is not None:
        append_text(cli_full_log_path, combined_output)
        write_external_cli_log_footer(
            cli_full_log_path,
            timestamp_end=datetime.now().isoformat(),
            returncode=completed.returncode,
            elapsed_seconds=time.monotonic() - started,
        )
    result.update(
        {
            "returncode": completed.returncode,
            "stdout_tail": tail_text(completed.stdout),
            "stderr_tail": tail_text(completed.stderr),
        }
    )
    return result


def run_agent_active(
    paper_folder: Path,
    stages: list[str] | None,
    agent: str,
    command_template: str | None,
    dry_run: bool,
    continue_on_error: bool,
    max_agent_retries: int,
    skip_valid: bool = True,
    stream_agent_output: bool = False,
) -> dict[str, Any]:
    if not paper_folder.exists():
        raise FileNotFoundError(f"Paper folder does not exist: {paper_folder}")
    if not has_manual_marker(paper_folder):
        raise RuntimeError(f"Refusing active agent run: missing {paper_folder / MANUAL_MARKER}")
    if max_agent_retries < 0:
        raise ValueError("--max-agent-retries must be >= 0")

    selected_stages = stages or DEFAULT_AGENT_ACTIVE_STAGES
    unknown = [stage for stage in selected_stages if stage not in STAGES]
    if unknown:
        raise ValueError(f"Unknown stages: {unknown}")

    summary: dict[str, Any] = {
        "status": "completed",
        "paper_folder": str(paper_folder),
        "agent": agent,
        "dry_run": dry_run,
        "stream_agent_output": stream_agent_output,
        "stages": [],
    }
    append_log(
        paper_folder,
        {
            "action": "run_agent_active_start",
            "agent": agent,
            "stages": selected_stages,
            "dry_run": dry_run,
            "continue_on_error": continue_on_error,
            "max_agent_retries": max_agent_retries,
            "skip_valid": skip_valid,
            "stream_agent_output": stream_agent_output,
        },
    )

    for stage in selected_stages:
        ok, messages = validate_stage(stage, paper_folder)
        validation = validation_result_dict(ok, messages)
        mode = STAGE_EXECUTION_MODE.get(stage, "legacy")
        print(f"[RUN_AGENT_ACTIVE] stage={stage} mode={mode}")
        print(f"[VALIDATE] stage={stage}")
        print(f"[VALIDATE] ok={str(ok).lower()} messages={messages}")
        if ok and skip_valid:
            stage_result = {"stage": stage, "status": "skipped_valid", "validation": validation}
            summary["stages"].append(stage_result)
            append_log(paper_folder, {"action": "stage_skip_valid", "stage": stage, "validation": validation})
            update_state(paper_folder, stage, "validated", validation)
            continue

        stage_result: dict[str, Any] = {
            "stage": stage,
            "status": "running",
            "mode": mode,
            "attempts": 0,
            "validation": validation,
        }
        last_agent_result: dict[str, Any] | None = None
        try:
            run_result = run_stage(stage, paper_folder, dry_run=dry_run)
            stage_result["run_result"] = run_result
            if dry_run:
                stage_result["status"] = "planned"
                summary["stages"].append(stage_result)
                continue

            if run_result.get("status") == "external_agent_required":
                task_file = Path(run_result["task_file"])
                stage_result["task_file"] = str(task_file)
                validation_feedback: dict[str, Any] | None = None
                for attempt in range(max_agent_retries + 1):
                    stage_result["attempts"] = attempt + 1
                    preview_template = command_template
                    if not preview_template:
                        if agent == "custom":
                            preview_template = ""
                        else:
                            preview_template = DEFAULT_AGENT_COMMAND_TEMPLATES[agent]
                    preview_prompt = build_agent_prompt(stage, paper_folder, task_file, validation_result=validation_feedback)
                    _preview_args, preview_command, preview_stdin_mode = build_agent_command_args(
                        preview_template,
                        preview_prompt,
                        task_file,
                        paper_folder,
                        stage,
                    )
                    print(f"[EXTERNAL_AGENT] command={preview_command}")
                    print(f"[EXTERNAL_AGENT] task_file={task_file}")
                    print(
                        "[EXTERNAL_AGENT] "
                        f"stdin_mode={str(preview_stdin_mode).lower()} "
                        f"prompt_length={len(preview_prompt)}"
                    )
                    agent_result = run_external_cli_agent(
                        agent,
                        task_file,
                        paper_folder,
                        stage,
                        command_template=command_template,
                        dry_run=dry_run,
                        validation_result=validation_feedback,
                        stream_output=stream_agent_output,
                        attempt=attempt + 1,
                    )
                    last_agent_result = agent_result
                    append_log(paper_folder, {"action": "external_agent_call", **agent_result})
                    if agent_result.get("returncode") not in {0, None}:
                        stage_result["status"] = "agent_failed"
                        stage_result["agent_result"] = agent_result
                    print(f"[VALIDATE] stage={stage}")
                    ok, messages = validate_stage(stage, paper_folder)
                    validation_feedback = validation_result_dict(ok, messages)
                    print(f"[VALIDATE] ok={str(ok).lower()} messages={messages}")
                    stage_result["validation"] = validation_feedback
                    append_log(
                        paper_folder,
                        {"action": "stage_validation", "stage": stage, "task_file": str(task_file), "validation": validation_feedback},
                    )
                    if ok:
                        stage_result["status"] = "completed"
                        stage_result["agent_result"] = agent_result
                        update_state(paper_folder, stage, "validated", validation_feedback)
                        break
                    if attempt < max_agent_retries:
                        append_validation_failure_feedback(task_file, validation_feedback, attempt + 1)
                        append_log(
                            paper_folder,
                            {
                                "action": "stage_retry",
                                "stage": stage,
                                "task_file": str(task_file),
                                "attempt": attempt + 1,
                                "validation": validation_feedback,
                            },
                        )
                else:
                    stage_result["status"] = "validation_failed"
            else:
                print(f"[VALIDATE] stage={stage}")
                ok, messages = validate_stage(stage, paper_folder)
                validation = validation_result_dict(ok, messages)
                print(f"[VALIDATE] ok={str(ok).lower()} messages={messages}")
                stage_result["validation"] = validation
                append_log(paper_folder, {"action": "stage_validation", "stage": stage, "validation": validation})
                if ok:
                    stage_result["status"] = "completed"
                    update_state(paper_folder, stage, "validated", validation)
                else:
                    stage_result["status"] = "validation_failed"

            if stage_result["status"] != "completed":
                failure_detail = {
                    "stage": stage,
                    "status": stage_result["status"],
                    "validation": stage_result.get("validation"),
                    "agent_result": last_agent_result,
                }
                update_state(paper_folder, stage, "failed", failure_detail)
                if not continue_on_error:
                    summary["status"] = "failed"
                    summary["stages"].append(stage_result)
                    append_log(paper_folder, {"action": "run_agent_active_done", **summary})
                    return summary
        except Exception as exc:
            stage_result.update(
                {
                    "status": "failed",
                    "error": str(exc),
                    "traceback": traceback.format_exc(),
                }
            )
            update_state(paper_folder, stage, "failed", stage_result)
            if not continue_on_error:
                summary["status"] = "failed"
                summary["stages"].append(stage_result)
                append_log(paper_folder, {"action": "run_agent_active_done", **summary})
                return summary

        summary["stages"].append(stage_result)

    if any(stage_result.get("status") not in {"completed", "skipped_valid", "planned"} for stage_result in summary["stages"]):
        summary["status"] = "completed_with_errors" if continue_on_error else "failed"
    append_log(paper_folder, {"action": "run_agent_active_done", **summary})
    return summary


def run_stage(
    stage: str,
    paper_folder: Path,
    dry_run: bool = False,
    skip_backup: bool = False,
    mode_override: str | None = None,
) -> dict[str, Any]:
    if stage not in STAGES:
        raise ValueError(f"Unknown stage: {stage}")
    if not paper_folder.exists():
        raise FileNotFoundError(f"Paper folder does not exist: {paper_folder}")
    if STAGES[stage].get("requires_manual_marker") and not has_manual_marker(paper_folder):
        raise RuntimeError(f"Refusing to run {stage}: missing {paper_folder / MANUAL_MARKER}")

    mode = mode_override or STAGE_EXECUTION_MODE.get(stage, "legacy")
    if mode not in VALID_STAGE_EXECUTION_MODES:
        raise ValueError(f"Invalid execution mode for {stage}: {mode}")
    script = PROJECT_ROOT / STAGES[stage]["script"]
    detail = {"stage": stage, "mode": mode, "script": str(script), "dry_run": dry_run}
    append_log(paper_folder, {"action": "stage_start", **detail})
    update_state(paper_folder, stage, "running" if not dry_run else "dry_run", detail)

    if dry_run:
        return {"planned": detail}

    if mode == "external_agent":
        try:
            final = create_external_agent_task(stage, paper_folder)
            append_log(paper_folder, {"action": "stage_external_agent_task_created", "stage": stage, **final})
            update_state(paper_folder, stage, "external_agent_required", final)
            return final
        except Exception as exc:
            final = {
                "status": "failed",
                "mode": mode,
                "error": str(exc),
                "traceback": traceback.format_exc(),
                "backups": [],
            }
            append_log(paper_folder, {"action": "stage_failed", "stage": stage, **final})
            update_state(paper_folder, stage, "failed", final)
            raise

    backups = [] if skip_backup else backup_outputs(stage, paper_folder)
    try:
        if mode == "heuristic":
            result = run_heuristic_stage(stage, paper_folder)
        elif mode == "legacy":
            result = run_legacy_stage(stage, paper_folder)
        else:
            raise ValueError(f"Unsupported execution mode for {stage}: {mode}")
        ok, messages = validate_stage(stage, paper_folder)
        status = "success" if ok else "validation_failed"
        final = {"status": status, "mode": mode, "validation": messages, "backups": backups, "result": result}
        append_log(paper_folder, {"action": "stage_complete", "stage": stage, **final})
        update_state(paper_folder, stage, status, final)
        return final
    except Exception as exc:
        final = {
            "status": "failed",
            "mode": mode,
            "error": str(exc),
            "traceback": traceback.format_exc(),
            "backups": backups,
        }
        append_log(paper_folder, {"action": "stage_failed", "stage": stage, **final})
        update_state(paper_folder, stage, "failed", final)
        raise


def main() -> int:
    parser = argparse.ArgumentParser(description="External CLI agent task runner for LNPDB extraction.")
    sub = parser.add_subparsers(dest="command", required=True)

    for command in ["observe", "next"]:
        p = sub.add_parser(command)
        p.add_argument("--paper-folder", required=True)

    p_inspect_ref = sub.add_parser("inspect-reference")
    p_inspect_ref.add_argument("--paper-folder", required=True)
    p_inspect_ref.add_argument("--output-json")

    p_validate = sub.add_parser("validate")
    p_validate.add_argument("--stage", required=True, choices=STAGE_ORDER)
    p_validate.add_argument("--paper-folder", required=True)

    p_run = sub.add_parser("run")
    p_run.add_argument("--stage", required=True, choices=STAGE_ORDER)
    p_run.add_argument("--paper-folder", required=True)
    p_run.add_argument("--dry-run", action="store_true")
    p_run.add_argument("--skip-backup", action="store_true")
    p_run.add_argument("--mode", choices=sorted(VALID_STAGE_EXECUTION_MODES))

    p_run_agent = sub.add_parser("run-agent-active")
    p_run_agent.add_argument("--paper-folder", required=True)
    p_run_agent.add_argument("--agent", choices=["codex", "claude", "custom"], default="codex")
    p_run_agent.add_argument("--agent-command", default=None)
    p_run_agent.add_argument("--stages", nargs="*", default=None, choices=STAGE_ORDER)
    p_run_agent.add_argument("--dry-run", action="store_true")
    p_run_agent.add_argument("--continue-on-error", action="store_true")
    p_run_agent.add_argument("--max-agent-retries", type=int, default=1)
    p_run_agent.add_argument("--no-skip-valid", action="store_true")
    p_run_agent.add_argument("--stream-agent-output", action="store_true")

    args = parser.parse_args()
    paper_folder = resolve_paper_folder(args.paper_folder)

    if args.command == "observe":
        print(json.dumps(observe(paper_folder), ensure_ascii=False, indent=2))
        return 0

    if args.command == "next":
        result = next_stage(paper_folder)
        append_log(paper_folder, {"action": "next", "result": result})
        update_state(paper_folder, result.get("next_stage"), "next_stage_selected", result)
        print(json.dumps(result, ensure_ascii=False, indent=2))
        return 0

    if args.command == "inspect-reference":
        reference_context = collect_reference_context_for_06(paper_folder)
        if args.output_json:
            write_json(Path(args.output_json), reference_context)
        result = {
            "ok": True,
            "summary": summarize_reference_context(reference_context),
            "output_json": args.output_json or "",
        }
        append_log(paper_folder, {"action": "inspect_reference", "result": result})
        print(json.dumps(result, ensure_ascii=False, indent=2))
        return 0

    if args.command == "validate":
        ok, messages = validate_stage(args.stage, paper_folder)
        result = {"ok": ok, "messages": messages}
        append_log(paper_folder, {"action": "validate", "stage": args.stage, "result": result})
        update_state(paper_folder, args.stage, "validated" if ok else "validation_failed", result)
        print(json.dumps(result, ensure_ascii=False, indent=2))
        return 0 if ok else 2

    if args.command == "run":
        result = run_stage(
            args.stage,
            paper_folder,
            dry_run=args.dry_run,
            skip_backup=args.skip_backup,
            mode_override=args.mode,
        )
        print(json.dumps(result, ensure_ascii=False, indent=2, default=str))
        return 0 if result.get("status") in {None, "success", "external_agent_required"} or "planned" in result else 2

    if args.command == "run-agent-active":
        result = run_agent_active(
            paper_folder=paper_folder,
            stages=args.stages,
            agent=args.agent,
            command_template=args.agent_command,
            dry_run=args.dry_run,
            continue_on_error=args.continue_on_error,
            max_agent_retries=args.max_agent_retries,
            skip_valid=not args.no_skip_valid,
            stream_agent_output=args.stream_agent_output,
        )
        print(json.dumps(result, ensure_ascii=False, indent=2, default=str))
        return 0 if result.get("status") in {"completed", "completed_with_errors"} else 2

    return 1


if __name__ == "__main__":
    raise SystemExit(main())
