import csv
import hashlib
import io
import json
import math
import os
import re
import sys
from pathlib import Path

import pandas as pd
from openpyxl.styles import Font, PatternFill

current_dir = Path(__file__).resolve().parent
parent_dir = current_dir.parent
if str(parent_dir) not in sys.path:
    sys.path.append(str(parent_dir))

from find_api import find_api_key_file, get_vertexai_client
from LLM_Batch import (
    append_batch_request,
    build_batch_request_metadata,
    build_generate_content_batch_request,
    count_requests_in_jsonl,
    create_batch_job_record,
    create_batch_request_file,
    download_batch_results,
    load_batch_results_as_map,
    poll_batch_job,
    submit_batch_job,
    upload_file_to_gcs,
)


DEFAULT_GCS_BATCH_BUCKET = "gs://lnpdb-articles-extraction-batch-results-geon"
TARGET_VISUAL_TYPES = ["barplot", "table", "chemical_structure", "heatmap"]
REQUIRED_RESULT_KEYS = [
    "Excel_File_Name",
    "Excel_Sheet_Name",
    "Block_ID",
    "Block_CSV_Path",
    "Item_ID",
    "Formulation_Name",
    "IL_name",
    "IL_molarratio",
    "HL_name",
    "HL_molarratio",
    "CHL_name",
    "CHL_molarratio",
    "PEG_name",
    "PEG_molarratio",
    "Fifth_component_name",
    "Fifth_component_molarratio",
    "IL_to_nucleicacid_massratio",
    "_reference",
]


def list_pdf_paths(folder: Path) -> list[Path]:
    return sorted(
        [
            f for f in folder.rglob("*.pdf")
            if f.is_file() and not f.name.startswith("~") and "Exp_Excel" not in f.parts
        ]
    )


def upload_pdfs_to_gcs(folder: Path, gcs_batch_bucket: str) -> list[dict]:
    uploaded = []
    for pdf_path in list_pdf_paths(folder):
        gcs_uri = f"{gcs_batch_bucket}/papers/{folder.name}/{pdf_path.name}"
        upload_file_to_gcs(pdf_path, gcs_uri)
        uploaded.append(
            {
                "local_path": str(pdf_path),
                "gcs_uri": gcs_uri,
                "mime_type": "application/pdf",
                "name": pdf_path.name,
            }
        )
    return uploaded


def build_pdf_file_parts(uploaded_pdfs: list[dict]) -> list[dict]:
    parts = []
    for item in uploaded_pdfs:
        gcs_uri = str(item.get("gcs_uri", "")).strip()
        mime_type = str(item.get("mime_type", "application/pdf")).strip() or "application/pdf"
        if not gcs_uri:
            continue
        parts.append({"fileData": {"fileUri": gcs_uri, "mimeType": mime_type}})
    return parts


def get_md_text(folder: Path) -> str:
    texts = []
    for path in folder.rglob("*.md"):
        if path.is_file() and not path.name.startswith("~") and "Exp_Excel" not in path.parts:
            try:
                texts.append(path.read_text(encoding="utf-8"))
            except UnicodeDecodeError:
                texts.append(path.read_text(encoding="utf-8", errors="replace"))
    return "\n\n".join(texts)


def get_block_csv_content(folder: Path, block_csv_path: str, excel_file: str = "", excel_sheet: str = ""):
    block_csv_path = str(block_csv_path or "").strip()

    if block_csv_path:
        csv_path = folder / block_csv_path
        if csv_path.exists():
            try:
                df = pd.read_csv(csv_path, dtype=str).fillna("")
                return df.to_csv(index=False)
            except Exception as e:
                print(f"  ! block csv 로드 실패 ({block_csv_path}): {e}")

    exp_folder = folder / "Exp_Excel"
    file_path = exp_folder / excel_file

    if not file_path.exists():
        return ""

    try:
        if file_path.suffix.lower() == ".csv":
            df = pd.read_csv(file_path, dtype=str).fillna("")
        else:
            xls = pd.ExcelFile(file_path)
            df = pd.read_excel(xls, sheet_name=excel_sheet, dtype=str).fillna("")
        return df.to_csv(index=False)
    except Exception as e:
        print(f"  ! fallback 시트 로드 실패 ({excel_file}[{excel_sheet}]): {e}")
        return ""


def estimate_text_size_for_gemini(text: str) -> dict:
    if text is None:
        text = ""
    char_count = len(text)
    word_count = len(text.split()) if text else 0
    approx_tokens_from_chars = math.ceil(char_count / 4) if char_count else 0
    approx_tokens_from_words = math.ceil(word_count * 1.3) if word_count else 0
    approx_tokens = max(approx_tokens_from_chars, approx_tokens_from_words)
    return {
        "char_count": char_count,
        "word_count": word_count,
        "approx_tokens": approx_tokens,
    }


def is_numeric_like(value: str) -> bool:
    if value is None:
        return False

    s = str(value).strip()
    if s == "":
        return False

    s_lower = s.lower()
    if s_lower in {"nan", "na", "n/a", "none", "null", "inf", "-inf"}:
        return True

    s_clean = s.replace(",", "")
    percent_clean = s_clean[:-1] if s_clean.endswith("%") else s_clean

    sci_pattern = r"^[+-]?(?:\d+(?:\.\d*)?|\.\d+)(?:[eE][+-]?\d+)?$"
    if re.fullmatch(sci_pattern, percent_clean):
        return True

    frac_pattern = r"^[+-]?\d+\s*/\s*[+-]?\d+$"
    if re.fullmatch(frac_pattern, s_clean):
        return True

    return False


def strip_numeric_cells_from_csv(csv_text: str) -> str:
    if not csv_text or not csv_text.strip():
        return csv_text

    try:
        input_io = io.StringIO(csv_text)
        reader = csv.reader(input_io)
        output_io = io.StringIO()
        writer = csv.writer(output_io, lineterminator="\n")

        for row_idx, row in enumerate(reader):
            if row_idx == 0:
                writer.writerow(row)
                continue

            cleaned_row = []
            for cell in row:
                cleaned_row.append("" if is_numeric_like(cell) else cell)
            writer.writerow(cleaned_row)

        cleaned_text = output_io.getvalue().strip()
        return cleaned_text if cleaned_text else csv_text
    except Exception:
        return csv_text


def maybe_count_tokens_with_client(client, model_name, text: str):
    try:
        if hasattr(client, "models") and hasattr(client.models, "count_tokens"):
            result = client.models.count_tokens(model=model_name, contents=[text])
            total_tokens = getattr(result, "total_tokens", None)
            if total_tokens is not None:
                return int(total_tokens)
    except Exception:
        pass
    return None


def prepare_csv_for_prompt(client, model_name, csv_text: str, stage_name: str, soft_token_limit: int = 120000):
    if csv_text is None:
        csv_text = ""

    estimated = estimate_text_size_for_gemini(csv_text)
    counted_tokens = maybe_count_tokens_with_client(client, model_name, csv_text)
    token_basis = counted_tokens if counted_tokens is not None else estimated["approx_tokens"]

    print(
        f"      · [{stage_name}] CSV 크기 점검: chars={estimated['char_count']:,}, "
        f"words={estimated['word_count']:,}, approx_tokens={estimated['approx_tokens']:,}"
        + (f" counted_tokens={counted_tokens:,}" if counted_tokens is not None else "")
    )

    if token_basis <= soft_token_limit:
        return csv_text

    print(f"      · [{stage_name}] CSV가 커서 숫자 셀 제거 버전을 생성합니다.")
    stripped_csv = strip_numeric_cells_from_csv(csv_text)

    stripped_estimated = estimate_text_size_for_gemini(stripped_csv)
    stripped_counted_tokens = maybe_count_tokens_with_client(client, model_name, stripped_csv)
    stripped_basis = stripped_counted_tokens if stripped_counted_tokens is not None else stripped_estimated["approx_tokens"]

    print(
        f"      · [{stage_name}] 숫자 제거 후: chars={stripped_estimated['char_count']:,}, "
        f"words={stripped_estimated['word_count']:,}, approx_tokens={stripped_estimated['approx_tokens']:,}"
        + (f" counted_tokens={stripped_counted_tokens:,}" if stripped_counted_tokens is not None else "")
    )

    if stripped_basis < token_basis:
        return stripped_csv
    return csv_text


def split_item_ids_to_rows(df: pd.DataFrame) -> pd.DataFrame:
    if df is None or df.empty or "Item_ID" not in df.columns:
        return df

    expanded_rows = []
    for _, row in df.iterrows():
        item_value = row.get("Item_ID", "")
        if pd.isna(item_value):
            expanded_rows.append(row.to_dict())
            continue

        item_text = str(item_value).strip()
        if not item_text:
            expanded_rows.append(row.to_dict())
            continue

        split_items = [part.strip() for part in item_text.split(",") if str(part).strip()]
        if not split_items:
            expanded_rows.append(row.to_dict())
            continue

        for single_item in split_items:
            new_row = row.to_dict()
            new_row["Item_ID"] = single_item
            expanded_rows.append(new_row)

    return pd.DataFrame(expanded_rows)


def make_formula_block_custom_id(block_csv_path: str) -> str:
    seed = str(block_csv_path).strip()
    return "formula_block__" + hashlib.sha1(seed.encode("utf-8")).hexdigest()[:12]


def get_fig_selection_value(row):
    manual = str(row.get("manual_select", "")).strip().lower()
    if manual in {"yes", "y", "1", "true", "o"}:
        return "yes"
    if manual in {"no", "n", "0", "false", "x"}:
        return "no"
    return str(row.get("need_for_lnpdb", "")).strip().lower()


def build_formula_block_prompt(
    excel_file: str,
    excel_sheet: str,
    block_id: str,
    block_csv_path: str,
    target_items: list[str],
    csv_content: str,
    request_id: str,
) -> str:
    target_items_text = "\n".join([f"  - {item}" for item in target_items]) if target_items else "  - 없음"
    return f"""
    당신은 지질나노입자(LNP) 제형 및 데이터 매핑 전문가입니다.
    제공된 [엑셀 block 전체 데이터]를 정밀 분석하여, 연관된 실험 그룹의 LNP 조성비를 추출하세요.

    [Request Tracking]
    - request_id: {request_id}

    [분석 대상 block 정보]
    - 파일명: {excel_file}
    - 시트명: {excel_sheet}
    - block_id: {block_id}
    - block_csv_path: {block_csv_path}
    - 이 block와 매핑된 Figure 목록:
{target_items_text}

    [Raw Experimental Data (Block CSV)]
    {csv_content}

    [조성비 추출 절차 및 규칙]
    1. [문맥 파악]: 먼저 표적 데이터를 확인하여, 대개 Figure별 실험에 사용된 LNP 제형 이름(예: C8-200, E4i, MC3 등)을 파악하세요.
    2. [본문 매칭]: 파악한 제형 이름을 바탕으로, MD 텍스트나 PDF를 검색하여 해당 제형의 정확한 지질 성분과 몰비율(Molar ratio)을 찾으세요.
    3. [다중 제형 분리]: block 내에 여러 제형(Formulation)이 비교되고 있다면, 각각을 별도의 JSON 객체로 분리하세요.
    4. [Figure별 강제 분리]: `Item_ID`는 반드시 한 JSON 객체당 정확히 하나의 figure만 가져야 합니다.
       - 절대로 `figure 3b, 3c, 3d`처럼 여러 figure를 한 문자열에 묶지 마세요.
       - Figure가 여러 개라면 figure마다 JSON 객체를 각각 따로 만드세요.
       - 즉, `figure 3b`, `figure 3c`, `figure 3d`는 각각 서로 다른 row가 되도록 별도 객체로 출력해야 합니다.
       - `Item_ID` 값은 반드시 위에 주어진 Figure 목록 중 하나를 그대로 사용하세요.
       - Figure별로 조성 정보가 동일해 보여도 row는 합치지 말고 figure마다 반복해서 따로 출력하세요.
    5. [투입 대비 비율 추출]: 반드시 `IL_to_nucleicacid_massratio`를 최대한 찾으세요.
       - 본문/캡션/본문에 lipid:nucleic acid mass ratio, ionizable lipid:nucleic acid mass ratio, lipid:RNA ratio, lipid:mRNA ratio, wt/wt ratio 등이 직접 제시되면 그 값을 `IL_to_nucleicacid_massratio`에 쓰세요.
       - 정확한 mass ratio 정보가 없으면 비슷한 개념의 값을 대신 써도 됩니다. 예: N/P ratio, charge ratio, (+/-) ratio.
       - 그 경우 값 안에 반드시 괄호로 출처 종류를 붙이세요.
         예: `6 (N/P ratio)`, `3 (charge ratio)`, `10:1 (lipid:mRNA mass ratio)`
       - 정확한 mass ratio가 있으면 괄호 없이 값만 쓰세요. 예: `10:1`
       - 관련 정보가 전혀 없으면 빈 문자열 `""`로 쓰세요.
    6. [근거 우선]: `_reference`에는 해당 조성비 또는 투입 대비 비율의 근거가 되는 직접 문장을 쓰세요.
    7. 최상위 JSON에 `request_id`를 반드시 그대로 복사하세요.
    8. Fifth_component_name, Fifth_component_molarratio
        - 기본 4성분(IL, HL, CHL, PEG) 외에 추가 성분이 존재하면 기록하세요.
        - 5번째 성분이 없으면 반드시 "N/A"로 기입하세요.
        - 빈 문자열 "", null, 생략은 허용하지 않습니다.

    반드시 아래 구조의 JSON 객체만 반환하세요.
    {{
      "request_id": "{request_id}",
      "results": [
        {{
          "Excel_File_Name": "{excel_file}",
          "Excel_Sheet_Name": "{excel_sheet}",
          "Block_ID": "{block_id}",
          "Block_CSV_Path": "{block_csv_path}",
          "Item_ID": "figure 2a",
          "Formulation_Name": "E4i-200",
          "IL_name": "E4i",
          "IL_molarratio": "50",
          "HL_name": "DOPE",
          "HL_molarratio": "10",
          "CHL_name": "Cholesterol",
          "CHL_molarratio": "38.5",
          "PEG_name": "DMG-PEG2000",
          "PEG_molarratio": "1.5",
          "Fifth_component_name": "N/A",
          "Fifth_component_molarratio": "N/A",
          "IL_to_nucleicacid_massratio": "6 (N/P ratio)",
          "_reference": "LNP formulations were prepared at a molar ratio..."
        }}
      ]
    }}
    """


def validate_formula_block_results(results, expected_items: list[str]):
    if not isinstance(results, list) or not results:
        raise ValueError("missing_results")

    normalized_expected = {str(x).strip().lower() for x in expected_items if str(x).strip()}
    for idx, row in enumerate(results):
        if not isinstance(row, dict):
            raise ValueError(f"result_not_dict::{idx}")
        missing_keys = [key for key in REQUIRED_RESULT_KEYS if key not in row]
        if missing_keys:
            raise ValueError(f"missing_keys::{idx}::{','.join(missing_keys)}")
        item_id = str(row.get("Item_ID", "")).strip()
        if not item_id:
            raise ValueError(f"missing_item_id::{idx}")
        if normalized_expected and item_id.lower() not in normalized_expected:
            raise ValueError(f"unexpected_item_id::{idx}::{item_id}")


def run_formula_block_batch(
    folder: Path,
    client,
    model_name: str,
    block_to_items: dict,
    block_to_meta: dict,
    shared_contents: list,
    task_suffix: str = "",
):
    request_file = create_batch_request_file(folder, f"formula_block_{folder.name}{task_suffix}")
    request_count = 0

    for block_csv_path, items in block_to_items.items():
        meta = block_to_meta.get(block_csv_path, {})
        excel_file = meta.get("excel_file", "")
        excel_sheet = meta.get("excel_sheet", "")
        block_id = meta.get("block_id", Path(block_csv_path).stem)
        full_csv_content = get_block_csv_content(folder, block_csv_path, excel_file, excel_sheet)
        if not full_csv_content:
            continue

        current_csv = prepare_csv_for_prompt(client, model_name, full_csv_content, stage_name="formula_sheet")
        custom_id = make_formula_block_custom_id(block_csv_path)
        prompt = build_formula_block_prompt(
            excel_file=excel_file,
            excel_sheet=excel_sheet,
            block_id=block_id,
            block_csv_path=block_csv_path,
            target_items=items,
            csv_content=current_csv,
            request_id=custom_id,
        )

        metadata = build_batch_request_metadata(
            task_name="extract_formula_by_block",
            model_name=model_name,
            custom_id=custom_id,
            stage_name="extract_formula_by_block",
            item_id=block_csv_path,
            paper_folder=str(folder),
            extra_metadata={
                "excel_file": excel_file,
                "excel_sheet": excel_sheet,
                "block_id": block_id,
                "block_csv_path": block_csv_path,
            },
        )
        request_body = build_generate_content_batch_request(
            model_name=model_name,
            contents=list(shared_contents),
            prompt_text=prompt,
            response_mime_type="application/json",
        )
        append_batch_request(request_file, custom_id, request_body, metadata)
        request_count += 1

    if request_count == 0 or count_requests_in_jsonl(request_file) == 0:
        return {}

    local_job_id = create_batch_job_record(
        paper_folder=folder,
        task_name="extract_formula_by_block",
        model_name=model_name,
        request_file=request_file,
        metadata={
            "request_count": count_requests_in_jsonl(request_file),
            "gcs_input_uri": f"{DEFAULT_GCS_BATCH_BUCKET}/batch/{request_file.name}",
            "gcs_output_uri_prefix": f"{DEFAULT_GCS_BATCH_BUCKET}/batch_output/{request_file.stem}",
        },
    )

    batch_job = submit_batch_job(
        client=client,
        paper_folder=folder,
        local_job_id=local_job_id,
        display_name=f"formula-block-{folder.name}{task_suffix}",
    )
    print(f"  · block batch 제출 완료: {batch_job.name}")

    finished_job = poll_batch_job(client, folder, local_job_id, poll_interval_seconds=30)
    state_name = getattr(getattr(finished_job, "state", None), "name", None) or str(getattr(finished_job, "state", "UNKNOWN"))
    print(f"  · block batch 종료 상태: {state_name}")
    if state_name != "JOB_STATE_SUCCEEDED":
        raise RuntimeError(f"extract_formula_by_block batch 실패: {state_name}")

    result_file = download_batch_results(client, folder, local_job_id)
    return load_batch_results_as_map(result_file)


def consume_formula_block_batch_results(batch_results_map, block_paths, block_to_items, block_to_meta, model_name: str):
    all_final_data = []
    usage_rows = []
    failed_blocks = []

    for block_csv_path in block_paths:
        custom_id = make_formula_block_custom_id(block_csv_path)
        row = batch_results_map.get(custom_id)
        meta_info = block_to_meta.get(block_csv_path, {})

        if not row:
            print(f"  ! batch 결과 누락: {block_csv_path}")
            failed_blocks.append(block_csv_path)
            continue
        if not row.get("success"):
            print(f"  ! batch success=false: {block_csv_path} | {row.get('error')}")
            failed_blocks.append(block_csv_path)
            continue

        response_text = str(row.get("response_text", "")).strip()
        if not response_text:
            print(f"  ! batch 빈 응답: {block_csv_path}")
            failed_blocks.append(block_csv_path)
            continue

        clean_text = response_text.replace("```json", "").replace("```", "").strip()
        try:
            payload = json.loads(clean_text)
            if not isinstance(payload, dict):
                raise ValueError("response_not_dict")

            response_request_id = str(payload.get("request_id", "")).strip()
            if not response_request_id:
                raise ValueError("missing_request_id")
            if response_request_id != custom_id:
                raise ValueError(f"request_id_mismatch::{response_request_id}")

            results = payload.get("results")
            validate_formula_block_results(results, block_to_items.get(block_csv_path, []))
            all_final_data.extend(results)

            cost_info = row.get("cost_info") or {}
            usage_rows.append(
                {
                    "block_csv_path": block_csv_path,
                    "task_name": "extract_formula_by_block",
                    "model_name": model_name,
                    "input_tokens": row.get("input_tokens"),
                    "output_tokens": row.get("output_tokens"),
                    "total_tokens": row.get("total_tokens"),
                    "billed_output_tokens": row.get("billed_output_tokens"),
                    "total_cost_usd": cost_info.get("total_cost_usd", 0.0),
                    "excel_file": meta_info.get("excel_file", ""),
                    "excel_sheet": meta_info.get("excel_sheet", ""),
                    "block_id": meta_info.get("block_id", ""),
                }
            )
        except Exception as e:
            print(f"  ! batch 파싱/검증 실패 ({block_csv_path}): {e}")
            failed_blocks.append(block_csv_path)

    return all_final_data, usage_rows, failed_blocks


def fig_main(paper_folder_str, model_name, api_json_name="vertex.json"):
    paper_folder = Path(paper_folder_str)
    print(f"🔬 [조성비 표 기반 추출] 폴더 분석 시작: {paper_folder.name}")

    csv_path = paper_folder / "fig_table_lnpdb_classified.csv"
    if not csv_path.exists():
        return

    df = pd.read_csv(csv_path, low_memory=False)
    df["_fig_select"] = df.apply(get_fig_selection_value, axis=1)
    df["_visual_type_norm"] = (
        df["visual_type"]
        .astype(str)
        .str.strip()
        .str.lower()
        .replace({"raw_data_table": "table"})
    )

    valid_items = set(
        df[
            (df["_fig_select"].isin(["yes", "maybe"]))
            & (df["_visual_type_norm"].isin(TARGET_VISUAL_TYPES))
        ]["item_id"].astype(str).str.lower().str.strip().tolist()
    )

    print(f"  - valid_items 개수: {len(valid_items)}")
    if valid_items:
        print("  - valid_items 예시:")
        for x in list(sorted(valid_items))[:30]:
            print(f"    * {x}")

    map_path = paper_folder / "excel_mapping.json"
    if not map_path.exists():
        return
    with open(map_path, "r", encoding="utf-8") as f:
        excel_mapping = json.load(f)

    classified_block_map = {}
    if "matched_block_csv_path" in df.columns:
        for _, row in df.iterrows():
            item_id = str(row.get("item_id", "")).lower().strip()
            if item_id not in valid_items:
                continue
            raw_paths = str(row.get("matched_block_csv_path", "")).strip()
            if not raw_paths:
                continue
            for p in [x.strip() for x in raw_paths.split(" | ") if str(x).strip()]:
                classified_block_map.setdefault(p, [])
                if item_id not in classified_block_map[p]:
                    classified_block_map[p].append(item_id)

    block_to_items = {}
    block_to_meta = {}
    for item_id, mappings in excel_mapping.items():
        clean_item_id = item_id.lower().strip()
        if clean_item_id in valid_items:
            for m in mappings:
                block_csv_path = str(m.get("block_csv_path", "")).strip()
                if not block_csv_path:
                    continue
                excel_file = str(m.get("excel_file", "")).strip()
                excel_sheet = str(m.get("excel_sheet", "")).strip()
                block_id = str(m.get("block_id", "")).strip()
                block_to_items.setdefault(block_csv_path, [])
                if clean_item_id not in block_to_items[block_csv_path]:
                    block_to_items[block_csv_path].append(clean_item_id)
                if block_csv_path not in block_to_meta:
                    block_to_meta[block_csv_path] = {
                        "excel_file": excel_file,
                        "excel_sheet": excel_sheet,
                        "block_id": block_id,
                        "block_csv_path": block_csv_path,
                    }

    for block_csv_path, items in classified_block_map.items():
        block_to_items.setdefault(block_csv_path, [])
        for item in items:
            if item not in block_to_items[block_csv_path]:
                block_to_items[block_csv_path].append(item)
        if block_csv_path not in block_to_meta:
            guessed_file = ""
            guessed_sheet = ""
            guessed_block_id = Path(block_csv_path).stem
            parts = Path(block_csv_path).parts
            if len(parts) >= 3:
                guessed_sheet = parts[-2]
                guessed_file = f"{parts[-3]}.xlsx"
            block_to_meta[block_csv_path] = {
                "excel_file": guessed_file,
                "excel_sheet": guessed_sheet,
                "block_id": guessed_block_id,
                "block_csv_path": block_csv_path,
            }

    print(f"  - 실제 유효 formula block CSV 개수: {len(block_to_items)}")
    if block_to_items:
        print("  - 실제 유효 formula block CSV 목록:")
        for p in sorted(block_to_items.keys()):
            print(f"    * {p}")

    if not block_to_items:
        return

    api_key_path = find_api_key_file(api_json_name)
    with open(api_key_path, "r", encoding="utf-8") as f:
        cred_data = json.load(f)

    project_id = cred_data.get("project_id")
    if not project_id:
        raise ValueError(f"서비스 계정 JSON에 project_id가 없습니다: {api_key_path}")

    print(f"🔧 Vertex 프로젝트 설정: {project_id}")
    client = get_vertexai_client(api_key_path, project=project_id)

    print("  - PDF 문서 업로드 및 markdown 텍스트 수집 중..")
    uploaded_pdfs = upload_pdfs_to_gcs(paper_folder, DEFAULT_GCS_BATCH_BUCKET)
    shared_contents = build_pdf_file_parts(uploaded_pdfs)
    md_text = get_md_text(paper_folder)
    if md_text.strip():
        shared_contents.append(md_text)

    batch_results_map = run_formula_block_batch(
        folder=paper_folder,
        client=client,
        model_name=model_name,
        block_to_items=block_to_items,
        block_to_meta=block_to_meta,
        shared_contents=shared_contents,
    )

    all_results, usage_rows, failed_blocks = consume_formula_block_batch_results(
        batch_results_map=batch_results_map,
        block_paths=list(block_to_items.keys()),
        block_to_items=block_to_items,
        block_to_meta=block_to_meta,
        model_name=model_name,
    )

    if failed_blocks:
        print(f"\n🔁 실패한 {len(failed_blocks)}개 block만 batch 재시도합니다..")
        retry_block_to_items = {k: block_to_items[k] for k in failed_blocks if k in block_to_items}
        retry_block_to_meta = {k: block_to_meta[k] for k in failed_blocks if k in block_to_meta}
        retry_results_map = run_formula_block_batch(
            folder=paper_folder,
            client=client,
            model_name=model_name,
            block_to_items=retry_block_to_items,
            block_to_meta=retry_block_to_meta,
            shared_contents=shared_contents,
            task_suffix="_retry",
        )
        retry_data, retry_usage_rows, retry_failed_blocks = consume_formula_block_batch_results(
            batch_results_map=retry_results_map,
            block_paths=failed_blocks,
            block_to_items=block_to_items,
            block_to_meta=block_to_meta,
            model_name=model_name,
        )
        all_results.extend(retry_data)
        usage_rows.extend(retry_usage_rows)
        if retry_failed_blocks:
            print(f"  ! 재시도 후에도 실패한 block {len(retry_failed_blocks)}개")

    if all_results:
        final_df = pd.DataFrame(all_results)
        final_df = split_item_ids_to_rows(final_df)
        if "Item_ID" in final_df.columns:
            final_df["Item_ID"] = final_df["Item_ID"].astype(str).str.strip()
        if "Excel_File_Name" in final_df.columns:
            sort_cols = [c for c in ["Excel_File_Name", "Excel_Sheet_Name", "Block_ID", "Item_ID"] if c in final_df.columns]
            final_df.sort_values(by=sort_cols, inplace=True)

        output_xlsx = paper_folder / f"3_{paper_folder.name}_formula_sheetbase.xlsx"

        with pd.ExcelWriter(output_xlsx, engine="openpyxl") as writer:
            final_df.to_excel(writer, index=False, sheet_name="Formula_Data")
            ws = writer.sheets["Formula_Data"]
            highlight = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")
            bold_font = Font(bold=True)
            target_cols = [
                c
                for c in final_df.columns
                if any(x in str(c) for x in ["name", "molarratio", "Formulation_Name", "IL_to_nucleicacid_massratio"])
            ]
            preferred_order = ["Excel_File_Name", "Excel_Sheet_Name", "Block_ID", "Block_CSV_Path", "Item_ID"]
            other_cols = [c for c in final_df.columns if c not in preferred_order]
            final_df = final_df[[c for c in preferred_order if c in final_df.columns] + other_cols]
            ws.delete_rows(1, ws.max_row)
            for r_idx, row_vals in enumerate([list(final_df.columns)] + final_df.astype(str).fillna("").values.tolist(), start=1):
                for c_idx, val in enumerate(row_vals, start=1):
                    ws.cell(row=r_idx, column=c_idx, value=val)
            for col_name in target_cols:
                if col_name in final_df.columns:
                    col_idx = final_df.columns.get_loc(col_name) + 1
                    for row in range(1, len(final_df) + 2):
                        cell = ws.cell(row=row, column=col_idx)
                        cell.fill = highlight
                        if row > 1:
                            cell.font = bold_font

        print(f"\n✅ 시트 기반 조성비 추출 완료: {output_xlsx}")

    if usage_rows:
        usage_csv_path = paper_folder / "30_formula_excel_batch_usage.csv"
        pd.DataFrame(usage_rows).to_csv(usage_csv_path, index=False, encoding="utf-8-sig")


if __name__ == "__main__":
    TEST_DIR = r"C:\Users\kogun\PycharmProjects\LNPDB_Articles_Extraction\Extraction_Examples\excel_o"
    model_name = "gemini-3.1-pro-preview"
    api_json_name = "vertex.json"
    try:
        fig_main(TEST_DIR, model_name, api_json_name)
    except Exception as e:
        print(f"오류: {e}")
