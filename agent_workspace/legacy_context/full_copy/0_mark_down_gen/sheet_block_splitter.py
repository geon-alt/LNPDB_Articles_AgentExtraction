from pathlib import Path

import pandas as pd
from openpyxl import load_workbook


def normalize_cell_text(v) -> str:
    if v is None:
        return ""
    s = str(v).strip()
    if s.lower() in {"nan", "none", "null"}:
        return ""
    return s


def is_numeric_like(value: str) -> bool:
    if value is None:
        return False
    s = str(value).strip()
    if s == "":
        return False
    s_lower = s.lower()
    if s_lower in {"nan", "na", "n/a", "none", "null", "inf", "-inf"}:
        return True
    s_clean = s.replace(",", "")
    percent_clean = s_clean[:-1] if s_clean.endswith("%") else s_clean
    import re
    sci_pattern = r'^[+-]?(?:\d+(?:\.\d*)?|\.\d+)(?:[eE][+-]?\d+)?$'
    if re.fullmatch(sci_pattern, percent_clean):
        return True
    frac_pattern = r'^[+-]?\d+\s*/\s*[+-]?\d+$'
    if re.fullmatch(frac_pattern, s_clean):
        return True
    return False


def load_sheet_df_and_ws(excel_path: Path, sheet_name: str):
    suffix = excel_path.suffix.lower()
    if suffix == ".csv":
        df = pd.read_csv(excel_path, dtype=str).fillna("")
        return df, None
    df = pd.read_excel(excel_path, sheet_name=sheet_name, dtype=str).fillna("")
    wb = load_workbook(excel_path, data_only=True)
    if sheet_name not in wb.sheetnames:
        raise ValueError(f"시트를 찾을 수 없습니다: {sheet_name}")
    ws = wb[sheet_name]
    return df, ws


def build_ws_grid(ws):
    max_row = ws.max_row
    max_col = ws.max_column
    merged_map = {}
    for mr in ws.merged_cells.ranges:
        min_col, min_row, max_col_r, max_row_r = mr.bounds
        anchor_val = ws.cell(min_row, min_col).value
        for r in range(min_row, max_row_r + 1):
            for c in range(min_col, max_col_r + 1):
                merged_map[(r, c)] = {
                    "anchor": (min_row, min_col),
                    "range": (min_row, max_row_r, min_col, max_col_r),
                    "value": anchor_val,
                }
    cells = {}
    for r in range(1, max_row + 1):
        for c in range(1, max_col + 1):
            cell = ws.cell(r, c)
            raw_value = cell.value
            merged_info = merged_map.get((r, c))
            display_value = merged_info["value"] if merged_info else raw_value
            display_text = normalize_cell_text(display_value)
            border = cell.border
            fill = cell.fill
            font = cell.font
            has_border = False
            if border:
                for side in [border.left, border.right, border.top, border.bottom]:
                    if side and side.style not in (None, ""):
                        has_border = True
                        break
            has_fill = False
            if fill and getattr(fill, "fill_type", None):
                if fill.fill_type not in (None, "", "none"):
                    has_fill = True
            is_bold = bool(font and getattr(font, "bold", False))
            is_merged = merged_info is not None
            is_numeric = is_numeric_like(display_text)
            occupied = bool(display_text)
            cells[(r, c)] = {
                "row": r,
                "col": c,
                "value": raw_value,
                "display_value": display_text,
                "is_empty": (display_text == ""),
                "is_numeric": is_numeric,
                "is_merged": is_merged,
                "merged_range": merged_info["range"] if merged_info else None,
                "has_border": has_border,
                "has_fill": has_fill,
                "is_bold": is_bold,
                "occupied": occupied,
            }
    return {"cells": cells, "max_row": max_row, "max_col": max_col}



def build_occupancy_mask(grid):
    max_row = grid["max_row"]
    max_col = grid["max_col"]
    mask = [[0] * (max_col + 1) for _ in range(max_row + 1)]
    for (r, c), info in grid["cells"].items():
        if info["occupied"]:
            mask[r][c] = 1
    return mask


def expand_component_coords(grid, coords):
    expanded = set(coords)
    for r, c in list(expanded):
        info = grid["cells"].get((r, c)) or {}
        merged_range = info.get("merged_range")
        if not merged_range:
            continue
        mr1, mr2, mc1, mc2 = merged_range
        for rr in range(mr1, mr2 + 1):
            for cc in range(mc1, mc2 + 1):
                expanded.add((rr, cc))
    return expanded


def extract_component_df_from_ws(ws, grid, coords):
    if not coords:
        return pd.DataFrame(), None

    expanded = expand_component_coords(grid, coords)
    bbox = trim_empty_edges(grid, bbox_from_coords(list(expanded)))
    if bbox["r1"] > bbox["r2"] or bbox["c1"] > bbox["c2"]:
        return pd.DataFrame(), bbox

    component_cells = set(expanded)
    full_df = extract_block_df_from_ws(ws, bbox).fillna("")
    rows = bbox["r2"] - bbox["r1"] + 1
    cols = bbox["c2"] - bbox["c1"] + 1
    canvas = pd.DataFrame("", index=range(rows), columns=range(cols))

    for abs_r, abs_c in component_cells:
        if not (bbox["r1"] <= abs_r <= bbox["r2"] and bbox["c1"] <= abs_c <= bbox["c2"]):
            continue
        local_r = abs_r - bbox["r1"]
        local_c = abs_c - bbox["c1"]
        if local_r < full_df.shape[0] and local_c < full_df.shape[1]:
            canvas.iat[local_r, local_c] = full_df.iat[local_r, local_c]

    return canvas, bbox


def build_numeric_mask(grid, bbox=None):
    max_row = grid["max_row"]
    max_col = grid["max_col"]
    mask = [[0] * (max_col + 1) for _ in range(max_row + 1)]
    if bbox is None:
        r1, r2, c1, c2 = 1, max_row, 1, max_col
    else:
        r1, r2, c1, c2 = bbox["r1"], bbox["r2"], bbox["c1"], bbox["c2"]
    for r in range(r1, r2 + 1):
        for c in range(c1, c2 + 1):
            info = grid["cells"][(r, c)]
            if info["is_numeric"]:
                mask[r][c] = 1
    return mask


def count_empty_rows_between(grid, upper_row: int, lower_row: int, c1: int, c2: int) -> int:
    if lower_row <= upper_row + 1:
        return 0
    gap = 0
    for rr in range(upper_row + 1, lower_row):
        has_any = any(grid["cells"][(rr, cc)]["occupied"] for cc in range(c1, c2 + 1))
        if has_any:
            break
        gap += 1
    return gap


def count_empty_cols_between(grid, left_col: int, right_col: int, r1: int, r2: int) -> int:
    if right_col <= left_col + 1:
        return 0
    gap = 0
    for cc in range(left_col + 1, right_col):
        has_any = any(grid["cells"][(rr, cc)]["occupied"] for rr in range(r1, r2 + 1))
        if has_any:
            break
        gap += 1
    return gap


def find_numeric_body_bbox(grid, bbox):
    numeric_mask = build_numeric_mask(grid, bbox)
    comps = connected_components(numeric_mask, grid["max_row"], grid["max_col"])
    inside = []
    for comp in comps:
        comp = [(r, c) for r, c in comp if bbox["r1"] <= r <= bbox["r2"] and bbox["c1"] <= c <= bbox["c2"]]
        if not comp:
            continue
        comp_bbox = bbox_from_coords(comp)
        n_rows = comp_bbox["r2"] - comp_bbox["r1"] + 1
        n_cols = comp_bbox["c2"] - comp_bbox["c1"] + 1
        score = len(comp)
        inside.append((score, n_rows, n_cols, comp_bbox))
    if not inside:
        return None
    inside.sort(key=lambda x: (x[0], x[1], x[2]), reverse=True)
    return inside[0][3]


def compute_relative_numeric_body_features(bbox, numeric_body_bbox):
    if not numeric_body_bbox:
        return {
            "has_numeric_body": False,
            "numeric_body_bbox": None,
            "title_to_body_gap_rows": None,
            "left_body_margin_cols": None,
        }
    return {
        "has_numeric_body": True,
        "numeric_body_bbox": numeric_body_bbox,
        "title_to_body_gap_rows": max(0, numeric_body_bbox["r1"] - bbox["r1"]),
        "left_body_margin_cols": max(0, numeric_body_bbox["c1"] - bbox["c1"]),
    }


def dilate_mask(mask, max_row, max_col, row_radius=0, col_radius=1):
    out = [[0] * (max_col + 1) for _ in range(max_row + 1)]
    for r in range(1, max_row + 1):
        for c in range(1, max_col + 1):
            if mask[r][c] != 1:
                continue
            for rr in range(max(1, r - row_radius), min(max_row, r + row_radius) + 1):
                for cc in range(max(1, c - col_radius), min(max_col, c + col_radius) + 1):
                    out[rr][cc] = 1
    return out


def connected_components(mask, max_row, max_col):
    visited = [[False] * (max_col + 1) for _ in range(max_row + 1)]
    components = []
    for r in range(1, max_row + 1):
        for c in range(1, max_col + 1):
            if mask[r][c] != 1 or visited[r][c]:
                continue
            stack = [(r, c)]
            visited[r][c] = True
            coords = []
            while stack:
                cr, cc = stack.pop()
                coords.append((cr, cc))
                for nr, nc in ((cr - 1, cc), (cr + 1, cc), (cr, cc - 1), (cr, cc + 1)):
                    if 1 <= nr <= max_row and 1 <= nc <= max_col and mask[nr][nc] == 1 and not visited[nr][nc]:
                        visited[nr][nc] = True
                        stack.append((nr, nc))
            components.append(coords)
    return components


def bbox_from_coords(coords):
    rows = [r for r, _ in coords]
    cols = [c for _, c in coords]
    return {"r1": min(rows), "r2": max(rows), "c1": min(cols), "c2": max(cols)}


# --- Start containment helpers ---

def is_bbox_contained(inner_bbox, outer_bbox):
    return (
        outer_bbox["r1"] <= inner_bbox["r1"] <= inner_bbox["r2"] <= outer_bbox["r2"]
        and outer_bbox["c1"] <= inner_bbox["c1"] <= inner_bbox["c2"] <= outer_bbox["c2"]
    )


def merge_contained_raw_elements(raw_elements):
    if not raw_elements:
        return []

    ordered = sorted(
        raw_elements,
        key=lambda x: (
            -((x["bbox"]["r2"] - x["bbox"]["r1"] + 1) * (x["bbox"]["c2"] - x["bbox"]["c1"] + 1)),
            x["bbox"]["r1"],
            x["bbox"]["c1"],
        )
    )

    kept = []
    for elem in ordered:
        contained_in_existing = False
        for parent in kept:
            if is_bbox_contained(elem["bbox"], parent["bbox"]):
                parent.setdefault("contained_element_ids", [])
                if elem.get("element_id"):
                    parent["contained_element_ids"].append(elem.get("element_id"))
                contained_in_existing = True
                break
        if not contained_in_existing:
            elem.setdefault("contained_element_ids", [])
            kept.append(elem)

    return sorted(kept, key=lambda x: (x["bbox"]["r1"], x["bbox"]["c1"]))

# --- End containment helpers ---


def trim_empty_edges(grid, bbox):
    r1, r2, c1, c2 = bbox["r1"], bbox["r2"], bbox["c1"], bbox["c2"]
    def row_has_content(r):
        return any(grid["cells"][(r, c)]["occupied"] for c in range(c1, c2 + 1))
    def col_has_content(c):
        return any(grid["cells"][(r, c)]["occupied"] for r in range(r1, r2 + 1))
    while r1 <= r2 and not row_has_content(r1):
        r1 += 1
    while r2 >= r1 and not row_has_content(r2):
        r2 -= 1
    while c1 <= c2 and not col_has_content(c1):
        c1 += 1
    while c2 >= c1 and not col_has_content(c2):
        c2 -= 1
    return {"r1": r1, "r2": r2, "c1": c1, "c2": c2}


def expand_block_context(grid, bbox, up_rows=2, left_cols=1):
    max_row = grid["max_row"]
    max_col = grid["max_col"]
    r1, r2, c1, c2 = bbox["r1"], bbox["r2"], bbox["c1"], bbox["c2"]
    for rr in range(max(1, r1 - up_rows), r1):
        row_text_cells = row_bold_cells = row_merged_cells = overlap_cols = 0
        for cc in range(c1, c2 + 1):
            info = grid["cells"][(rr, cc)]
            if info["display_value"]:
                row_text_cells += 1
            if info["is_bold"]:
                row_bold_cells += 1
            if info["is_merged"]:
                row_merged_cells += 1
            if info["occupied"]:
                overlap_cols += 1
        if row_text_cells > 0 and (row_bold_cells > 0 or row_merged_cells > 0 or overlap_cols >= max(1, (c2 - c1 + 1) // 3)):
            r1 = rr
    for cc in range(max(1, c1 - left_cols), c1):
        col_text_cells = 0
        for rr in range(r1, r2 + 1):
            if grid["cells"][(rr, cc)]["display_value"]:
                col_text_cells += 1
        if col_text_cells > 0:
            c1 = cc
    return {"r1": max(1, r1), "r2": min(max_row, r2), "c1": max(1, c1), "c2": min(max_col, c2)}


def extract_block_df_from_ws(ws, bbox):
    r1, r2, c1, c2 = bbox["r1"], bbox["r2"], bbox["c1"], bbox["c2"]
    merged_map = {}
    for mr in ws.merged_cells.ranges:
        min_col, min_row, max_col_r, max_row_r = mr.bounds
        anchor_val = ws.cell(min_row, min_col).value
        for r in range(min_row, max_row_r + 1):
            for c in range(min_col, max_col_r + 1):
                merged_map[(r, c)] = anchor_val
    rows = []
    for r in range(r1, r2 + 1):
        row_vals = []
        for c in range(c1, c2 + 1):
            v = merged_map[(r, c)] if (r, c) in merged_map else ws.cell(r, c).value
            row_vals.append("" if v is None else str(v))
        rows.append(row_vals)
    return pd.DataFrame(rows).fillna("")


def infer_block_title(ws, bbox):
    r1, r2, c1, c2 = bbox["r1"], bbox["r2"], bbox["c1"], bbox["c2"]
    title_candidates = []
    for rr in range(r1, min(r1 + 2, r2) + 1):
        texts = []
        for cc in range(c1, c2 + 1):
            s = normalize_cell_text(ws.cell(rr, cc).value)
            if s:
                texts.append(s)
        if texts:
            title_candidates.append(" | ".join(texts[:3]))
    return title_candidates[0][:120] if title_candidates else f"R{r1}_C{c1}"


def compute_block_features(grid, bbox):
    r1, r2, c1, c2 = bbox["r1"], bbox["r2"], bbox["c1"], bbox["c2"]
    n_rows = r2 - r1 + 1
    n_cols = c2 - c1 + 1
    total = max(1, n_rows * n_cols)
    nonempty = numeric = text_cells = border_cells = fill_cells = bold_cells = merged_cells = 0
    row_nonempty_counts = []
    col_nonempty_counts = []
    for r in range(r1, r2 + 1):
        row_count = 0
        for c in range(c1, c2 + 1):
            info = grid["cells"][(r, c)]
            if info["occupied"]:
                nonempty += 1
                row_count += 1
            if info["is_numeric"]:
                numeric += 1
            if info["display_value"] and not info["is_numeric"]:
                text_cells += 1
            if info["has_border"]:
                border_cells += 1
            if info["has_fill"]:
                fill_cells += 1
            if info["is_bold"]:
                bold_cells += 1
            if info["is_merged"]:
                merged_cells += 1
        row_nonempty_counts.append(row_count)
    for c in range(c1, c2 + 1):
        col_count = 0
        for r in range(r1, r2 + 1):
            if grid["cells"][(r, c)]["occupied"]:
                col_count += 1
        col_nonempty_counts.append(col_count)
    empty_rows_inside = sum(1 for x in row_nonempty_counts if x == 0)
    empty_cols_inside = sum(1 for x in col_nonempty_counts if x == 0)
    header_like_rows = 0
    for r in range(r1, min(r1 + 3, r2) + 1):
        txt = num = bold = 0
        for c in range(c1, c2 + 1):
            info = grid["cells"][(r, c)]
            if info["display_value"] and not info["is_numeric"]:
                txt += 1
            if info["is_numeric"]:
                num += 1
            if info["is_bold"]:
                bold += 1
        if txt >= 1 and (bold >= 1 or txt >= num):
            header_like_rows += 1
    title_like_top_row = False
    top_text = top_merged = top_bold = 0
    for c in range(c1, c2 + 1):
        info = grid["cells"][(r1, c)]
        if info["display_value"]:
            top_text += 1
        if info["is_merged"]:
            top_merged += 1
        if info["is_bold"]:
            top_bold += 1
    title_like_top_row = (top_text > 0 and (top_merged > 0 or top_bold > 0))

    numeric_body_bbox = find_numeric_body_bbox(grid, bbox)
    numeric_body_features = compute_relative_numeric_body_features(bbox, numeric_body_bbox)

    return {
        "n_rows": n_rows, "n_cols": n_cols,
        "nonempty_ratio": round(nonempty / total, 4),
        "numeric_ratio": round(numeric / total, 4),
        "text_ratio": round(text_cells / total, 4),
        "border_ratio": round(border_cells / total, 4),
        "fill_ratio": round(fill_cells / total, 4),
        "bold_ratio": round(bold_cells / total, 4),
        "merged_ratio": round(merged_cells / total, 4),
        "empty_rows_inside": empty_rows_inside,
        "empty_cols_inside": empty_cols_inside,
        "header_like_rows": header_like_rows,
        "title_like_top_row": title_like_top_row,
        **numeric_body_features,
    }


def classify_block_route(block_obj):
    f = block_obj["features"]
    if f["n_rows"] < 1 or f["n_cols"] < 1:
        return "skip", "too_small"
    if f["nonempty_ratio"] < 0.08:
        return "skip", "too_sparse"
    if f["merged_ratio"] >= 0.08 or f["header_like_rows"] >= 2 or f["empty_cols_inside"] >= 1 or (f["title_like_top_row"] and f["n_rows"] >= 5):
        return "41", "complex_layout"
    return "40", "regular_table"


def build_element_metadata(grid, elements):
    ordered = sorted(elements, key=lambda x: (x["bbox"]["r1"], x["bbox"]["c1"]))
    for idx, elem in enumerate(ordered, 1):
        bbox = elem["bbox"]
        prev_elem = ordered[idx - 2] if idx > 1 else None
        next_elem = ordered[idx] if idx < len(ordered) else None

        gap_rows_from_prev = None
        gap_cols_from_prev = None
        if prev_elem is not None:
            prev_bbox = prev_elem["bbox"]
            gap_rows_from_prev = count_empty_rows_between(
                grid,
                prev_bbox["r2"],
                bbox["r1"],
                min(prev_bbox["c1"], bbox["c1"]),
                max(prev_bbox["c2"], bbox["c2"]),
            )
            gap_cols_from_prev = count_empty_cols_between(
                grid,
                prev_bbox["c2"],
                bbox["c1"],
                min(prev_bbox["r1"], bbox["r1"]),
                max(prev_bbox["r2"], bbox["r2"]),
            )

        elem["reading_order_index"] = idx
        elem["prev_element_id"] = prev_elem.get("element_id") if prev_elem else None
        elem["next_element_id"] = next_elem.get("element_id") if next_elem else None
        elem["gap_rows_from_prev"] = gap_rows_from_prev
        elem["gap_cols_from_prev"] = gap_cols_from_prev
        elem["element_hint"] = {
            "has_numeric_body": elem["features"].get("has_numeric_body", False),
            "title_to_body_gap_rows": elem["features"].get("title_to_body_gap_rows"),
            "left_body_margin_cols": elem["features"].get("left_body_margin_cols"),
        }
    return ordered


def group_elements_by_numeric_body(elements):
    grouped = []
    current = []
    for elem in elements:
        if not current:
            current = [elem]
            continue

        prev = current[-1]
        same_table_candidate = False
        if prev["features"].get("has_numeric_body") and elem["features"].get("has_numeric_body"):
            prev_gap = elem.get("gap_rows_from_prev")
            prev_col_gap = elem.get("gap_cols_from_prev")
            if prev_gap is not None and prev_gap <= 2:
                same_table_candidate = True
            if prev_col_gap is not None and prev_col_gap == 0 and abs(prev["bbox"]["r1"] - elem["bbox"]["r1"]) <= 2:
                same_table_candidate = True

        if same_table_candidate:
            current.append(elem)
        else:
            grouped.append(current)
            current = [elem]

    if current:
        grouped.append(current)

    return grouped


# --- Inserted helper functions ---


def union_bbox(bboxes):
    rows1 = [b["r1"] for b in bboxes]
    rows2 = [b["r2"] for b in bboxes]
    cols1 = [b["c1"] for b in bboxes]
    cols2 = [b["c2"] for b in bboxes]
    return {
        "r1": min(rows1),
        "r2": max(rows2),
        "c1": min(cols1),
        "c2": max(cols2),
    }

# --- Inserted helper functions ---

def union_bbox_from_elements(elements):
    return union_bbox([x["bbox"] for x in elements])


def recompute_group_bbox(grid, group, include_intermediate_elements: bool = True):
    base_bbox = union_bbox_from_elements(group)
    r1, r2, c1, c2 = base_bbox["r1"], base_bbox["r2"], base_bbox["c1"], base_bbox["c2"]

    if include_intermediate_elements:
        group_rows = [x["bbox"]["r1"] for x in group] + [x["bbox"]["r2"] for x in group]
        group_cols = [x["bbox"]["c1"] for x in group] + [x["bbox"]["c2"] for x in group]
        r1, r2 = min(group_rows), max(group_rows)
        c1, c2 = min(group_cols), max(group_cols)

    merged_bbox = trim_empty_edges(grid, {"r1": r1, "r2": r2, "c1": c1, "c2": c2})
    return merged_bbox


def build_element_api_payload(elements):
    payload = []
    for elem in elements:
        print(elem.get("element_id"))
        payload.append({
            "element_id": elem.get("element_id"),
            "reading_order_index": elem.get("reading_order_index"),
            "prev_element_id": elem.get("prev_element_id"),
            "next_element_id": elem.get("next_element_id"),
            "gap_rows_from_prev": elem.get("gap_rows_from_prev"),
            "gap_cols_from_prev": elem.get("gap_cols_from_prev"),
            "title_guess": elem.get("title_guess"),
            "bbox": elem.get("bbox"),
            "features": {
                "n_rows": elem["features"].get("n_rows"),
                "n_cols": elem["features"].get("n_cols"),
                "nonempty_ratio": elem["features"].get("nonempty_ratio"),
                "numeric_ratio": elem["features"].get("numeric_ratio"),
                "header_like_rows": elem["features"].get("header_like_rows"),
                "title_like_top_row": elem["features"].get("title_like_top_row"),
                "has_numeric_body": elem["features"].get("has_numeric_body"),
                "numeric_body_bbox": elem["features"].get("numeric_body_bbox"),
                "title_to_body_gap_rows": elem["features"].get("title_to_body_gap_rows"),
                "left_body_margin_cols": elem["features"].get("left_body_margin_cols"),
            },
        })
    return payload


def merge_element_group(ws, excel_path: Path, sheet_name: str, group_idx: int, group):
    grid = build_ws_grid(ws)
    merged_bbox = recompute_group_bbox(grid, group, include_intermediate_elements=True)
    merged_df = extract_block_df_from_ws(ws, merged_bbox)
    merged_bbox = trim_empty_edges(grid, merged_bbox)
    merged_df = extract_block_df_from_ws(ws, merged_bbox)
    title_guess = next((x.get("title_guess") for x in group if x.get("title_guess")), sheet_name)
    source_element_ids = []
    for x in group:
        if x.get("element_id"):
            source_element_ids.append(x.get("element_id"))
        source_element_ids.extend(x.get("contained_element_ids", []))
    source_reading_order = [x.get("reading_order_index") for x in group if x.get("reading_order_index") is not None]
    route_hints = [x.get("route_hint") for x in group if x.get("route_hint")]
    route_reasons = [x.get("route_reason") for x in group if x.get("route_reason")]
    merged_features = dict(group[0].get("features", {})) if group else {}
    merged_features["group_element_count"] = len(group)
    merged_features["source_element_ids"] = source_element_ids
    merged_features["source_reading_order"] = source_reading_order
    merged_features["group_bbox"] = merged_bbox
    merged_features["recomputed_from_group"] = True

    if any(h == "41" for h in route_hints):
        route_hint = "41"
        route_reason = "api_or_heuristic_group_contains_complex_element"
    else:
        route_hint = route_hints[0] if route_hints else "40"
        route_reason = route_reasons[0] if route_reasons else "regular_table"

    return {
        "block_id": f"{Path(excel_path).stem}__{sheet_name}__block_{group_idx:03d}",
        "group_id": f"{Path(excel_path).stem}__{sheet_name}__group_{group_idx:03d}",
        "element_id": f"{Path(excel_path).stem}__{sheet_name}__merged_{group_idx:03d}",
        "title_guess": title_guess,
        "bbox": merged_bbox,
        "df": merged_df,
        "csv_text": merged_df.to_csv(index=False),
        "features": merged_features,
        "route_hint": route_hint,
        "route_reason": route_reason,
        "reading_order_index": min(source_reading_order) if source_reading_order else 1,
        "prev_element_id": group[0].get("prev_element_id") if group else None,
        "next_element_id": group[-1].get("next_element_id") if group else None,
        "gap_rows_from_prev": group[0].get("gap_rows_from_prev") if group else None,
        "gap_cols_from_prev": group[0].get("gap_cols_from_prev") if group else None,
        "element_hint": group[0].get("element_hint") if group else {
            "has_numeric_body": False,
            "title_to_body_gap_rows": None,
            "left_body_margin_cols": None
        },
        "source_element_ids": source_element_ids,
        "api_element_payload": build_element_api_payload(group),
    }

def apply_resolved_grouping(ws, excel_path: Path, sheet_name: str, raw_elements, resolved_groups=None):
    element_map = {x.get("element_id"): x for x in raw_elements}

    if resolved_groups:
        grouped_elements = []
        used = set()

        for grp in resolved_groups:
            elem_ids = [x for x in grp if x in element_map]
            if not elem_ids:
                continue

            group = [element_map[eid] for eid in elem_ids]
            grouped_elements.append(group)
            used.update(elem_ids)

        # grouping 결과에 포함되지 않은 element는 단독 그룹으로 유지
        for elem in raw_elements:
            eid = elem.get("element_id")
            if eid not in used:
                grouped_elements.append([elem])
    else:
        # 핵심 수정:
        # resolved_groups가 없을 때는 휴리스틱 merge 금지
        grouped_elements = [[elem] for elem in raw_elements]

    final_blocks = []
    for group_idx, group in enumerate(grouped_elements, 1):
        final_blocks.append(
            merge_element_group(ws, excel_path, sheet_name, group_idx, group)
        )
    return final_blocks


def split_sheet_into_blocks(excel_path: Path, sheet_name: str, resolved_groups=None):
    df, ws = load_sheet_df_and_ws(excel_path, sheet_name)
    if ws is None:
        return [{
            "block_id": f"{Path(excel_path).stem}__{sheet_name}__block_001",
            "element_id": f"{Path(excel_path).stem}__{sheet_name}__element_001",
            "group_id": f"{Path(excel_path).stem}__{sheet_name}__group_001",
            "title_guess": sheet_name,
            "bbox": {"r1": 1, "r2": len(df), "c1": 1, "c2": max(1, df.shape[1])},
            "df": df.copy(),
            "csv_text": df.to_csv(index=False),
            "features": {
                "n_rows": int(df.shape[0]), "n_cols": int(df.shape[1]), "nonempty_ratio": 1.0 if not df.empty else 0.0,
                "numeric_ratio": 0.0, "text_ratio": 1.0 if not df.empty else 0.0, "border_ratio": 0.0,
                "fill_ratio": 0.0, "bold_ratio": 0.0, "merged_ratio": 0.0, "empty_rows_inside": 0,
                "empty_cols_inside": 0, "header_like_rows": 1, "title_like_top_row": False,
                "has_numeric_body": False, "numeric_body_bbox": None, "title_to_body_gap_rows": None,
                "left_body_margin_cols": None,
            },
            "route_hint": "41", "route_reason": "csv_fallback",
            "reading_order_index": 1,
            "prev_element_id": None,
            "next_element_id": None,
            "gap_rows_from_prev": None,
            "gap_cols_from_prev": None,
            "element_hint": {"has_numeric_body": False, "title_to_body_gap_rows": None, "left_body_margin_cols": None},
            "source_element_ids": [f"{Path(excel_path).stem}__{sheet_name}__element_001"],
            "api_element_payload": [],
        }]
    grid = build_ws_grid(ws)
    mask = dilate_mask(build_occupancy_mask(grid), grid["max_row"], grid["max_col"], row_radius=0, col_radius=0)
    comps = connected_components(mask, grid["max_row"], grid["max_col"])
    raw_elements = []
    for comp in comps:
        element_df, bbox = extract_component_df_from_ws(ws, grid, comp)
        if bbox is None:
            continue
        bbox = trim_empty_edges(grid, bbox)
        if bbox["r1"] > bbox["r2"] or bbox["c1"] > bbox["c2"]:
            continue
        features = compute_block_features(grid, bbox)
        if features["nonempty_ratio"] < 0.03:
            continue
        if features["n_rows"] < 1 or features["n_cols"] < 1:
            continue
        if not (
            features.get("has_numeric_body")
            or features.get("title_like_top_row")
            or features.get("header_like_rows", 0) >= 1
            or features.get("text_ratio", 0) >= 0.08
        ):
            continue
        route_hint, route_reason = classify_block_route({"bbox": bbox, "features": features})
        raw_elements.append({
            "bbox": bbox,
            "df": element_df,
            "csv_text": element_df.to_csv(index=False),
            "features": features,
            "route_hint": route_hint,
            "route_reason": route_reason,
            "title_guess": "", # infer_block_title(ws, bbox),
        })
    raw_elements = sorted(raw_elements, key=lambda x: (x["bbox"]["r1"], x["bbox"]["c1"]))
    for i, elem in enumerate(raw_elements, 1):
        elem["element_id"] = f"{Path(excel_path).stem}__{sheet_name}__element_{i:03d}"

    raw_elements = merge_contained_raw_elements(raw_elements)
    raw_elements = build_element_metadata(grid, raw_elements)

    # 핵심 수정:
    # resolved_groups가 아직 없으면 raw element를 그대로 반환
    if resolved_groups is None:
        final_blocks = []
        for i, elem in enumerate(raw_elements, 1):
            final_blocks.append({
                "block_id": f"{Path(excel_path).stem}__{sheet_name}__block_{i:03d}",
                "group_id": f"{Path(excel_path).stem}__{sheet_name}__group_{i:03d}",
                "element_id": elem.get("element_id"),
                "title_guess": elem.get("title_guess"),
                "bbox": elem.get("bbox"),
                "df": elem.get("df"),
                "csv_text": elem.get("csv_text"),
                "features": elem.get("features"),
                "route_hint": elem.get("route_hint"),
                "route_reason": elem.get("route_reason"),
                "reading_order_index": elem.get("reading_order_index"),
                "prev_element_id": elem.get("prev_element_id"),
                "next_element_id": elem.get("next_element_id"),
                "gap_rows_from_prev": elem.get("gap_rows_from_prev"),
                "gap_cols_from_prev": elem.get("gap_cols_from_prev"),
                "element_hint": elem.get("element_hint"),
                "source_element_ids": ([elem.get("element_id")] if elem.get("element_id") else []) + elem.get("contained_element_ids", []),
                "api_element_payload": build_element_api_payload([elem]),
            })
    else:
        final_blocks = apply_resolved_grouping(
            ws,
            excel_path,
            sheet_name,
            raw_elements,
            resolved_groups=resolved_groups,
        )

    if not final_blocks:
        fallback_df = extract_block_df_from_ws(ws, {"r1": 1, "r2": ws.max_row, "c1": 1, "c2": ws.max_column})
        final_blocks = [{
            "block_id": f"{Path(excel_path).stem}__{sheet_name}__block_001",
            "element_id": f"{Path(excel_path).stem}__{sheet_name}__element_001",
            "group_id": f"{Path(excel_path).stem}__{sheet_name}__group_001",
            "title_guess": sheet_name,
            "bbox": {"r1": 1, "r2": ws.max_row, "c1": 1, "c2": ws.max_column},
            "df": fallback_df,
            "csv_text": fallback_df.to_csv(index=False),
            "features": {
                "n_rows": ws.max_row, "n_cols": ws.max_column, "nonempty_ratio": 1.0 if ws.max_row and ws.max_column else 0.0,
                "numeric_ratio": 0.0, "text_ratio": 1.0, "border_ratio": 0.0, "fill_ratio": 0.0, "bold_ratio": 0.0,
                "merged_ratio": 0.0, "empty_rows_inside": 0, "empty_cols_inside": 0, "header_like_rows": 1,
                "title_like_top_row": False, "has_numeric_body": False, "numeric_body_bbox": None,
                "title_to_body_gap_rows": None, "left_body_margin_cols": None,
            },
            "route_hint": "41", "route_reason": "fallback_full_sheet",
            "reading_order_index": 1,
            "prev_element_id": None,
            "next_element_id": None,
            "gap_rows_from_prev": None,
            "gap_cols_from_prev": None,
            "element_hint": {"has_numeric_body": False, "title_to_body_gap_rows": None, "left_body_margin_cols": None},
            "source_element_ids": [f"{Path(excel_path).stem}__{sheet_name}__element_001"],
            "api_element_payload": [],
        }]
    return final_blocks
