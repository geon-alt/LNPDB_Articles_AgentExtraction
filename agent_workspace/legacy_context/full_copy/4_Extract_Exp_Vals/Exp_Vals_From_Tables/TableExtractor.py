import streamlit as st
import cv2
import numpy as np
import pandas as pd
from PIL import Image
import json
import hashlib
from io import BytesIO
from streamlit_drawable_canvas import st_canvas
import sys
from pathlib import Path
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

# ============================================================
# 실행 방법:
#   pip install streamlit opencv-python pillow pandas openpyxl
#   pip install streamlit-drawable-canvas
#   python -m streamlit run table_extractor.py
# ============================================================

# =========================
# Gemini API 설정
# =========================
CURRENT_DIR = Path(__file__).resolve().parent

PROJECT_ROOT = None
for p in [CURRENT_DIR, *CURRENT_DIR.parents]:
    if (p / "find_api.py").exists():
        PROJECT_ROOT = p
        if str(p) not in sys.path:
            sys.path.insert(0, str(p))
        break

API_IMPORT_ERROR = None
try:
    from find_api import find_api_key_file, get_vertexai_client
    API_AVAILABLE = True
except Exception as e:
    API_AVAILABLE = False
    API_IMPORT_ERROR = str(e)

# =========================
# Streamlit 기본 설정
# =========================
st.set_page_config(layout="wide", page_title="Image → Excel Table Extractor")

# =========================
# 세션 상태 초기화
# =========================
def init_session_state():
    defaults = {
        "uploaded_file_hash": None,
        "uploaded_file_name": None,
        "image_cv": None,
        "image_pil": None,
        "crop_region": None,         # 드래그로 선택한 표 영역 (x, y, w, h)
        "df_extracted": None,        # 추출된 표 데이터
        "extraction_done": False,
        "pan_x": 0,
        "canvas_display_width": 700,
        "zoom_pct": 100,
        "extraction_log": "",
    }
    for k, v in defaults.items():
        if k not in st.session_state:
            st.session_state[k] = v


init_session_state()

# =========================
# 유틸 함수
# =========================
def file_to_hash(file_bytes: bytes) -> str:
    return hashlib.md5(file_bytes).hexdigest()

def load_image_from_bytes(file_bytes: bytes):
    np_arr = np.frombuffer(file_bytes, dtype=np.uint8)
    image_cv = cv2.imdecode(np_arr, cv2.IMREAD_COLOR)
    if image_cv is None:
        raise ValueError("이미지를 디코딩하지 못했습니다.")
    image_pil = Image.open(BytesIO(file_bytes)).convert("RGB")
    return image_cv, image_pil

def crop_image(image_cv, region):
    """region: (x, y, w, h) → cropped numpy array"""
    x, y, w, h = region
    H, W = image_cv.shape[:2]
    x1 = max(0, x)
    y1 = max(0, y)
    x2 = min(W, x + w)
    y2 = min(H, y + h)
    return image_cv[y1:y2, x1:x2]

def pil_to_bytes(pil_img: Image.Image, fmt="PNG") -> bytes:
    buf = BytesIO()
    pil_img.save(buf, format=fmt)
    return buf.getvalue()

# =========================
# Gemini OCR 분석
# =========================
def run_gemini_table_ocr(pil_img: Image.Image):
    """
    Gemini에게 표 이미지를 보내고 JSON 2D 배열로 반환받습니다.
    반환 형식: {"headers": [...], "rows": [[...], [...]]}
    """
    if not API_AVAILABLE:
        return None, f"Gemini API 모듈 import 실패: {API_IMPORT_ERROR}"
    try:
        key_path = find_api_key_file("phy1-491103-7c82edc452c9.json")
        client = get_vertexai_client(key_path)

        prompt = """
You are an expert OCR and table extraction assistant.
Analyze the table in this image and extract ALL cell contents precisely.

Return ONLY valid JSON — no markdown fences, no explanation.
Format:
{
  "headers": ["Col1", "Col2", "Col3"],
  "rows": [
    ["val1", "val2", "val3"],
    ["val4", "val5", "val6"]
  ]
}

Rules:
- If the table has NO header row, use "headers": []
- Preserve numbers exactly as shown (do not convert units)
- Preserve merged cell text — repeat the value in each logical cell
- If a cell is empty, use ""
- Preserve special characters (±, ×, %, etc.)
- Output must be parseable by json.loads()
"""
        res = client.models.generate_content(
            model="gemini-2.5-flash",
            contents=[prompt, pil_img]
        )
        text = res.text.strip().replace("```json", "").replace("```", "").strip()
        data = json.loads(text)
        return data, None
    except json.JSONDecodeError as e:
        return None, f"JSON 파싱 오류: {e}\n원본 응답: {text[:300]}"
    except Exception as e:
        return None, f"Gemini 분석 실패: {e}"

# =========================
# Excel 파일 생성
# =========================
def df_to_excel_bytes(df: pd.DataFrame, sheet_name="Sheet1", has_header=True) -> bytes:
    """DataFrame → openpyxl 워크북 → bytes"""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = sheet_name

    # 스타일 정의
    header_fill = PatternFill("solid", fgColor="1F4E79")
    header_font = Font(name="Arial", bold=True, color="FFFFFF", size=11)
    body_font = Font(name="Arial", size=10)
    center_align = Alignment(horizontal="center", vertical="center", wrap_text=True)
    left_align = Alignment(horizontal="left", vertical="center", wrap_text=True)
    thin = Side(style="thin", color="AAAAAA")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    alt_fill = PatternFill("solid", fgColor="EEF3FA")

    if has_header:
        # 헤더 행 쓰기
        for col_idx, col_name in enumerate(df.columns, 1):
            cell = ws.cell(row=1, column=col_idx, value=str(col_name))
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = center_align
            cell.border = border
        start_row = 2
    else:
        start_row = 1

    # 데이터 행 쓰기
    for row_idx, row_data in enumerate(df.itertuples(index=False), start=start_row):
        fill = alt_fill if (row_idx % 2 == 0) else PatternFill()
        for col_idx, value in enumerate(row_data, 1):
            cell = ws.cell(row=row_idx, column=col_idx, value=value)
            cell.font = body_font
            cell.fill = fill
            cell.alignment = left_align
            cell.border = border

    # 열 너비 자동 조정
    for col in ws.columns:
        max_len = 0
        col_letter = col[0].column_letter
        for cell in col:
            try:
                if cell.value:
                    max_len = max(max_len, len(str(cell.value)))
            except Exception:
                pass
        ws.column_dimensions[col_letter].width = min(max(max_len + 2, 10), 40)

    # 행 높이
    for row in ws.iter_rows():
        ws.row_dimensions[row[0].row].height = 18

    ws.freeze_panes = "A2" if has_header else None

    buf = BytesIO()
    wb.save(buf)
    return buf.getvalue()

# =========================
# 표시용 이미지 빌드
# =========================
def build_display_image():
    """원본 이미지에 선택 영역 박스를 그려 반환"""
    if st.session_state.image_cv is None:
        return None
    img = st.session_state.image_cv.copy()
    if st.session_state.crop_region is not None:
        x, y, w, h = st.session_state.crop_region
        cv2.rectangle(img, (x, y), (x + w, y + h), (0, 165, 255), 3)
    return cv2.cvtColor(img, cv2.COLOR_BGR2RGB)

# =========================
# 업로드 처리
# =========================
def handle_upload(uploaded_file):
    if uploaded_file is None:
        return
    file_bytes = uploaded_file.getvalue()
    file_hash = file_to_hash(file_bytes)
    if st.session_state.uploaded_file_hash != file_hash or st.session_state.image_cv is None:
        try:
            image_cv, image_pil = load_image_from_bytes(file_bytes)
            st.session_state.image_cv = image_cv
            st.session_state.image_pil = image_pil
            st.session_state.uploaded_file_hash = file_hash
            st.session_state.uploaded_file_name = uploaded_file.name
            st.session_state.crop_region = None
            st.session_state.df_extracted = None
            st.session_state.extraction_done = False
            st.session_state.extraction_log = ""
            st.rerun()
        except Exception as e:
            st.error(f"이미지 로딩 오류: {e}")

# ============================================================
# ─── UI ──────────────────────────────────────────────────
# ============================================================
st.title("📋 Image → Excel  Table Extractor")
st.caption("표가 포함된 이미지를 업로드하고, 드래그로 영역을 지정하면 자동으로 엑셀 파일로 추출해드립니다.")

# ── 사이드바 ────────────────────────────────────────────────
with st.sidebar:
    st.header("⚙️ 설정")

    uploaded_file = st.file_uploader("1. 이미지 업로드", type=["png", "jpg", "jpeg", "bmp", "tiff", "webp"])
    handle_upload(uploaded_file)

    st.divider()

    st.subheader("2. 표시 설정")
    st.session_state.canvas_display_width = st.slider(
        "화면 너비 (px)", 400, 1200, st.session_state.canvas_display_width, 50
    )
    st.session_state.zoom_pct = st.slider(
        "원본 줌 (%)", 50, 300, st.session_state.zoom_pct, 10
    )

    st.divider()

    st.subheader("3. 추출 옵션")
    has_header = st.checkbox("첫 행을 헤더로 처리", value=True)
    sheet_name = st.text_input("시트 이름", value="Sheet1")
    figure_name = st.text_input("파일명 (확장자 제외)", value="extracted_table", placeholder="예: Figure3_Table")

    st.divider()

    if st.session_state.crop_region is not None:
        x, y, w, h = st.session_state.crop_region
        st.success(f"✅ 선택 영역: x={x}, y={y}, w={w}, h={h}")
        if st.button("🔄 영역 초기화", use_container_width=True):
            st.session_state.crop_region = None
            st.session_state.df_extracted = None
            st.session_state.extraction_done = False
            st.rerun()
    else:
        st.info("오른쪽 이미지에서 드래그하여 표 영역을 지정하세요.\n전체 이미지가 표라면 아래 버튼을 클릭하세요.")
        if st.session_state.image_cv is not None:
            if st.button("🖼️ 이미지 전체를 표 영역으로 사용", use_container_width=True):
                h_img, w_img = st.session_state.image_cv.shape[:2]
                st.session_state.crop_region = (0, 0, w_img, h_img)
                st.rerun()

    st.divider()

    # 추출 실행 버튼
    extract_btn = st.button(
        "🤖 Gemini AI로 표 추출 실행",
        type="primary",
        use_container_width=True,
        disabled=(st.session_state.crop_region is None or st.session_state.image_cv is None),
    )

# ── 메인 영역 ────────────────────────────────────────────────
if st.session_state.image_cv is None:
    st.info("👈 왼쪽 사이드바에서 이미지를 업로드해 주세요.")
    st.stop()

# ── Gemini 추출 실행 ─────────────────────────────────────────
if extract_btn:
    with st.spinner("🔬 Gemini AI가 표를 분석하고 있습니다..."):
        cropped_cv = crop_image(st.session_state.image_cv, st.session_state.crop_region)
        cropped_pil = Image.fromarray(cv2.cvtColor(cropped_cv, cv2.COLOR_BGR2RGB))

        data, err = run_gemini_table_ocr(cropped_pil)

        if err:
            st.session_state.extraction_log = f"❌ {err}"
            st.session_state.df_extracted = None
            st.session_state.extraction_done = False
        else:
            headers = data.get("headers", [])
            rows = data.get("rows", [])

            if not rows:
                st.session_state.extraction_log = "⚠️ 표 데이터를 찾지 못했습니다. 영역을 다시 지정해보세요."
                st.session_state.df_extracted = None
                st.session_state.extraction_done = False
            else:
                # 열 수 통일 (들쭉날쭉한 행 패딩)
                max_cols = max(len(r) for r in rows)
                if headers:
                    max_cols = max(max_cols, len(headers))
                    headers += [""] * (max_cols - len(headers))
                    col_names = headers
                else:
                    col_names = [f"Col{i+1}" for i in range(max_cols)]

                padded_rows = [r + [""] * (max_cols - len(r)) for r in rows]
                df = pd.DataFrame(padded_rows, columns=col_names)

                st.session_state.df_extracted = df
                st.session_state.extraction_done = True
                st.session_state.extraction_log = f"✅ {len(df)}행 × {len(df.columns)}열 추출 완료"

    st.rerun()

# ── 상단: 이미지 + 미리보기 레이아웃 ────────────────────────
tab_image, tab_preview = st.tabs(["🖼️ 이미지 영역 지정", "📊 추출 결과 & 다운로드"])

# ============================================================
# 탭 1: 이미지 드래그
# ============================================================
with tab_image:
    col_info, col_canvas = st.columns([1, 2], gap="medium")

    with col_info:
        st.subheader("사용 방법")
        st.markdown("""
1. **드래그**로 표 영역을 선택하세요.
2. 선택 후 사이드바의 **"Gemini AI로 표 추출 실행"** 버튼을 누르세요.
3. 결과 탭에서 확인·수정 후 **엑셀 다운로드**를 받으세요.

> 전체 이미지가 표라면 사이드바의  
> **"이미지 전체를 표 영역으로 사용"** 버튼을 클릭하세요.
""")

        if st.session_state.crop_region:
            x, y, w, h = st.session_state.crop_region
            st.success(f"선택 영역\n- 위치: ({x}, {y})\n- 크기: {w} × {h} px")

            # 크롭된 미리보기
            cropped = crop_image(st.session_state.image_cv, st.session_state.crop_region)
            cropped_pil = Image.fromarray(cv2.cvtColor(cropped, cv2.COLOR_BGR2RGB))
            st.image(cropped_pil, caption="선택 영역 미리보기", use_column_width=True)

        if st.session_state.extraction_log:
            if "✅" in st.session_state.extraction_log:
                st.success(st.session_state.extraction_log)
            elif "⚠️" in st.session_state.extraction_log:
                st.warning(st.session_state.extraction_log)
            else:
                st.error(st.session_state.extraction_log)

    with col_canvas:
        st.subheader("이미지 (드래그하여 영역 선택)")

        img_rgb = build_display_image()
        orig_h, orig_w = img_rgb.shape[:2]

        zoom_factor = st.session_state.zoom_pct / 100.0
        view_w = int(orig_w / zoom_factor)
        view_w = min(view_w, orig_w)

        max_pan = max(0, orig_w - view_w)
        pan_x = 0
        if max_pan > 0:
            pan_x = st.slider("↔️ 가로 스크롤", 0, max_pan, 0, key="img_pan_slider")

        cropped_view = img_rgb[:, pan_x: pan_x + view_w]
        pil_bg = Image.fromarray(cropped_view)

        canvas_w = st.session_state.canvas_display_width
        canvas_h = int(pil_bg.height * (canvas_w / pil_bg.width))

        canvas_result = st_canvas(
            fill_color="rgba(255, 165, 0, 0.15)",
            stroke_width=2,
            stroke_color="#FF6600",
            background_image=pil_bg,
            update_streamlit=False,
            height=canvas_h,
            width=canvas_w,
            drawing_mode="rect",
            key="table_area_canvas",
        )

        if canvas_result.json_data is not None:
            objs = canvas_result.json_data.get("objects", [])
            if objs:
                last = objs[-1]
                if last["type"] == "rect":
                    scale = view_w / canvas_w
                    rx = int(last["left"] * scale) + pan_x
                    ry = int(last["top"] * scale)
                    rw = int(last["width"] * scale)
                    rh = int(last["height"] * scale)

                    if rw > 10 and rh > 10:
                        new_region = (rx, ry, rw, rh)
                        if st.session_state.crop_region != new_region:
                            st.session_state.crop_region = new_region
                            st.session_state.df_extracted = None
                            st.session_state.extraction_done = False
                            st.rerun()

# ============================================================
# 탭 2: 추출 결과 & 다운로드
# ============================================================
with tab_preview:
    if not st.session_state.extraction_done or st.session_state.df_extracted is None:
        st.info("아직 추출된 데이터가 없습니다. 이미지 영역을 선택한 뒤 사이드바의 추출 버튼을 눌러주세요.")
    else:
        df = st.session_state.df_extracted

        st.subheader(f"추출 결과 — {len(df)}행 × {len(df.columns)}열")
        st.caption("셀을 직접 클릭하여 수정할 수 있습니다.")

        # ── 편집 가능한 데이터 에디터 ────────────────────────
        edited_df = st.data_editor(
            df,
            use_container_width=True,
            num_rows="dynamic",
            hide_index=True,
            key="result_editor",
        )
        st.session_state.df_extracted = edited_df.copy()

        st.divider()

        # ── 열 추가/삭제 도구 ────────────────────────────────
        with st.expander("🔧 열 관리 (추가 / 삭제)"):
            mc1, mc2 = st.columns(2)
            with mc1:
                new_col_name = st.text_input("추가할 열 이름", key="new_col_input")
                if st.button("열 추가", use_container_width=True):
                    if new_col_name and new_col_name not in st.session_state.df_extracted.columns:
                        st.session_state.df_extracted[new_col_name] = ""
                        st.rerun()
            with mc2:
                del_col = st.selectbox("삭제할 열", options=st.session_state.df_extracted.columns.tolist(), key="del_col_sel")
                if st.button("열 삭제", use_container_width=True):
                    st.session_state.df_extracted = st.session_state.df_extracted.drop(columns=[del_col])
                    st.rerun()

        st.divider()

        # ── 다운로드 버튼 ───────────────────────────────────
        col_dl1, col_dl2 = st.columns(2)

        with col_dl1:
            excel_bytes = df_to_excel_bytes(
                st.session_state.df_extracted,
                sheet_name=sheet_name,
                has_header=has_header,
            )
            st.download_button(
                label="📥 엑셀(.xlsx)로 다운로드",
                data=excel_bytes,
                file_name=f"{figure_name}.xlsx",
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                use_container_width=True,
                type="primary",
            )

        with col_dl2:
            csv_bytes = st.session_state.df_extracted.to_csv(index=False).encode("utf-8-sig")
            st.download_button(
                label="📄 CSV로 다운로드",
                data=csv_bytes,
                file_name=f"{figure_name}.csv",
                mime="text/csv",
                use_container_width=True,
            )

        # ── 원본 크롭 이미지도 나란히 비교 ──────────────────
        if st.session_state.crop_region is not None:
            st.divider()
            st.subheader("원본 표 이미지 비교")
            cropped_compare = crop_image(st.session_state.image_cv, st.session_state.crop_region)
            st.image(
                cv2.cvtColor(cropped_compare, cv2.COLOR_BGR2RGB),
                caption="선택된 표 영역 (원본)",
                use_column_width=True,
            )

# ── 하단 안내 ────────────────────────────────────────────────
with st.expander("📖 패키지 설치 방법"):
    st.code("""
pip install streamlit opencv-python pillow pandas openpyxl
pip install streamlit-drawable-canvas
python -m streamlit run table_extractor.py
    """, language="bash")