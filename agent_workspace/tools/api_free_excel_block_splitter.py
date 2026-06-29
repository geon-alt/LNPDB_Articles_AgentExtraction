"""API-free Excel block splitter for external-agent stage 03.

This helper intentionally avoids Gemini/Vertex/LLM modules and the legacy
03_split_excel_blocks_batch.py script. It wraps the deterministic workbook
layout utilities in 0_mark_down_gen/sheet_block_splitter.py and writes the
stage artifacts expected by Agent_Task_Runner validation.
"""

from __future__ import annotations

import argparse
import contextlib
import csv
import importlib.util
import io
import json
import re
import shutil
from datetime import datetime, timezone
from pathlib import Path
from typing import Any

from openpyxl import load_workbook


PROJECT_ROOT = Path(__file__).resolve().parents[2]
SPLITTER_PATH = PROJECT_ROOT / "0_mark_down_gen" / "sheet_block_splitter.py"

INVENTORY_FIELDS = [
    "excel_file",
    "excel_sheet",
    "block_id",
    "group_id",
    "element_id",
    "block_csv_path",
    "block_meta_path",
    "block_type",
]

ALLOWED_BLOCK_TYPES = {
    "title_and_table",
    "table_body",
    "table_title",
    "multi_table",
    "note",
    "other",
}


def load_splitter_module() -> Any:
    spec = importlib.util.spec_from_file_location("api_free_sheet_block_splitter", SPLITTER_PATH)
    if spec is None or spec.loader is None:
        raise RuntimeError(f"Could not load splitter module from {SPLITTER_PATH}")
    module = importlib.util.module_from_spec(spec)
    spec.loader.exec_module(module)
    return module


def safe_name(value: str) -> str:
    cleaned = re.sub(r"[^A-Za-z0-9_.-]+", "_", value.strip())
    cleaned = cleaned.strip("._")
    return cleaned or "unnamed"


def rel_to_paper(path: Path, paper_folder: Path) -> str:
    try:
        return path.resolve().relative_to(paper_folder.resolve()).as_posix()
    except ValueError:
        return str(path)


def utc_now() -> str:
    return datetime.now(timezone.utc).isoformat(timespec="seconds")


def backup_existing_outputs(paper_folder: Path) -> list[str]:
    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    backup_root = paper_folder / "_agent_backups" / f"03_split_excel_blocks_batch_{timestamp}"
    backed_up: list[str] = []
    for name in ("excel_block_inventory.csv", "three_core_result_all.json", "Exp_Excel_Blocks"):
        src = paper_folder / name
        if not src.exists():
            continue
        backup_root.mkdir(parents=True, exist_ok=True)
        dst = backup_root / name
        if src.is_dir():
            shutil.copytree(src, dst)
        else:
            shutil.copy2(src, dst)
        backed_up.append(rel_to_paper(src, paper_folder))
    return backed_up


def json_sanitize(value: Any) -> Any:
    if isinstance(value, dict):
        return {str(k): json_sanitize(v) for k, v in value.items()}
    if isinstance(value, (list, tuple)):
        return [json_sanitize(v) for v in value]
    try:
        import pandas as pd

        if value is pd.NA:
            return None
    except Exception:
        pass
    if hasattr(value, "item"):
        try:
            return value.item()
        except Exception:
            pass
    return value


def nonempty_texts(df: Any, limit: int = 20) -> list[str]:
    texts: list[str] = []
    for row in df.fillna("").astype(str).values.tolist():
        for cell in row:
            text = str(cell).strip()
            if text and text.lower() not in {"nan", "none", "null"}:
                texts.append(text)
                if len(texts) >= limit:
                    return texts
    return texts


def numeric_like(text: str) -> bool:
    s = str(text).strip().replace(",", "")
    if s.endswith("%"):
        s = s[:-1]
    if not s:
        return False
    try:
        float(s)
        return True
    except ValueError:
        return False


def row_has_mixed_header(row: list[str]) -> bool:
    values = [str(v).strip() for v in row if str(v).strip()]
    if not values:
        return False
    text_count = sum(1 for value in values if not numeric_like(value))
    numeric_count = len(values) - text_count
    return text_count > 0 and numeric_count > 0


def classify_block(block: dict[str, Any]) -> tuple[str, str, str, bool]:
    df = block.get("df")
    features = block.get("features") or {}
    if df is None:
        return "other", "low", "No dataframe was returned for this block.", True

    matrix = df.fillna("").astype(str).values.tolist()
    texts = nonempty_texts(df, 80)
    numeric_count = sum(1 for text in texts if numeric_like(text))
    text_count = len(texts) - numeric_count
    row_count = int(features.get("n_rows") or len(matrix))
    col_count = int(features.get("n_cols") or (max((len(r) for r in matrix), default=0)))
    numeric_ratio = float(features.get("numeric_ratio") or 0.0)
    text_ratio = float(features.get("text_ratio") or 0.0)
    header_like_rows = int(features.get("header_like_rows") or 0)
    title_like_top_row = bool(features.get("title_like_top_row"))
    empty_rows_inside = int(features.get("empty_rows_inside") or 0)
    empty_cols_inside = int(features.get("empty_cols_inside") or 0)
    merged_ratio = float(features.get("merged_ratio") or 0.0)
    first_rows = matrix[: min(3, len(matrix))]
    mixed_header = any(row_has_mixed_header(row) for row in first_rows)
    figure_labels = sum(1 for text in texts if re.search(r"\bfig(?:ure)?\.?\s*\d+", text, re.I))
    group_labels = sum(
        1
        for text in texts
        if not numeric_like(text)
        and any(token in text.lower() for token in ("pbs", "control", "treated", "mc3", "g0-", "fluc", "il-12"))
    )

    if len(texts) <= 3 and numeric_count == 0:
        return "table_title", "high", "Block contains only a short text label/title and no numeric cells.", False

    if numeric_count == 0 and text_count > 0:
        if row_count <= 4:
            return "table_title", "medium", "Text-only block is short and functions as a table or panel label.", False
        return "note", "medium", "Text-only block spans multiple rows and is better treated as a note.", False

    if (
        (figure_labels >= 2 and numeric_count >= 6)
        or (empty_cols_inside >= 2 and group_labels >= 2 and numeric_count >= 8)
        or (row_count >= 18 and col_count >= 10 and empty_rows_inside >= 1 and numeric_count >= 10)
    ):
        return "multi_table", "medium", "Block contains multiple panel/group labels or internal separators with numeric data.", False

    if numeric_count >= 3 and (
        title_like_top_row
        or header_like_rows >= 1
        or mixed_header
        or group_labels >= 1
        or merged_ratio > 0
    ):
        return "title_and_table", "high", "Header/group text is present with a numeric table body.", False

    if numeric_count >= max(3, text_count) and numeric_ratio >= max(0.25, text_ratio):
        return "table_body", "high", "Numeric cells dominate and no clear separate title was detected.", False

    if row_count <= 4 and text_count >= numeric_count:
        return "note", "medium", "Small text-heavy block with limited numeric content.", False

    return "other", "low", "Mixed or sparse content could not be assigned to a narrower block type.", True


def write_csv_matrix(path: Path, df: Any) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    df.fillna("").to_csv(path, index=False, header=False, encoding="utf-8-sig")


def write_csv_dicts(path: Path, rows: list[dict[str, Any]], fieldnames: list[str]) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    with path.open("w", encoding="utf-8-sig", newline="") as handle:
        writer = csv.DictWriter(handle, fieldnames=fieldnames, extrasaction="ignore")
        writer.writeheader()
        for row in rows:
            writer.writerow({field: row.get(field, "") for field in fieldnames})


def write_json(path: Path, data: Any) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    path.write_text(json.dumps(json_sanitize(data), ensure_ascii=False, indent=2), encoding="utf-8")


def split_sheet(splitter: Any, excel_file: Path, sheet_name: str) -> tuple[list[dict[str, Any]], str]:
    capture = io.StringIO()
    with contextlib.redirect_stdout(capture):
        blocks = splitter.split_sheet_into_blocks(excel_file, sheet_name)
    return blocks, capture.getvalue().strip()


def process_workbooks(paper_folder: Path) -> dict[str, Any]:
    marker = paper_folder / ".manual_select_review_done"
    if not marker.exists():
        raise FileNotFoundError(f"Missing required manual review marker: {marker}")
    exp_excel = paper_folder / "Exp_Excel"
    if not exp_excel.is_dir():
        raise FileNotFoundError(f"Missing Exp_Excel folder: {exp_excel}")

    backups = backup_existing_outputs(paper_folder)
    block_root = paper_folder / "Exp_Excel_Blocks"
    block_root.mkdir(parents=True, exist_ok=True)

    splitter = load_splitter_module()
    inventory_rows: list[dict[str, Any]] = []
    summary_rows: list[dict[str, Any]] = []
    excel_files = sorted(p for p in exp_excel.glob("*.xlsx") if p.is_file())

    for excel_file in excel_files:
        wb = load_workbook(excel_file, read_only=True, data_only=True)
        for sheet_name in wb.sheetnames:
            sheet_summary: dict[str, Any] = {
                "excel_file": rel_to_paper(excel_file, paper_folder),
                "excel_sheet": sheet_name,
                "result": {
                    "status": "ok",
                    "method": "openpyxl_layout_splitter_api_free",
                    "reasoning": [],
                    "blocks": [],
                },
            }
            try:
                blocks, captured_stdout = split_sheet(splitter, excel_file, sheet_name)
            except Exception as exc:
                sheet_summary["result"] = {
                    "status": "failed",
                    "method": "openpyxl_layout_splitter_api_free",
                    "manual_required": True,
                    "failure_cause": str(exc),
                    "changed_outputs": "No block files were written for this failed sheet.",
                }
                summary_rows.append(sheet_summary)
                continue

            sheet_summary["result"]["reasoning"].append(
                "Split candidate blocks using deterministic occupied-cell components, merged-cell expansion, "
                "blank row/column separation, border/fill/bold/header features, and numeric-density features."
            )
            if captured_stdout:
                sheet_summary["result"]["helper_stdout_suppressed"] = captured_stdout.splitlines()[:20]

            for local_index, block in enumerate(blocks, start=1):
                df = block.get("df")
                if df is None or df.empty:
                    continue
                block_type, confidence, reason, manual_required = classify_block(block)
                if block_type not in ALLOWED_BLOCK_TYPES:
                    block_type = "other"
                    confidence = "low"
                    manual_required = True
                    reason = "Classifier returned a non-allowed type; normalized to other."

                canonical_block_id = f"block_{len(inventory_rows) + 1:04d}"
                rel_dir = Path("Exp_Excel_Blocks") / safe_name(excel_file.stem) / safe_name(sheet_name)
                block_csv = paper_folder / rel_dir / f"{canonical_block_id}.csv"
                block_meta = paper_folder / rel_dir / f"{canonical_block_id}.json"
                write_csv_matrix(block_csv, df)

                texts = nonempty_texts(df, 12)
                features = block.get("features") or {}
                meta = {
                    "excel_file": rel_to_paper(excel_file, paper_folder),
                    "excel_sheet": sheet_name,
                    "block_id": canonical_block_id,
                    "source_block_id": block.get("block_id", ""),
                    "group_id": block.get("group_id", ""),
                    "element_id": block.get("element_id", ""),
                    "reading_order_index": block.get("reading_order_index", local_index),
                    "bbox": block.get("bbox", {}),
                    "block_type": block_type,
                    "classification_confidence": confidence,
                    "manual_required": manual_required,
                    "classification_reason": reason,
                    "evidence": {
                        "nonempty_sample": texts,
                        "numeric_count_in_sample": sum(1 for text in texts if numeric_like(text)),
                        "features": features,
                        "route_hint": block.get("route_hint", ""),
                        "route_reason": block.get("route_reason", ""),
                        "source_element_ids": block.get("source_element_ids", []),
                    },
                    "created_by": "api_free_excel_block_splitter.py",
                    "created_at": utc_now(),
                }
                write_json(block_meta, meta)

                row = {
                    "excel_file": rel_to_paper(excel_file, paper_folder),
                    "excel_sheet": sheet_name,
                    "block_id": canonical_block_id,
                    "group_id": block.get("group_id", ""),
                    "element_id": block.get("element_id", ""),
                    "block_csv_path": rel_to_paper(block_csv, paper_folder),
                    "block_meta_path": rel_to_paper(block_meta, paper_folder),
                    "block_type": block_type,
                }
                inventory_rows.append(row)
                sheet_summary["result"]["blocks"].append(
                    {
                        **row,
                        "source_block_id": block.get("block_id", ""),
                        "bbox": block.get("bbox", {}),
                        "classification_confidence": confidence,
                        "manual_required": manual_required,
                        "classification_reason": reason,
                        "evidence_text": " | ".join(texts[:8]),
                    }
                )

            sheet_summary["result"]["block_count"] = len(sheet_summary["result"]["blocks"])
            if not sheet_summary["result"]["blocks"]:
                sheet_summary["result"]["manual_required"] = True
                sheet_summary["result"]["reasoning"].append("No non-empty blocks were produced for this sheet.")
            summary_rows.append(sheet_summary)

    write_csv_dicts(paper_folder / "excel_block_inventory.csv", inventory_rows, INVENTORY_FIELDS)
    write_json(paper_folder / "three_core_result_all.json", summary_rows)
    return {
        "excel_files": len(excel_files),
        "sheets": len(summary_rows),
        "blocks": len(inventory_rows),
        "backed_up": backups,
        "outputs": [
            rel_to_paper(paper_folder / "excel_block_inventory.csv", paper_folder),
            rel_to_paper(paper_folder / "three_core_result_all.json", paper_folder),
            rel_to_paper(block_root, paper_folder),
        ],
    }


def main() -> None:
    parser = argparse.ArgumentParser(description="Split Excel sheets into API-free table blocks.")
    parser.add_argument("--paper-folder", required=True, help="Target paper folder")
    args = parser.parse_args()
    result = process_workbooks(Path(args.paper_folder))
    print(json.dumps(result, ensure_ascii=False, indent=2))


if __name__ == "__main__":
    main()
